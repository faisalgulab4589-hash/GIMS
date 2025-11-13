import sqlite3
from datetime import datetime
from openpyxl import load_workbook
from PyQt5.QtWidgets import QFileDialog, QMessageBox

# Assuming db.py and config.py are in the same directory or accessible
from db import get_connection, init_db
from config import DB_NAME

def import_students_from_excel(parent=None):
    file_path, _ = QFileDialog.getOpenFileName(
        parent, "Select Excel File", "", "Excel Files (*.xlsx *.xls)"
    )
    if not file_path:
        QMessageBox.information(parent, "Import Cancelled", "No Excel file selected.")
        return

    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
    except Exception as e:
        QMessageBox.critical(parent, "Error Reading Excel", f"Could not open or read Excel file: {e}")
        return

    # Assuming the first row contains headers
    headers = [cell.value for cell in sheet[1]]
    
    # Define mapping from Excel headers to database column names
    # Supports multiple variations of header names for flexibility
    column_mapping = {
        # Admission Number
        'Admission No': 'admission_no',
        'admission_no': 'admission_no',
        'Admission Number': 'admission_no',
        # Name
        'Name': 'name',
        'Student Name': 'name',
        # Father's Name
        'Father\'s Name': 'father_name',
        'Father Name': 'father_name',
        'Fathers Name': 'father_name',
        # Address
        'Address': 'address',
        # Date of Birth
        'DOB': 'dob',
        'Date of Birth': 'dob',
        # Gender
        'Gender': 'gender',
        # Nationality
        'Nationality': 'nationality',
        # District
        'District': 'district',
        # Phone
        'Phone': 'phone',
        'Phone #': 'phone',
        'Phone Number': 'phone',
        # SMS Phone
        'SMS Phone': 'sms_phone',
        'SMS Phone #': 'sms_phone',
        'SMS Number': 'sms_phone',
        'SMS': 'sms_phone',
        # Campus
        'Campus': 'campus',
        # Board
        'Board': 'board',
        # Technology/Program
        'Technology': 'technology',
        'Technology/Program': 'technology',
        'Program': 'technology',
        'Course': 'technology',
        # Semester
        'Semester': 'semester',
        'Semester/Year': 'semester',
        'Session': 'semester',
        'Year': 'semester',
        # Status
        'Status': 'status',
        'Student Status': 'status',
        # Student Type
        'Student Type': 'student_type',
        'Type': 'student_type',
        # Remarks
        'Remarks': 'remarks',
        'Remarks & Notes': 'remarks',
        'Notes': 'remarks',
        'Comments': 'remarks',
        # 'Photo Path' will not be imported from Excel, handled by Biodata form
        # 'Created At' will be set automatically
    }

    # Reverse mapping for easier lookup
    db_to_excel_map = {v: k for k, v in column_mapping.items()}

    # Note: We no longer require all headers to be present. The import will gracefully handle missing columns.
    # Only Admission No is truly required to identify the student record.

    imported_count = 0
    updated_count = 0
    skipped_count = 0
    error_messages = []

    # Assuming 'parent' is an instance of BioDataForm
    if parent is None or not hasattr(parent, 'save_student'):
        QMessageBox.critical(None, "Import Error", "Invalid parent object provided for Excel import. Must be a BioDataForm instance.")
        return

    for row_idx in range(2, sheet.max_row + 1): # Start from second row (after headers)
        row_data = {}
        for col_idx, header in enumerate(headers):
            if header is None:
                continue
            # Try exact match first
            if header in column_mapping:
                db_col_name = column_mapping[header]
                cell_value = sheet.cell(row=row_idx, column=col_idx + 1).value
                # Only set if not already set (to avoid overwriting with None)
                if db_col_name not in row_data or row_data[db_col_name] is None:
                    row_data[db_col_name] = cell_value
            else:
                # Try case-insensitive match
                header_lower = str(header).lower().strip()
                for excel_header, db_col_name in column_mapping.items():
                    if str(excel_header).lower().strip() == header_lower:
                        cell_value = sheet.cell(row=row_idx, column=col_idx + 1).value
                        # Only set if not already set (to avoid overwriting with None)
                        if db_col_name not in row_data or row_data[db_col_name] is None:
                            row_data[db_col_name] = cell_value
                        break

        admission_no = str(row_data.get('admission_no', '')).strip()
        if not admission_no:
            skipped_count += 1
            error_messages.append(f"Row {row_idx}: Skipped due to missing Admission No.")
            continue

        # Populate BioDataForm fields
        parent.admission_no_field.setText(admission_no)
        parent.name.setText(str(row_data.get('name') or ''))
        parent.fname.setText(str(row_data.get('father_name') or ''))
        parent.address.setPlainText(str(row_data.get('address') or ''))

        dob_excel = row_data.get('dob')
        if isinstance(dob_excel, datetime):
            parent.dob.setDate(dob_excel.date())
        elif isinstance(dob_excel, str) and dob_excel:
            try:
                parent.dob.setDate(datetime.strptime(dob_excel, '%Y-%m-%d').date())
            except ValueError:
                try:
                    parent.dob.setDate(datetime.strptime(dob_excel, '%d-%m-%Y').date())
                except ValueError:
                    error_messages.append(f"Row {row_idx} (Admission No: {admission_no}): Invalid DOB format '{dob_excel}'. Expected YYYY-MM-DD or DD-MM-YYYY. Using current date.")
                    parent.dob.setDate(datetime.now().date())
        else:
            parent.dob.setDate(datetime.now().date()) # Default to current date if invalid or missing

        # Set combo box values, handling cases where value might not exist in dropdown
        def set_combo_text(combo_box, text):
            if not text:
                combo_box.setCurrentIndex(0)
                return
            text_str = str(text).strip()
            if not text_str:
                combo_box.setCurrentIndex(0)
                return
            index = combo_box.findText(text_str)
            if index != -1:
                combo_box.setCurrentIndex(index)
            else:
                # Add the new item if it's not empty
                combo_box.addItem(text_str)
                combo_box.setCurrentText(text_str)

        set_combo_text(parent.gender, row_data.get('gender'))
        set_combo_text(parent.nationality, row_data.get('nationality'))
        set_combo_text(parent.district, row_data.get('district'))
        set_combo_text(parent.campus, row_data.get('campus'))
        set_combo_text(parent.board, row_data.get('board'))
        set_combo_text(parent.technology, row_data.get('technology'))
        set_combo_text(parent.semester, row_data.get('semester'))
        set_combo_text(parent.status_combo, row_data.get('status', 'Active'))
        set_combo_text(parent.student_type, row_data.get('student_type', 'Paid'))

        # Set phone fields - handle None values properly
        phone_val = row_data.get('phone')
        parent.phone.setText(str(phone_val) if phone_val else '')

        sms_val = row_data.get('sms_phone')
        parent.sms_phone.setText(str(sms_val) if sms_val else '')

        # Photo path is not imported from Excel, it's handled by the form's upload mechanism
        parent.file_path = None # Ensure no old photo path is carried over
        parent.show_preview() # Clear any existing preview

        # Attempt to save the student using the BioDataForm's method
        try:
            # The save_student method will handle validation and DB interaction
            # It also shows QMessageBox for success/failure, so we'll suppress those for batch import
            # and collect our own counts/errors.
            # To avoid multiple QMessageBox popups, we'll temporarily override it.
            original_info = QMessageBox.information
            original_warning = QMessageBox.warning
            original_critical = QMessageBox.critical
            original_question = QMessageBox.question # Store original question function

            # Mock QMessageBox to capture results without showing popups
            class MockMessageBox:
                def __init__(self):
                    self.last_message = None
                    self.last_title = None
                def information(self_mock, parent_mock, title, message):
                    self_mock.last_message = message
                    self_mock.last_title = title
                def warning(self_mock, parent_mock, title, message):
                    self_mock.last_message = message
                    self_mock.last_title = title
                def critical(self_mock, parent_mock, title, message):
                    self_mock.last_message = message
                    self_mock.last_title = title
                def question(self_mock, parent_mock, title, message, buttons):
                    # For updates, always say Yes to proceed with update
                    return QMessageBox.Yes

            mock_msg_box = MockMessageBox()
            QMessageBox.information = mock_msg_box.information
            QMessageBox.warning = mock_msg_box.warning
            QMessageBox.critical = mock_msg_box.critical
            QMessageBox.question = mock_msg_box.question # Override question for updates

            # Call save_student, but catch phone validation errors and retry with empty phones
            try:
                parent.save_student()
            except Exception as save_error:
                # If phone validation fails, try again with empty phone fields
                error_msg = str(save_error)
                if 'phone' in error_msg.lower() or 'sms' in error_msg.lower():
                    # Clear phone fields and retry
                    parent.phone.setText('')
                    parent.sms_phone.setText('')
                    parent.save_student()
                else:
                    raise

            # Restore original QMessageBox functions
            QMessageBox.information = original_info
            QMessageBox.warning = original_warning
            QMessageBox.critical = original_critical
            QMessageBox.question = original_question # Restore original question

            if "updated successfully" in mock_msg_box.last_message:
                updated_count += 1
            elif "saved successfully" in mock_msg_box.last_message:
                imported_count += 1
            else:
                skipped_count += 1
                error_messages.append(f"Row {row_idx} (Admission No: {admission_no}): Save failed - {mock_msg_box.last_message}")

        except Exception as e:
            skipped_count += 1
            error_messages.append(f"Row {row_idx} (Admission No: {admission_no}): Error calling save_student - {e}")
        finally:
            # Ensure QMessageBox is restored even if an error occurs
            QMessageBox.information = original_info
            QMessageBox.warning = original_warning
            QMessageBox.critical = original_critical
            QMessageBox.question = original_question


    summary = (
        f"✅ Excel Import Complete!\n\n"
        f"📊 Results:\n"
        f"  • Successfully Imported: {imported_count}\n"
        f"  • Successfully Updated: {updated_count}\n"
        f"  • Skipped: {skipped_count}\n"
        f"  • Total Processed: {imported_count + updated_count + skipped_count}\n"
    )
    if error_messages:
        summary += f"\n⚠️ Errors encountered ({len(error_messages)}):\n" + "\n".join(error_messages[:10])
        if len(error_messages) > 10:
            summary += f"\n... and {len(error_messages) - 10} more errors"
        QMessageBox.warning(parent, "Import Summary with Errors", summary)
    else:
        QMessageBox.information(parent, "✅ Import Successful", summary)

if __name__ == '__main__':
    # This block is for testing the import script independently
    # Ensure the database is initialized before attempting to import
    init_db()
    from PyQt5.QtWidgets import QApplication
    app = QApplication([]) # Initialize QApplication once
    import_students_from_excel(None) # Pass None as parent for standalone execution
    app.exec_() # Start the Qt event loop
