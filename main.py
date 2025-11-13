import os
import re
import sqlite3
import calendar
from difflib import SequenceMatcher
from collections import defaultdict
from flask import Flask, send_file, jsonify, request, redirect, url_for, session, render_template, flash, abort
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
import db
import io
from openpyxl import Workbook
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
import pandas as pd # For reading excel file in Flask
from db import get_connection # Ensure get_connection is imported
from config import DB_NAME # Ensure DB_NAME is imported
from werkzeug.security import generate_password_hash, check_password_hash
import functools
import json
try:
    import bcrypt
except ImportError:  # pragma: no cover - optional dependency
    bcrypt = None
    print("Warning: bcrypt module not found. Falling back to Werkzeug PBKDF2 hashing.")
from rbac_constants import DEFAULT_MODULES, ROUTE_PERMISSION_RULES

app = Flask(__name__, template_folder='templates')
import secrets
app.secret_key = secrets.token_hex(16) # Generate a random 16-byte (32-character) hex string
app.config['MAX_CONTENT_LENGTH'] = 10 * 1024 * 1024  # 10MB max file size
SESSION_TIMEOUT_MINUTES = 15
app.permanent_session_lifetime = timedelta(minutes=SESSION_TIMEOUT_MINUTES)

# Initialize the database
with app.app_context():
    db.init_db()

# Decorator to check if user is logged in
def login_required(view):
    @functools.wraps(view)
    def wrapped_view(**kwargs):
        if 'logged_in' not in session or not session['logged_in']:
            # Check if this is an API request (JSON expected)
            if request.path.startswith('/api/'):
                return jsonify({'status': 'error', 'message': 'Unauthorized. Please login first.'}), 401
            return redirect(url_for('login'))
        return view(**kwargs)
    return wrapped_view

# Decorator to check if user is admin
def admin_required(view):
    @functools.wraps(view)
    def wrapped_view(**kwargs):
        if not user_is_admin():
            # Check if this is an API request (JSON expected)
            if request.path.startswith('/api/'):
                return jsonify({'status': 'error', 'message': 'Unauthorized. Admin privileges required.'}), 403
            flash('Unauthorized access. Admin privileges required.', 'danger')
            return redirect(url_for('index'))
        return view(**kwargs)
    return wrapped_view


@app.before_request
def enforce_session_timeout():
    """Logout inactive sessions after the configured idle window."""
    if not session.get('logged_in'):
        return
    now = datetime.utcnow()
    last_activity = session.get('last_activity')
    if last_activity:
        try:
            last_active = datetime.fromisoformat(last_activity)
        except ValueError:
            last_active = now
        if now - last_active > timedelta(minutes=SESSION_TIMEOUT_MINUTES):
            logout_current_user(reason='Session expired', silent=True)
            if request.path.startswith('/api/'):
                return jsonify({'status': 'error', 'message': 'Session expired. Please login again.'}), 401
            flash('Session expired due to inactivity. Please login again.', 'warning')
            return redirect(url_for('login'))
    session['last_activity'] = now.isoformat()


@app.before_request
def enforce_module_permissions():
    """Apply module guards declared in ROUTE_PERMISSION_RULES."""
    if not session.get('logged_in') or session.get('auth_system') != 'rbac':
        return
    path = request.path or ''
    if path.startswith('/static'):
        return
    for rule in ROUTE_PERMISSION_RULES:
        if any(path.startswith(prefix) for prefix in rule.get('prefixes', [])):
            if not user_has_module(rule.get('module')):
                return forbidden_response(rule.get('module'))
            break


@app.context_processor
def inject_user_metadata():
    """Expose frequently used session values to templates."""
    return {
        'current_user_full_name': session.get('full_name') or session.get('teacher_name'),
        'current_role_label': session.get('role_label') or (session.get('role') or '').title(),
        'allowed_modules': session.get('module_permissions') or [],
        'module_payload': session.get('module_payload') or [],
        'last_login_at': session.get('last_login_at')
    }


AUDIT_EXCLUDED_ENDPOINTS = {
    'admin_create_user',
    'admin_update_user',
    'admin_update_user_status',
    'admin_reset_user_password',
    'admin_delete_user',
    'admin_create_role',
    'admin_update_role',
    'admin_delete_role'
}

EXPORT_ENDPOINT_LOGS = {
    'export_report1_excel': ('reports', 'Exported organization summary (Excel)'),
    'export_report1_pdf': ('reports', 'Exported organization summary (PDF)'),
    'export_report2_excel': ('reports', 'Exported semester analytics (Excel)'),
    'export_report2_pdf': ('reports', 'Exported semester analytics (PDF)'),
    'export_report3_excel': ('reports', 'Exported technology analytics (Excel)'),
    'export_report3_pdf': ('reports', 'Exported technology analytics (PDF)'),
    'export_deductions_excel': ('payroll', 'Exported deductions register (Excel)'),
    'export_deductions_pdf': ('payroll', 'Exported deductions register (PDF)'),
    'export_daily_attendance_pdf': ('attendance', 'Exported daily attendance report (PDF)'),
    'export_monthly_attendance_pdf': ('attendance', 'Exported monthly attendance report (PDF)'),
    'export_monthly_attendance_excel': ('attendance', 'Exported monthly attendance report (Excel)'),
    'export_cards_to_pdf': ('documents', 'Generated student cards PDF'),
    'export_exam_results_pdf': ('dmc', 'Exported exam results PDF')
}


@app.after_request
def audit_mutating_requests(response):
    """Automatically log successful mutating operations for RBAC users."""
    if (
        session.get('auth_system') == 'rbac'
        and session.get('user_id')
        and request.method in {'POST', 'PUT', 'DELETE'}
        and response.status_code < 400
    ):
        endpoint = request.endpoint or ''
        if endpoint not in AUDIT_EXCLUDED_ENDPOINTS:
            module_key = resolve_module_from_path(request.path)
            if module_key:
                action = 'delete' if request.method == 'DELETE' else ('create' if request.method == 'POST' else 'update')
                log_user_action(action, module_key=module_key, description=f'{request.method} {request.path}')
    if response.status_code < 400:
        export_meta = EXPORT_ENDPOINT_LOGS.get(request.endpoint or '')
        if export_meta and session.get('auth_system') == 'rbac' and session.get('user_id'):
            log_user_action('export', module_key=export_meta[0], description=export_meta[1])
    return response

def can_view_student_dmc(student_id):
    """Admins/teachers or the student (when logged in) can view the DMC."""
    if session.get('role') in ('admin', 'teacher'):
        return True
    if session.get('student_logged_in') and session.get('student_id') == student_id:
        return True
    return False

def parse_multi_value(value):
    """Convert stored JSON/comma string into a clean list."""
    if not value:
        return []
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return []
        try:
            parsed = json.loads(value)
            if isinstance(parsed, list):
                return [str(item).strip() for item in parsed if str(item).strip()]
        except json.JSONDecodeError:
            pass
        if ',' in value:
            return [item.strip() for item in value.split(',') if item.strip()]
        return [value]
    return []

def serialize_multi_value(values):
    """Serialize a multi-select list to JSON."""
    if not values:
        return json.dumps([])
    cleaned = [str(item).strip() for item in values if str(item).strip()]
    return json.dumps(cleaned)

def join_display(values):
    """Return a human-friendly comma-separated string."""
    return ', '.join(values or [])

def normalize_phone(value):
    """Return only numeric phone digits (max 15)."""
    digits = re.sub(r'\D', '', value or '')
    return digits[:15]

def normalize_cnic(value):
    """Return CNIC in standard 13-digit format with dashes when possible."""
    digits = re.sub(r'\D', '', value or '')[:13]
    if len(digits) == 13:
        return f"{digits[:5]}-{digits[5:12]}-{digits[12:]}"
    return digits

def teacher_email_exists(conn, email, exclude_id=None):
    """Check if an email is already associated with a teacher."""
    if not email:
        return False
    query = 'SELECT id FROM teachers WHERE lower(email) = lower(?)'
    params = [email]
    if exclude_id:
        query += ' AND id != ?'
        params.append(exclude_id)
    return conn.execute(query, params).fetchone() is not None

def ensure_inventory_permission(required_roles=None):
    """Centralized authorization guard for inventory resources."""
    roles = required_roles or INVENTORY_FULL_ACCESS_ROLES
    user_role = session.get('role')
    if user_role in roles:
        return None
    message = 'Unauthorized. Inventory access required.'
    if request.path.startswith('/api/'):
        return jsonify({'status': 'error', 'message': message}), 403
    flash(message, 'danger')
    return redirect(url_for('dashboard'))

def parse_decimal(value):
    """Safely convert user input into a rounded float."""
    try:
        return round(float(value), 2)
    except (TypeError, ValueError):
        return 0.0

def inventory_counts(conn):
    """Return summarized metrics for inventory dashboard."""
    total_items = conn.execute('SELECT COUNT(*) AS c FROM inventory_items').fetchone()['c']
    low_stock = conn.execute(
        'SELECT COUNT(*) AS c FROM inventory_items WHERE quantity_available < ? AND status NOT IN (?, ?)',
        (INVENTORY_LOW_STOCK_THRESHOLD, *INVENTORY_INACTIVE_STATUSES)
    ).fetchone()['c']
    expiring = conn.execute(
        '''
            SELECT COUNT(*) AS c FROM inventory_items
            WHERE expiry_date IS NOT NULL
              AND expiry_date != ''
              AND DATE(expiry_date) <= DATE('now', ?)
              AND DATE(expiry_date) >= DATE('now')
        ''',
        (f'+{INVENTORY_EXPIRY_WINDOW_DAYS} days',)
    ).fetchone()['c']
    expired = conn.execute(
        '''
            SELECT COUNT(*) AS c FROM inventory_items
            WHERE expiry_date IS NOT NULL
              AND expiry_date != ''
              AND DATE(expiry_date) < DATE('now')
        '''
    ).fetchone()['c']
    return {
        'total_items': total_items,
        'low_stock_items': low_stock,
        'expiring_items': expiring,
        'expired_items': expired
    }

def inventory_item_payload(data):
    """Normalize dict payload for insert/update operations."""
    quantity_total = int(data.get('quantity_total') or data.get('quantity') or 0)
    quantity_available = int(data.get('quantity_available') or quantity_total)
    unit_price = parse_decimal(data.get('unit_price'))
    total_price = round(quantity_total * unit_price, 2)
    return {
        'item_name': (data.get('item_name') or '').strip(),
        'category': (data.get('category') or '').strip(),
        'brand': (data.get('brand') or '').strip(),
        'model_no': (data.get('model_no') or '').strip(),
        'quantity_total': quantity_total,
        'quantity_available': quantity_available,
        'purchase_date': (data.get('purchase_date') or '').strip() or None,
        'expiry_date': (data.get('expiry_date') or '').strip() or None,
        'unit_price': unit_price,
        'total_price': total_price,
        'location': (data.get('location') or '').strip(),
        'status': (data.get('status') or 'Active').strip() or 'Active',
        'remarks': (data.get('remarks') or '').strip()
    }

def serialize_inventory_item(row):
    """Convert sqlite row to dict with parsed helper values."""
    if not row:
        return None
    item = dict(row)
    item['quantity_total'] = int(item.get('quantity_total') or 0)
    item['quantity_available'] = int(item.get('quantity_available') or 0)
    item['unit_price'] = parse_decimal(item.get('unit_price'))
    item['total_price'] = parse_decimal(item.get('total_price'))
    item['is_low_stock'] = item['quantity_available'] < INVENTORY_LOW_STOCK_THRESHOLD and item.get('status') not in INVENTORY_INACTIVE_STATUSES
    if item.get('expiry_date'):
        try:
            expiry_dt = datetime.fromisoformat(item['expiry_date'])
            delta = (expiry_dt.date() - datetime.now().date()).days
            item['days_to_expiry'] = delta
            item['is_expired'] = delta < 0
            item['is_expiring_soon'] = 0 <= delta <= INVENTORY_EXPIRY_WINDOW_DAYS
        except ValueError:
            item['days_to_expiry'] = None
            item['is_expired'] = False
            item['is_expiring_soon'] = False
    else:
        item['days_to_expiry'] = None
        item['is_expired'] = False
        item['is_expiring_soon'] = False
    return item

def serialize_issue_row(row):
    if not row:
        return None
    issue = dict(row)
    issue['quantity'] = int(issue.get('quantity') or 0)
    issue['quantity_returned'] = int(issue.get('quantity_returned') or 0)
    return issue

def maybe_create_auto_request(conn, item_dict):
    """Create a low-stock request when quantity dips below threshold."""
    if not item_dict or item_dict.get('quantity_available', 0) >= INVENTORY_LOW_STOCK_THRESHOLD:
        return
    existing = conn.execute(
        '''
            SELECT request_id FROM inventory_requests
            WHERE LOWER(item_name) = LOWER(?) AND status IN ('Pending','Approved')
            ORDER BY request_date DESC LIMIT 1
        ''',
        (item_dict.get('item_name'),)
    ).fetchone()
    if existing:
        return
    now = datetime.now().isoformat()
    conn.execute(
        '''
            INSERT INTO inventory_requests (requested_by, requested_role, item_name, quantity, status, request_date, remarks)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''',
        ('System', 'Automation', item_dict.get('item_name'), INVENTORY_LOW_STOCK_THRESHOLD * 2, 'Pending', now,
         'Auto-generated request due to low stock')
    )

def calculate_grade_from_percentage(percentage, passing_threshold=None):
    """Return grade string based on percentage."""
    if percentage is None:
        return '-'
    if percentage >= 90:
        return 'A+'
    if percentage >= 80:
        return 'A'
    if percentage >= 70:
        return 'B'
    if percentage >= 60:
        return 'C'
    if percentage >= 50:
        return 'D'
    if passing_threshold is not None and percentage >= passing_threshold:
        return 'E'
    return 'F'

def determine_result_status(obtained_marks, total_marks, passing_marks=None):
    if total_marks is None or total_marks == 0:
        return 'Pending'
    percentage = (obtained_marks / total_marks * 100) if total_marks else 0
    threshold = passing_marks if passing_marks is not None else 50
    return 'Pass' if percentage >= threshold else 'Fail'

DEDUCTION_TYPE_OPTIONS = ('Late', 'Absent', 'Leave without Pay', 'Other')

INVENTORY_ISSUE_TYPES = ('Student', 'Teacher', 'Department')
INVENTORY_INACTIVE_STATUSES = ('Damaged', 'Disposed')
INVENTORY_LOW_STOCK_THRESHOLD = 10
INVENTORY_EXPIRY_WINDOW_DAYS = 30
INVENTORY_FULL_ACCESS_ROLES = {'admin', 'store_incharge'}
INVENTORY_REPORT_ROLES = INVENTORY_FULL_ACCESS_ROLES | {'principal'}


def hash_password(password):
    """Hash plaintext password using bcrypt."""
    if not password:
        raise ValueError("Password cannot be empty")
    if bcrypt:
        return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    return generate_password_hash(password)


def verify_password(password, hashed):
    """Safely verify a bcrypt hashed password."""
    if not password or not hashed:
        return False
    if hashed.startswith('pbkdf2:'):
        return check_password_hash(hashed, password)
    if not bcrypt:
        return check_password_hash(hashed, password)
    try:
        return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))
    except (ValueError, TypeError):
        return False


def canonical_role_key(role_name):
    """Return a normalized role key."""
    if not role_name:
        return ''
    return re.sub(r'\W+', '_', role_name.strip().lower())


def fetch_role_modules(conn, role_id):
    """Return modules assigned to a role."""
    rows = conn.execute(
        '''
            SELECT m.module_key, m.label
            FROM role_permissions rp
            JOIN access_modules m ON rp.module_id = m.id
            WHERE rp.role_id = ? AND rp.granted = 1
            ORDER BY m.label
        ''',
        (role_id,)
    ).fetchall()
    return [dict(row) for row in rows]


def build_user_session(user_row, role_modules):
    """Populate the Flask session with RBAC user context."""
    session.clear()
    session['logged_in'] = True
    session['auth_system'] = 'rbac'
    session['user_id'] = user_row['id']
    session['username'] = user_row['username']
    session['full_name'] = user_row['full_name']
    session['last_login_at'] = user_row.get('last_login_at')
    role_label = user_row.get('role_name') or ''
    role_key = canonical_role_key(role_label)
    session['role_label'] = role_label
    session['role_key'] = role_key
    session['role'] = role_key or session.get('role')
    session['is_admin'] = role_key == 'admin'
    module_keys = [module['module_key'] for module in role_modules]
    session['module_permissions'] = module_keys
    session['module_payload'] = role_modules
    session['last_activity'] = datetime.utcnow().isoformat()
    session.permanent = True


def current_user_id():
    """Return the active RBAC user id if available."""
    return session.get('user_id')


def user_is_admin():
    """Return True if the logged-in user has admin privileges."""
    if session.get('auth_system') == 'rbac':
        return bool(session.get('is_admin'))
    return (session.get('role') or '').lower() == 'admin'


def user_has_module(module_key):
    """Check if current RBAC session has access to a module."""
    if not module_key:
        return True
    if user_is_admin():
        return True
    if session.get('auth_system') != 'rbac':
        return True
    allowed = session.get('module_permissions') or []
    return module_key in allowed


def forbidden_response(module_key):
    """Return a formatted forbidden response based on request type."""
    readable = (module_key or 'requested').replace('_', ' ').title()
    message = f'Access to {readable} module is not permitted.'
    if request.path.startswith('/api/'):
        return jsonify({'status': 'error', 'message': message}), 403
    flash(message, 'danger')
    return redirect(url_for('index'))


def module_required(module_key):
    """Decorator to enforce that user has module access."""
    def decorator(view):
        @functools.wraps(view)
        def wrapped_view(*args, **kwargs):
            if not user_has_module(module_key):
                return forbidden_response(module_key)
            return view(*args, **kwargs)
        return wrapped_view
    return decorator


def log_user_action(action, module_key=None, description=None, entity_type=None, entity_id=None, metadata=None):
    """Write a structured activity record for the logged-in RBAC user."""
    user_id = session.get('user_id')
    if not user_id:
        return
    conn = None
    try:
        conn = get_connection()
        conn.execute(
            '''
                INSERT INTO user_activity_log (
                    user_id, username, role_snapshot, action, module_key,
                    entity_type, entity_id, description, metadata, ip_address, created_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''',
            (
                user_id,
                session.get('username'),
                session.get('role_label') or session.get('role'),
                action,
                module_key,
                entity_type,
                str(entity_id) if entity_id is not None else None,
                description,
                json.dumps(metadata) if metadata else None,
                request.remote_addr,
                datetime.utcnow().isoformat()
            )
        )
        conn.commit()
    except Exception as exc:
        print(f"Activity log error: {exc}")
    finally:
        if conn:
            conn.close()


def logout_current_user(reason=None, silent=False):
    """Clear session data and optionally record logout activity."""
    if session.get('auth_system') == 'rbac' and session.get('user_id'):
        log_user_action('logout', description=reason or 'User logged out')
    keys_to_clear = [
        'logged_in', 'auth_system', 'user_id', 'username', 'full_name', 'last_login_at',
        'role', 'role_label', 'role_key', 'is_admin', 'module_permissions',
        'module_payload', 'teacher_id', 'teacher_name', 'permissions', 'employee_id',
        'subject', 'technology', 'assigned_semesters', 'student_logged_in',
        'student_id', 'last_activity', 'teacher_role'
    ]
    for key in keys_to_clear:
        session.pop(key, None)
    session.permanent = False
    if not silent:
        flash('You have been logged out.', 'info')


def list_access_modules(conn):
    """Return all registered modules ordered by label."""
    return [
        dict(row) for row in conn.execute(
            'SELECT id, module_key, label, description FROM access_modules ORDER BY label'
        ).fetchall()
    ]


def modules_by_role(conn):
    """Return a mapping of role_id -> [module dict]."""
    mapping = defaultdict(list)
    rows = conn.execute(
        '''
            SELECT rp.role_id, m.module_key, m.label
            FROM role_permissions rp
            JOIN access_modules m ON rp.module_id = m.id
            WHERE rp.granted = 1
            ORDER BY m.label
        '''
    ).fetchall()
    for row in rows:
        mapping[row['role_id']].append({'module_key': row['module_key'], 'label': row['label']})
    return mapping


def extract_module_keys(raw_modules):
    """Normalize module selection from form/JSON payloads."""
    if raw_modules is None:
        return []
    if isinstance(raw_modules, (list, tuple)):
        return [m for m in raw_modules if m]
    if isinstance(raw_modules, str):
        if raw_modules.strip().startswith('['):
            try:
                parsed = json.loads(raw_modules)
                if isinstance(parsed, list):
                    return [m for m in parsed if m]
            except json.JSONDecodeError:
                pass
        return [m.strip() for m in raw_modules.split(',') if m.strip()]
    return []


def resolve_module_from_path(path):
    """Resolve a module key based on the configured route rules."""
    if not path:
        return None
    for rule in ROUTE_PERMISSION_RULES:
        for prefix in rule.get('prefixes', []):
            if path.startswith(prefix):
                return rule.get('module')
    return None

def normalize_deduction_type(value):
    """Return a valid deduction type label."""
    if not value:
        return 'Other'
    value = str(value).strip()
    if not value:
        return 'Other'
    for option in DEDUCTION_TYPE_OPTIONS:
        if value.lower() == option.lower():
            return option
    lowered = value.lower()
    if 'late' in lowered:
        return 'Late'
    if 'absent' in lowered:
        return 'Absent'
    if 'leave' in lowered:
        return 'Leave without Pay'
    return 'Other'

def security_deduction_amount(employee_row):
    """Return the monthly security deduction configured for an employee."""
    if not employee_row:
        return 0.0
    mode = (employee_row.get('security_mode') or '').lower()
    amount = employee_row.get('security_amount') or 0
    if mode == 'monthly' and amount:
        return float(amount)
    return 0.0

def sum_manual_deductions(cur, employee_id, month, year):
    """Return the total manual deduction amount recorded for an employee/month."""
    result = cur.execute(
        '''
            SELECT COALESCE(SUM(amount), 0) AS total
            FROM employee_deductions
            WHERE employee_id = ? AND month = ? AND year = ?
        ''',
        (employee_id, month, year)
    ).fetchone()
    return float(result['total']) if result else 0.0

def upsert_employee_payroll(cur, employee_row, month, year):
    """Create or update the payroll record for an employee with latest deductions."""
    if not employee_row:
        return

    employee_id = employee_row['id']
    month = int(month)
    year = int(year)

    existing = cur.execute(
        'SELECT id, allowances FROM payroll WHERE employee_id = ? AND month = ? AND year = ?',
        (employee_id, month, year)
    ).fetchone()

    allowances = float(existing['allowances']) if existing and existing['allowances'] is not None else 0.0
    basic_salary = float(employee_row.get('basic_salary') or 0.0)
    manual_deductions = sum_manual_deductions(cur, employee_id, month, year)
    security_amount = security_deduction_amount(employee_row)
    total_deductions = manual_deductions + security_amount
    net_salary = max(basic_salary + allowances - total_deductions, 0.0)
    now = datetime.now().isoformat()

    if existing:
        cur.execute(
            '''
                UPDATE payroll
                SET basic_salary = ?,
                    allowances = ?,
                    deductions = ?,
                    net_salary = ?,
                    generated_date = ?
                WHERE id = ?
            ''',
            (basic_salary, allowances, total_deductions, net_salary, now, existing['id'])
        )
    else:
        cur.execute(
            '''
                INSERT INTO payroll (employee_id, month, year, basic_salary, allowances, deductions, net_salary, generated_date)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''',
            (employee_id, month, year, basic_salary, allowances, total_deductions, net_salary, now)
        )

def fetch_deductions_data(employee_id=None, employee_name=None, father_name=None,
                          month=None, year=None, campus=None):
    """Shared helper to retrieve employee deduction records."""
    conn = get_connection()
    cur = conn.cursor()
    query = '''
        SELECT ed.id, ed.employee_id, e.name as employee_name, e.father_name,
               e.campus, e.basic_salary AS employee_basic_salary,
               ed.month, ed.year, ed.amount,
               ed.days_deducted, ed.reason, ed.deduction_type, ed.created_at,
               d.name as department_name, des.name as designation_name,
               p.basic_salary AS payroll_basic_salary,
               p.allowances AS payroll_allowances,
               p.deductions AS payroll_total_deductions,
               p.net_salary AS payroll_net_salary
        FROM employee_deductions ed
        JOIN employees e ON ed.employee_id = e.id
        LEFT JOIN departments d ON e.department_id = d.id
        LEFT JOIN designations des ON e.designation_id = des.id
        LEFT JOIN payroll p ON p.employee_id = ed.employee_id
            AND p.month = ed.month AND p.year = ed.year
        WHERE 1=1
    '''
    params = []

    if employee_id:
        query += ' AND ed.employee_id = ?'
        params.append(int(employee_id))
    if employee_name:
        query += ' AND LOWER(e.name) LIKE ?'
        params.append(f'%{employee_name.lower()}%')
    if father_name:
        query += ' AND LOWER(e.father_name) LIKE ?'
        params.append(f'%{father_name.lower()}%')
    if month:
        query += ' AND ed.month = ?'
        params.append(int(month))
    if year:
        query += ' AND ed.year = ?'
        params.append(int(year))
    if campus:
        query += ' AND e.campus = ?'
        params.append(campus)

    query += ' ORDER BY e.name, ed.year DESC, ed.month DESC'
    cur.execute(query, params)
    rows = [dict(row) for row in cur.fetchall()]
    conn.close()
    
    formatted_rows = []
    for row in rows:
        base_salary = row.get('payroll_basic_salary')
        if base_salary is None:
            base_salary = row.get('employee_basic_salary') or 0.0
        allowances = row.get('payroll_allowances') or 0.0
        salary_before = float(base_salary or 0.0) + float(allowances or 0.0)
        payroll_total_deductions = row.get('payroll_total_deductions')
        amount = float(row.get('amount') or 0.0)
        if payroll_total_deductions is not None:
            salary_after = salary_before - float(payroll_total_deductions or 0.0)
        else:
            salary_after = salary_before - amount

        row['salary_before'] = round(salary_before, 2)
        row['salary_after'] = round(max(salary_after, 0.0), 2)
        row['deduction_type'] = row.get('deduction_type') or 'Other'
        row['entry_date'] = row.get('created_at')
        formatted_rows.append(row)
    return formatted_rows

def fetch_exam_results_rows(cursor, exam_id, campus=None, technology=None):
    """Reusable helper to grab exam results with optional student filters."""
    query = '''
        SELECT mr.*, s.name, s.admission_no, s.father_name, s.campus, s.semester, s.technology
        FROM midterm_results mr
        JOIN students s ON mr.student_id = s.id
        WHERE mr.exam_id = ?
    '''
    params = [exam_id]
    if campus:
        query += ' AND s.campus = ?'
        params.append(campus)
    if technology:
        query += ' AND s.technology = ?'
        params.append(technology)

    query += ' ORDER BY mr.obtained_marks DESC'
    cursor.execute(query, params)
    return [dict(row) for row in cursor.fetchall()]

@app.route("/login", methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        username = (request.form['username'] or '').strip()
        password = request.form['password']

        conn = get_connection()
        try:
            user = conn.execute(
                '''
                    SELECT u.*, r.name AS role_name
                    FROM users u
                    JOIN user_roles r ON u.role_id = r.id
                    WHERE LOWER(u.username) = LOWER(?)
                    LIMIT 1
                ''',
                (username,)
            ).fetchone()

            if user:
                now = datetime.utcnow()
                now_iso = now.isoformat()
                status = (user['status'] or 'Active').lower()
                if status != 'active':
                    error = 'Your account is inactive. Please contact the administrator.'
                else:
                    suspended_until = user['suspended_until']
                    if suspended_until:
                        try:
                            suspend_dt = datetime.fromisoformat(suspended_until)
                            if suspend_dt > now:
                                formatted = suspend_dt.strftime('%d %b %Y %H:%M')
                                error = f'Account suspended until {formatted}.'
                        except ValueError:
                            pass
                if not error:
                    if verify_password(password, user['password_hash']):
                        conn.execute(
                            'UPDATE users SET login_attempts = 0, last_login_at = ?, last_login_ip = ? WHERE id = ?',
                            (now_iso, request.remote_addr, user['id'])
                        )
                        conn.commit()
                        modules = fetch_role_modules(conn, user['role_id'])
                        build_user_session(dict(user), modules)
                        log_user_action('login', description='Successful login')
                        flash('Logged in successfully!', 'success')
                        return redirect(url_for('index'))
                    else:
                        conn.execute(
                            'UPDATE users SET login_attempts = login_attempts + 1, last_failed_login_at = ? WHERE id = ?',
                            (now_iso, user['id'])
                        )
                        conn.commit()
                        error = 'Invalid username or password. Please try again.'
            else:
                teacher = conn.execute('SELECT * FROM teachers WHERE username = ? AND status = ?', (username, 'Active')).fetchone()
                if teacher and check_password_hash(teacher['password_hash'], password):
                    permissions = conn.execute(
                        'SELECT permission_name FROM teacher_permissions WHERE teacher_id = ? AND granted = 1',
                        (teacher['id'],)
                    ).fetchall()
                    permission_list = [p['permission_name'] for p in permissions]
                    
                    teacher_dict = dict(teacher)
                    assigned_semesters = parse_multi_value(teacher_dict.get('assigned_semesters'))
                    technology_assignments = parse_multi_value(teacher_dict.get('technology_assignments') or teacher_dict.get('technology'))
                    teacher_dict['assigned_semesters'] = assigned_semesters
                    teacher_dict['technology_assignments'] = technology_assignments
                    teacher_dict['technology'] = join_display(technology_assignments)
                    session.clear()
                    session['logged_in'] = True
                    session['auth_system'] = 'legacy_teacher'
                    session['teacher_id'] = teacher_dict['id']
                    session['username'] = teacher_dict['username']
                    session['teacher_name'] = teacher_dict['name']
                    session['role'] = teacher_dict['role']
                    session['permissions'] = permission_list
                    session['employee_id'] = teacher_dict.get('employee_id')
                    session['subject'] = teacher_dict.get('subject')
                    session['technology'] = technology_assignments
                    session['assigned_semesters'] = assigned_semesters
                    session['last_activity'] = datetime.utcnow().isoformat()
                    
                    conn.execute(
                        'INSERT INTO teacher_activity_log (teacher_id, activity_type, activity_description, created_at) VALUES (?, ?, ?, ?)',
                        (teacher_dict['id'], 'login', f'Teacher logged in from {request.remote_addr}', datetime.now().isoformat())
                    )
                    conn.commit()
                    
                    flash('Logged in successfully!', 'success')
                    return redirect(url_for('index'))
                else:
                    error = 'Invalid Credentials or Account Deactivated. Please try again.'
        finally:
            conn.close()
    return render_template('login.html', error=error)

@app.route("/teacher/login", methods=['GET', 'POST'])
def teacher_login():
    """Separate login page for teachers"""
    error = None
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        conn = get_connection()
        try:
            teacher = conn.execute('SELECT * FROM teachers WHERE username = ? AND status = ? AND role = ?', 
                                  (username, 'Active', 'teacher')).fetchone()
            
            if teacher and check_password_hash(teacher['password_hash'], password):
                permissions = conn.execute(
                    'SELECT permission_name FROM teacher_permissions WHERE teacher_id = ? AND granted = 1',
                    (teacher['id'],)
                ).fetchall()
                permission_list = [p['permission_name'] for p in permissions]
                
                teacher_dict = dict(teacher)
                assigned_semesters = parse_multi_value(teacher_dict.get('assigned_semesters'))
                technology_assignments = parse_multi_value(teacher_dict.get('technology_assignments') or teacher_dict.get('technology'))
                teacher_dict['assigned_semesters'] = assigned_semesters
                teacher_dict['technology_assignments'] = technology_assignments
                teacher_dict['technology'] = join_display(technology_assignments)
                session.clear()
                session['logged_in'] = True
                session['auth_system'] = 'legacy_teacher'
                session['teacher_id'] = teacher_dict['id']
                session['username'] = teacher_dict['username']
                session['teacher_name'] = teacher_dict['name']
                session['role'] = teacher_dict['role']
                session['permissions'] = permission_list
                session['employee_id'] = teacher_dict.get('employee_id')
                session['subject'] = teacher_dict.get('subject')
                session['technology'] = technology_assignments
                session['assigned_semesters'] = assigned_semesters
                session['last_activity'] = datetime.utcnow().isoformat()
                session.permanent = True
                
                conn.execute(
                    'INSERT INTO teacher_activity_log (teacher_id, activity_type, activity_description, ip_address, created_at) VALUES (?, ?, ?, ?, ?)',
                    (teacher_dict['id'], 'login', 'Teacher logged in', request.remote_addr, datetime.now().isoformat())
                )
                conn.commit()
                
                flash('Logged in successfully!', 'success')
                return redirect(url_for('teacher_dashboard'))
            else:
                error = 'Invalid Credentials or Account Deactivated. Please contact administrator.'
        finally:
            conn.close()
    return render_template('teacher_login.html', error=error)


def dashboard_context():
    """Shared context for dashboard rendering."""
    return {
        'allowed_modules': session.get('module_permissions') or [],
        'module_payload': session.get('module_payload') or [],
        'last_login_at': session.get('last_login_at'),
        'role_label': session.get('role_label') or session.get('role')
    }

@app.route("/")
@login_required
def index():
    return render_template('dashboard.html', **dashboard_context())

@app.route("/dashboard")
@login_required
def dashboard():
    return render_template('dashboard.html', **dashboard_context())

@app.route("/reports_page")
@login_required
@module_required('reports')
def reports_page():
    return redirect(url_for('index', _anchor='reports'))

@app.route("/free_students_report")
@login_required
@module_required('reports')
def free_students_report():
    return render_template('free_students_report.html')

@app.route('/logout')
@login_required
def logout():
    logout_current_user()
    return redirect(url_for('login'))

@app.route('/change_password', methods=['POST'])
@login_required
def change_password():
    try:
        data = request.get_json() if request.is_json else request.form
        current_password = data.get('current_password') or data.get('current-password')
        new_password = data.get('new_password') or data.get('new-password')
        confirm_password = data.get('confirm_password') or data.get('confirm-password')

        if not all([current_password, new_password, confirm_password]):
            if request.is_json:
                return jsonify({'status': 'error', 'message': 'All fields are required'}), 400
            flash('All fields are required.', 'danger')
            return redirect(url_for('index'))

        if session.get('auth_system') == 'rbac' and session.get('user_id'):
            conn = get_connection()
            try:
                user = conn.execute('SELECT password_hash FROM users WHERE id = ?', (session['user_id'],)).fetchone()
                if not user or not verify_password(current_password, user['password_hash']):
                    raise ValueError('Incorrect current password.')
                if new_password != confirm_password:
                    raise ValueError('New passwords do not match.')
                if len(new_password) < 6:
                    raise ValueError('Password must be at least 6 characters long.')

                conn.execute(
                    'UPDATE users SET password_hash = ?, updated_at = ? WHERE id = ?',
                    (hash_password(new_password), datetime.utcnow().isoformat(), session['user_id'])
                )
                conn.commit()
                log_user_action('update', description='Updated own password')
            finally:
                conn.close()
        else:
            teacher_id = session.get('teacher_id')
            conn = get_connection()
            try:
                teacher = conn.execute('SELECT password_hash FROM teachers WHERE id = ?', (teacher_id,)).fetchone()
            finally:
                conn.close()

            if not teacher or not check_password_hash(teacher['password_hash'], current_password):
                if request.is_json:
                    return jsonify({'status': 'error', 'message': 'Incorrect current password'}), 400
                flash('Incorrect current password.', 'danger')
                return redirect(url_for('index'))
            if new_password != confirm_password:
                if request.is_json:
                    return jsonify({'status': 'error', 'message': 'New passwords do not match'}), 400
                flash('New passwords do not match.', 'danger')
                return redirect(url_for('index'))
            if len(new_password) < 6:
                if request.is_json:
                    return jsonify({'status': 'error', 'message': 'Password must be at least 6 characters long'}), 400
                flash('Password must be at least 6 characters long.', 'danger')
                return redirect(url_for('index'))

            conn = get_connection()
            conn.execute('UPDATE teachers SET password_hash = ?, updated_at = ? WHERE id = ?',
                         (generate_password_hash(new_password), datetime.now().isoformat(), teacher_id))
            conn.commit()
            conn.close()

        if request.is_json:
            return jsonify({'status': 'success', 'message': 'Password changed successfully!'})
        flash('Password changed successfully!', 'success')
        return redirect(url_for('index'))
    except ValueError as err:
        if request.is_json:
            return jsonify({'status': 'error', 'message': str(err)}), 400
        flash(str(err), 'danger')
        return redirect(url_for('index'))
    except Exception as e:
        print(f"Error changing password: {e}")
        if request.is_json:
            return jsonify({'status': 'error', 'message': 'An unexpected error occurred'}), 500
        flash('An error occurred. Please try again.', 'danger')
        return redirect(url_for('index'))

@app.route('/reset_password', methods=['POST'])
@login_required
@admin_required
def reset_password():
    """Reset password for any user (Admin only)"""
    try:
        data = request.get_json() if request.is_json else request.form
        username = data.get('username')
        new_password = data.get('new_password') or data.get('new-password')
        confirm_password = data.get('confirm_password') or data.get('confirm-password')

        if not all([username, new_password, confirm_password]):
            if request.is_json:
                return jsonify({'status': 'error', 'message': 'All fields are required'}), 400
            flash('All fields are required.', 'danger')
            return redirect(url_for('index'))

        if new_password != confirm_password:
            if request.is_json:
                return jsonify({'status': 'error', 'message': 'Passwords do not match'}), 400
            flash('Passwords do not match.', 'danger')
            return redirect(url_for('index'))

        if len(new_password) < 6:
            if request.is_json:
                return jsonify({'status': 'error', 'message': 'Password must be at least 6 characters long'}), 400
            flash('Password must be at least 6 characters long.', 'danger')
            return redirect(url_for('index'))

        conn = get_connection()
        teacher = conn.execute('SELECT id FROM teachers WHERE username = ?', (username,)).fetchone()
        
        if not teacher:
            conn.close()
            if request.is_json:
                return jsonify({'status': 'error', 'message': 'User not found'}), 404
            flash('User not found.', 'danger')
            return redirect(url_for('index'))

        new_password_hash = generate_password_hash(new_password)
        conn.execute('UPDATE teachers SET password_hash = ?, updated_at = ? WHERE username = ?',
                     (new_password_hash, datetime.now().isoformat(), username))
        conn.commit()
        conn.close()
        
        if request.is_json:
            return jsonify({'status': 'success', 'message': f'Password reset successfully for user: {username}'})
        flash(f'Password reset successfully for user: {username}', 'success')
        return redirect(url_for('index'))
    except Exception as e:
        print(f"Error resetting password: {e}")
        if request.is_json:
            return jsonify({'status': 'error', 'message': str(e)}), 500
        flash('An error occurred. Please try again.', 'danger')
        return redirect(url_for('index'))


@app.route('/api/me', methods=['GET'])
@login_required
def api_me():
    """Return the current session's metadata."""
    return jsonify({
        'status': 'success',
        'data': {
            'full_name': session.get('full_name') or session.get('teacher_name'),
            'username': session.get('username'),
            'role': session.get('role_label') or session.get('role'),
            'modules': session.get('module_permissions') or [],
            'last_login_at': session.get('last_login_at'),
            'is_admin': user_is_admin()
        }
    })


@app.route('/api/me/modules', methods=['GET'])
@login_required
def api_me_modules():
    """Return module payload for front-end permission toggling."""
    modules = session.get('module_payload') or []
    return jsonify({'status': 'success', 'modules': modules, 'allowed': session.get('module_permissions') or []})

# ==================== USER & ROLE MANAGEMENT (ADMIN PANEL) ====================

def admin_success(message, endpoint, extra=None):
    if request.is_json:
        payload = {'status': 'success', 'message': message}
        if extra:
            payload.update(extra)
        return jsonify(payload)
    flash(message, 'success')
    return redirect(url_for(endpoint))


def admin_error(message, endpoint, code=400):
    if request.is_json:
        return jsonify({'status': 'error', 'message': message}), code
    flash(message, 'danger')
    return redirect(url_for(endpoint))


def parse_date_filter(value, is_end=False):
    if not value:
        return None
    try:
        parsed = datetime.fromisoformat(value)
    except ValueError:
        try:
            parsed = datetime.strptime(value, '%Y-%m-%d')
        except ValueError:
            return None
    if is_end:
        parsed = parsed + timedelta(days=1)
    return parsed.isoformat()


@app.route('/admin/users')
@login_required
@admin_required
def admin_users():
    conn = get_connection()
    try:
        users = conn.execute(
            '''
                SELECT u.*, r.name AS role_name
                FROM users u
                JOIN user_roles r ON u.role_id = r.id
                ORDER BY LOWER(u.full_name)
            '''
        ).fetchall()
        roles = conn.execute('SELECT id, name, description FROM user_roles ORDER BY name').fetchall()
        modules = list_access_modules(conn)
        role_module_map = modules_by_role(conn)
    finally:
        conn.close()

    status_counts = defaultdict(int)
    user_payload = []
    for user in users:
        user_dict = dict(user)
        user_dict['modules'] = role_module_map.get(user['role_id'], [])
        status_counts[user_dict.get('status') or 'Unknown'] += 1
        user_payload.append(user_dict)

    return render_template(
        'admin_users.html',
        users=user_payload,
        roles=[dict(role) for role in roles],
        modules=modules,
        status_counts=dict(status_counts)
    )


@app.route('/admin/users/create', methods=['POST'])
@login_required
@admin_required
def admin_create_user():
    data = request.get_json() if request.is_json else request.form
    full_name = (data.get('full_name') or '').strip()
    username = (data.get('username') or '').strip()
    role_id = data.get('role_id')
    status = (data.get('status') or 'Active').title()
    password = data.get('password')

    if not all([full_name, username, role_id]):
        return admin_error('Full name, username, and role are required.', 'admin_users')

    try:
        role_id = int(role_id)
    except (TypeError, ValueError):
        return admin_error('Invalid role selected.', 'admin_users')

    generated_password = None
    if not password:
        generated_password = secrets.token_urlsafe(8)
        password = generated_password

    conn = get_connection()
    try:
        existing = conn.execute('SELECT id FROM users WHERE LOWER(username) = LOWER(?)', (username,)).fetchone()
        if existing:
            return admin_error('Username already exists.', 'admin_users')
        now = datetime.utcnow().isoformat()
        conn.execute(
            '''INSERT INTO users (full_name, username, password_hash, role_id, status, created_at, updated_at)
               VALUES (?, ?, ?, ?, ?, ?, ?)''',
            (full_name, username, hash_password(password), role_id, status, now, now)
        )
        user_id = conn.execute('SELECT last_insert_rowid() AS id').fetchone()['id']
        conn.commit()
    finally:
        conn.close()

    log_user_action('create', module_key='admin', description=f'Created user {username}', entity_type='user', entity_id=user_id)
    message = 'User account created successfully.'
    extra = {}
    if generated_password:
        message += f' Temporary password: {generated_password}'
        extra['temporary_password'] = generated_password
    return admin_success(message, 'admin_users', extra=extra)


@app.route('/admin/users/<int:user_id>/update', methods=['POST'])
@login_required
@admin_required
def admin_update_user(user_id):
    data = request.get_json() if request.is_json else request.form
    full_name = (data.get('full_name') or '').strip()
    username = (data.get('username') or '').strip()
    role_id = data.get('role_id')
    status = (data.get('status') or 'Active').title()

    if not all([full_name, username, role_id]):
        return admin_error('Full name, username, and role are required.', 'admin_users')

    try:
        role_id = int(role_id)
    except (TypeError, ValueError):
        return admin_error('Invalid role selected.', 'admin_users')

    conn = get_connection()
    try:
        existing = conn.execute(
            'SELECT id FROM users WHERE LOWER(username) = LOWER(?) AND id != ?',
            (username, user_id)
        ).fetchone()
        if existing:
            return admin_error('Username already exists.', 'admin_users')
        conn.execute(
            '''UPDATE users
               SET full_name = ?, username = ?, role_id = ?, status = ?, updated_at = ?
               WHERE id = ?''',
            (full_name, username, role_id, status, datetime.utcnow().isoformat(), user_id)
        )
        conn.commit()
    finally:
        conn.close()

    log_user_action('update', module_key='admin', description=f'Updated user {username}', entity_type='user', entity_id=user_id)
    return admin_success('User updated successfully.', 'admin_users')


@app.route('/admin/users/<int:user_id>/status', methods=['POST'])
@login_required
@admin_required
def admin_update_user_status(user_id):
    data = request.get_json() if request.is_json else request.form
    status = (data.get('status') or 'Active').title()
    suspended_until = parse_date_filter(data.get('suspended_until'), is_end=False)

    conn = get_connection()
    try:
        conn.execute(
            'UPDATE users SET status = ?, suspended_until = ?, updated_at = ? WHERE id = ?',
            (status, suspended_until, datetime.utcnow().isoformat(), user_id)
        )
        if conn.total_changes == 0:
            return admin_error('User not found.', 'admin_users', code=404)
        conn.commit()
    finally:
        conn.close()

    log_user_action('update', module_key='admin', description=f'Changed user status to {status}', entity_type='user', entity_id=user_id)
    return admin_success('User status updated.', 'admin_users')


@app.route('/admin/users/<int:user_id>/reset-password', methods=['POST'])
@login_required
@admin_required
def admin_reset_user_password(user_id):
    data = request.get_json() if request.is_json else request.form
    password = data.get('password')
    generated_password = None
    if not password:
        generated_password = secrets.token_urlsafe(8)
        password = generated_password

    conn = get_connection()
    try:
        conn.execute(
            'UPDATE users SET password_hash = ?, updated_at = ? WHERE id = ?',
            (hash_password(password), datetime.utcnow().isoformat(), user_id)
        )
        if conn.total_changes == 0:
            return admin_error('User not found.', 'admin_users', code=404)
        conn.commit()
    finally:
        conn.close()

    log_user_action('update', module_key='admin', description='Reset user password', entity_type='user', entity_id=user_id)
    message = 'Password reset successfully.'
    extra = {}
    if generated_password:
        message += f' New password: {generated_password}'
        extra['temporary_password'] = generated_password
    return admin_success(message, 'admin_users', extra=extra)


@app.route('/admin/users/<int:user_id>/delete', methods=['POST'])
@login_required
@admin_required
def admin_delete_user(user_id):
    if session.get('user_id') == user_id:
        return admin_error('You cannot delete your own account.', 'admin_users')
    conn = get_connection()
    try:
        conn.execute('DELETE FROM users WHERE id = ?', (user_id,))
        if conn.total_changes == 0:
            return admin_error('User not found.', 'admin_users', code=404)
        conn.commit()
    finally:
        conn.close()
    log_user_action('delete', module_key='admin', description='Deleted user account', entity_type='user', entity_id=user_id)
    return admin_success('User deleted successfully.', 'admin_users')


@app.route('/admin/roles')
@login_required
@admin_required
def admin_roles():
    conn = get_connection()
    try:
        roles = conn.execute('SELECT * FROM user_roles ORDER BY name').fetchall()
        modules = list_access_modules(conn)
        role_module_map = modules_by_role(conn)
    finally:
        conn.close()
    payload = []
    for role in roles:
        role_dict = dict(role)
        role_dict['modules'] = role_module_map.get(role['id'], [])
        payload.append(role_dict)
    return render_template('admin_roles.html', roles=payload, modules=modules)


@app.route('/admin/roles/create', methods=['POST'])
@login_required
@admin_required
def admin_create_role():
    data = request.get_json() if request.is_json else request.form
    name = (data.get('name') or '').strip()
    description = (data.get('description') or '').strip()
    module_keys_source = data.get('modules')
    if hasattr(data, 'getlist'):
        modules_from_form = data.getlist('modules')
        if modules_from_form:
            module_keys_source = modules_from_form
    module_keys = extract_module_keys(module_keys_source)

    if not name:
        return admin_error('Role name is required.', 'admin_roles')

    conn = get_connection()
    try:
        existing = conn.execute('SELECT id FROM user_roles WHERE LOWER(name) = LOWER(?)', (name,)).fetchone()
        if existing:
            return admin_error('Role name already exists.', 'admin_roles')
        now = datetime.utcnow().isoformat()
        conn.execute(
            'INSERT INTO user_roles (name, description, is_system, created_at, updated_at) VALUES (?, ?, 0, ?, ?)',
            (name, description, now, now)
        )
        role_id = conn.execute('SELECT last_insert_rowid() AS id').fetchone()['id']
        module_lookup = {module['module_key']: module['id'] for module in list_access_modules(conn)}
        for key in module_keys:
            module_id = module_lookup.get(key)
            if module_id:
                conn.execute(
                    'INSERT INTO role_permissions (role_id, module_id, granted, created_at, updated_at) VALUES (?, ?, 1, ?, ?)',
                    (role_id, module_id, now, now)
                )
        conn.commit()
    finally:
        conn.close()

    log_user_action('create', module_key='admin', description=f'Created role {name}', entity_type='role', entity_id=role_id)
    return admin_success('Role created successfully.', 'admin_roles')


@app.route('/admin/roles/<int:role_id>/update', methods=['POST'])
@login_required
@admin_required
def admin_update_role(role_id):
    data = request.get_json() if request.is_json else request.form
    name = (data.get('name') or '').strip()
    description = (data.get('description') or '').strip()
    module_keys_source = data.get('modules')
    if hasattr(data, 'getlist'):
        modules_from_form = data.getlist('modules')
        if modules_from_form:
            module_keys_source = modules_from_form
    module_keys = extract_module_keys(module_keys_source)

    if not name:
        return admin_error('Role name is required.', 'admin_roles')

    conn = get_connection()
    try:
        existing = conn.execute('SELECT * FROM user_roles WHERE id = ?', (role_id,)).fetchone()
        if not existing:
            return admin_error('Role not found.', 'admin_roles', code=404)
        duplicate = conn.execute(
            'SELECT id FROM user_roles WHERE LOWER(name) = LOWER(?) AND id != ?',
            (name, role_id)
        ).fetchone()
        if duplicate:
            return admin_error('Another role already uses this name.', 'admin_roles')

        now = datetime.utcnow().isoformat()
        conn.execute(
            'UPDATE user_roles SET name = ?, description = ?, updated_at = ? WHERE id = ?',
            (name, description, now, role_id)
        )
        conn.execute('DELETE FROM role_permissions WHERE role_id = ?', (role_id,))
        module_lookup = {module['module_key']: module['id'] for module in list_access_modules(conn)}
        for key in module_keys:
            module_id = module_lookup.get(key)
            if module_id:
                conn.execute(
                    'INSERT INTO role_permissions (role_id, module_id, granted, created_at, updated_at) VALUES (?, ?, 1, ?, ?)',
                    (role_id, module_id, now, now)
                )
        conn.commit()
    finally:
        conn.close()

    log_user_action('update', module_key='admin', description=f'Updated role {name}', entity_type='role', entity_id=role_id)
    return admin_success('Role updated successfully.', 'admin_roles')


@app.route('/admin/roles/<int:role_id>/delete', methods=['POST'])
@login_required
@admin_required
def admin_delete_role(role_id):
    conn = get_connection()
    try:
        role = conn.execute('SELECT * FROM user_roles WHERE id = ?', (role_id,)).fetchone()
        if not role:
            return admin_error('Role not found.', 'admin_roles', code=404)
        if role['is_system']:
            return admin_error('System roles cannot be deleted.', 'admin_roles')
        linked_user = conn.execute('SELECT id FROM users WHERE role_id = ? LIMIT 1', (role_id,)).fetchone()
        if linked_user:
            return admin_error('Cannot delete a role that is assigned to users.', 'admin_roles')
        role_name = role['name']
        conn.execute('DELETE FROM user_roles WHERE id = ?', (role_id,))
        conn.commit()
    finally:
        conn.close()
    log_user_action('delete', module_key='admin', description=f'Deleted role {role_name}', entity_type='role', entity_id=role_id)
    return admin_success('Role deleted successfully.', 'admin_roles')


@app.route('/admin/activity-log')
@login_required
@admin_required
def admin_activity_log():
    filters = {
        'module': request.args.get('module'),
        'username': request.args.get('username'),
        'action': request.args.get('action'),
        'start': parse_date_filter(request.args.get('start')),
        'end': parse_date_filter(request.args.get('end'), is_end=True)
    }
    query = 'SELECT * FROM user_activity_log WHERE 1=1'
    params = []
    if filters['module']:
        query += ' AND module_key = ?'
        params.append(filters['module'])
    if filters['username']:
        query += ' AND LOWER(username) = LOWER(?)'
        params.append(filters['username'])
    if filters['action']:
        query += ' AND LOWER(action) = LOWER(?)'
        params.append(filters['action'])
    if filters['start']:
        query += ' AND datetime(created_at) >= datetime(?)'
        params.append(filters['start'])
    if filters['end']:
        query += ' AND datetime(created_at) <= datetime(?)'
        params.append(filters['end'])
    query += ' ORDER BY datetime(created_at) DESC LIMIT 500'

    conn = get_connection()
    try:
        logs = [dict(row) for row in conn.execute(query, params).fetchall()]
    finally:
        conn.close()

    return render_template(
        'activity_log.html',
        logs=logs,
        modules=DEFAULT_MODULES,
        filters=request.args
    )


@app.route('/api/activity-log', methods=['GET'])
@login_required
@admin_required
def api_activity_log():
    module_key = request.args.get('module')
    try:
        limit = int(request.args.get('limit', 200))
    except (TypeError, ValueError):
        limit = 200
    limit = max(1, min(limit, 500))
    query = 'SELECT * FROM user_activity_log WHERE 1=1'
    params = []
    if module_key:
        query += ' AND module_key = ?'
        params.append(module_key)
    query += ' ORDER BY datetime(created_at) DESC LIMIT ?'
    params.append(limit)

    conn = get_connection()
    try:
        logs = [dict(row) for row in conn.execute(query, params).fetchall()]
    finally:
        conn.close()

    return jsonify({'status': 'success', 'logs': logs})

UPLOAD_FOLDER = os.path.join(app.root_path, 'uploads')
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

TEACHER_PERMISSION_OPTIONS = [
    {'id': 'student_management', 'label': 'Student Management'},
    {'id': 'mark_attendance', 'label': 'Attendance / Mark Attendance'},
    {'id': 'attendance_reports', 'label': 'Attendance Reports'},
    {'id': 'upload_questions', 'label': 'Question Bank / Upload Questions'},
    {'id': 'create_exam', 'label': 'Exam Creation'},
    {'id': 'exam_attendance', 'label': 'Exam Attendance'},
    {'id': 'view_results', 'label': 'Results & Reports'},
    {'id': 'leave_management', 'label': 'Leave Management'},
    {'id': 'payroll', 'label': 'Payroll / Finance'},
    {'id': 'hr_section', 'label': 'HR Section'},
    {'id': 'sms_module', 'label': 'SMS & Communication'}
]
VALID_TEACHER_PERMISSIONS = {perm['id'] for perm in TEACHER_PERMISSION_OPTIONS}

# ==================== TEACHER MANAGEMENT (ADMIN PANEL) ====================

@app.route("/admin/teachers")
@admin_required
def admin_teachers():
    conn = get_connection()
    # Get all teachers with employee info
    teachers = conn.execute('''
        SELECT t.id, t.username, t.name, t.role, t.assigned_semesters, t.employee_id, 
               t.subject, t.technology, t.technology_assignments, t.status, t.email,
               t.phone, t.cnic,
               e.name as employee_name, e.department_id, d.name as department_name,
               e.contact as employee_contact, e.cnic as employee_cnic
        FROM teachers t
        LEFT JOIN employees e ON t.employee_id = e.id
        LEFT JOIN departments d ON e.department_id = d.id
        ORDER BY t.name
    ''').fetchall()
    
    # Get filter options
    technologies = [row['name'] for row in conn.execute('SELECT name FROM technologies').fetchall()]
    semesters = [row['name'] for row in conn.execute('SELECT name FROM semesters').fetchall()]
    departments = [row['name'] for row in conn.execute('SELECT name FROM departments').fetchall()]
    
    conn.close()
    
    # Convert teachers to dicts with permissions
    teachers_list = []
    for teacher in teachers:
        teacher_dict = dict(teacher)
        teacher_dict['assigned_semesters'] = parse_multi_value(teacher_dict.get('assigned_semesters'))
        teacher_dict['assigned_semesters_str'] = join_display(teacher_dict['assigned_semesters'])
        technology_list = parse_multi_value(teacher_dict.get('technology_assignments') or teacher_dict.get('technology'))
        teacher_dict['technology_assignments'] = technology_list
        teacher_dict['technology'] = join_display(technology_list)
        
        # Get permissions for this teacher
        conn = get_connection()
        permissions = conn.execute(
            'SELECT permission_name FROM teacher_permissions WHERE teacher_id = ? AND granted = 1',
            (teacher['id'],)
        ).fetchall()
        teacher_dict['permissions'] = [p['permission_name'] for p in permissions]
        conn.close()
        
        teachers_list.append(teacher_dict)
    
    total_teachers = len(teachers_list)
    active_teachers = sum(1 for teacher in teachers_list if (teacher.get('status') or '').lower() == 'active')
    inactive_teachers = total_teachers - active_teachers
    unique_semesters = sorted({sem for teacher in teachers_list for sem in (teacher.get('assigned_semesters') or [])})
    unique_technologies = sorted({tech for teacher in teachers_list for tech in (teacher.get('technology_assignments') or [])})
    
    return render_template('admin_teachers.html', 
                          teachers=teachers_list,
                          technologies=technologies,
                          semesters=semesters,
                          departments=departments,
                          total_teachers=total_teachers,
                          active_teachers=active_teachers,
                          inactive_teachers=inactive_teachers,
                          unique_semesters=unique_semesters,
                          unique_technologies=unique_technologies)

@app.route("/admin/teachers/add", methods=['GET', 'POST'])
@admin_required
def add_teacher():
    conn = get_connection()
    semesters = [row['name'] for row in conn.execute('SELECT name FROM semesters').fetchall()]
    technologies = [row['name'] for row in conn.execute('SELECT name FROM technologies').fetchall()]
    employees = conn.execute('SELECT id, name, email, department_id FROM employees WHERE status = "Active"').fetchall()
    conn.close()

    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        name = request.form['name']
        role = request.form.get('role', 'teacher')
        email = (request.form.get('email', '') or '').strip()
        phone = normalize_phone(request.form.get('phone', ''))
        cnic = normalize_cnic(request.form.get('cnic', ''))
        employee_id = request.form.get('employee_id', '') or None
        subject = request.form.get('subject', '')
        assigned_semesters = [sem.strip() for sem in request.form.getlist('assigned_semesters') if sem.strip()]
        assigned_semesters = list(dict.fromkeys(assigned_semesters))
        technology_list = request.form.getlist('technologies') or request.form.getlist('technology')
        technology_list = [tech.strip() for tech in technology_list if tech and tech.strip()]
        technology_list = list(dict.fromkeys(technology_list))
        permissions = request.form.getlist('permissions')  # Get list of selected permissions

        if not username or not password or not name:
            flash('Username, password, and name are required.', 'danger')
            return render_template('add_edit_teacher.html', 
                                 semesters=semesters, 
                                 technologies=technologies,
                                 employees=employees,
                                 teacher=None)
        if not assigned_semesters:
            flash('Please select at least one semester.', 'danger')
            return render_template('add_edit_teacher.html', 
                                 semesters=semesters, 
                                 technologies=technologies,
                                 employees=employees,
                                 teacher=None)
        if not technology_list:
            flash('Please select at least one technology/program.', 'danger')
            return render_template('add_edit_teacher.html', 
                                 semesters=semesters, 
                                 technologies=technologies,
                                 employees=employees,
                                 teacher=None)

        password_hash = generate_password_hash(password)
        assigned_semesters_json = serialize_multi_value(assigned_semesters)
        technology_assignments_json = serialize_multi_value(technology_list)
        technology_display = join_display(technology_list)

        conn = get_connection()
        try:
            if teacher_email_exists(conn, email):
                flash('This teacher already exists in the system.', 'warning')
                return render_template('add_edit_teacher.html', 
                                     semesters=semesters, 
                                     technologies=technologies,
                                     employees=employees,
                                     teacher=None)
            cursor = conn.cursor()
            cursor.execute(
                '''INSERT INTO teachers (username, password_hash, name, role, assigned_semesters, 
                   employee_id, subject, technology, technology_assignments, status, email, phone, cnic, created_at, updated_at) 
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                (username, password_hash, name, role, assigned_semesters_json, employee_id, 
                 subject, technology_display, technology_assignments_json, 'Active', email or None,
                 phone or None, cnic or None, datetime.now().isoformat(), datetime.now().isoformat())
            )
            teacher_id = cursor.lastrowid
            
            # Save permissions
            for permission in permissions:
                conn.execute(
                    'INSERT INTO teacher_permissions (teacher_id, permission_name, granted, created_at, updated_at) VALUES (?, ?, ?, ?, ?)',
                    (teacher_id, permission, 1, datetime.now().isoformat(), datetime.now().isoformat())
                )
            
            conn.commit()
            flash(f'Teacher {name} added successfully!', 'success')
            return redirect(url_for('admin_teachers'))
        except sqlite3.IntegrityError:
            flash('Username already exists. Please choose a different username.', 'danger')
        except Exception as e:
            flash(f'Error adding teacher: {str(e)}', 'danger')
        finally:
            conn.close()
    
    return render_template('add_edit_teacher.html', 
                          semesters=semesters,
                          technologies=technologies,
                          employees=employees,
                          teacher=None)

@app.route("/admin/teachers/edit/<int:teacher_id>", methods=['GET', 'POST'])
@admin_required
def edit_teacher(teacher_id):
    conn = get_connection()
    teacher = conn.execute('''
        SELECT t.*, e.name as employee_name 
        FROM teachers t 
        LEFT JOIN employees e ON t.employee_id = e.id 
        WHERE t.id = ?
    ''', (teacher_id,)).fetchone()
    semesters = [row['name'] for row in conn.execute('SELECT name FROM semesters').fetchall()]
    technologies = [row['name'] for row in conn.execute('SELECT name FROM technologies').fetchall()]
    employees = conn.execute('SELECT id, name, email, department_id FROM employees WHERE status = "Active"').fetchall()
    
    # Get current permissions
    current_permissions = conn.execute(
        'SELECT permission_name FROM teacher_permissions WHERE teacher_id = ? AND granted = 1',
        (teacher_id,)
    ).fetchall()
    current_permission_list = [p['permission_name'] for p in current_permissions]
    
    conn.close()

    if not teacher:
        flash('Teacher not found.', 'danger')
        return redirect(url_for('admin_teachers'))

    teacher_dict = dict(teacher)
    teacher_dict['assigned_semesters'] = parse_multi_value(teacher_dict.get('assigned_semesters'))
    teacher_dict['technology_assignments'] = parse_multi_value(teacher_dict.get('technology_assignments') or teacher_dict.get('technology'))
    teacher_dict['technology'] = join_display(teacher_dict['technology_assignments'])
    teacher_dict['permissions'] = current_permission_list

    if request.method == 'POST':
        name = request.form['name']
        role = request.form.get('role', 'teacher')
        email = (request.form.get('email', '') or '').strip()
        phone = normalize_phone(request.form.get('phone', ''))
        cnic = normalize_cnic(request.form.get('cnic', ''))
        employee_id = request.form.get('employee_id', '') or None
        subject = request.form.get('subject', '')
        assigned_semesters = [sem.strip() for sem in request.form.getlist('assigned_semesters') if sem.strip()]
        assigned_semesters = list(dict.fromkeys(assigned_semesters))
        technology_list = request.form.getlist('technologies') or request.form.getlist('technology')
        technology_list = [tech.strip() for tech in technology_list if tech and tech.strip()]
        technology_list = list(dict.fromkeys(technology_list))
        permissions = request.form.getlist('permissions')
        password = request.form.get('password')
        status = request.form.get('status', 'Active')

        if not name:
            flash('Teacher name is required.', 'danger')
            return render_template('add_edit_teacher.html', 
                                 semesters=semesters,
                                 technologies=technologies,
                                 employees=employees,
                                 teacher=teacher_dict)
        if not assigned_semesters:
            flash('Please select at least one semester.', 'danger')
            return render_template('add_edit_teacher.html', 
                                 semesters=semesters,
                                 technologies=technologies,
                                 employees=employees,
                                 teacher=teacher_dict)
        if not technology_list:
            flash('Please select at least one technology/program.', 'danger')
            return render_template('add_edit_teacher.html', 
                                 semesters=semesters,
                                 technologies=technologies,
                                 employees=employees,
                                 teacher=teacher_dict)

        assigned_semesters_json = serialize_multi_value(assigned_semesters)
        technology_assignments_json = serialize_multi_value(technology_list)
        technology_display = join_display(technology_list)
        update_fields = ['name = ?', 'role = ?', 'assigned_semesters = ?', 'employee_id = ?', 
                        'subject = ?', 'technology = ?', 'technology_assignments = ?', 'status = ?', 'email = ?', 'phone = ?', 'cnic = ?', 'updated_at = ?']
        update_params = [name, role, assigned_semesters_json, employee_id, subject, technology_display, technology_assignments_json, 
                        status, email or None, phone or None, cnic or None, datetime.now().isoformat()]

        if password:
            password_hash = generate_password_hash(password)
            update_fields.append('password_hash = ?')
            update_params.append(password_hash)
        
        update_params.append(teacher_id)

        conn = get_connection()
        try:
            if teacher_email_exists(conn, email, exclude_id=teacher_id):
                flash('This teacher already exists in the system.', 'warning')
                return render_template('add_edit_teacher.html', 
                                      semesters=semesters,
                                      technologies=technologies,
                                      employees=employees,
                                      teacher=teacher_dict)
            conn.execute(
                f"UPDATE teachers SET {', '.join(update_fields)} WHERE id = ?",
                update_params
            )
            
            # Update permissions - delete all and re-insert
            conn.execute('DELETE FROM teacher_permissions WHERE teacher_id = ?', (teacher_id,))
            for permission in permissions:
                conn.execute(
                    'INSERT INTO teacher_permissions (teacher_id, permission_name, granted, created_at, updated_at) VALUES (?, ?, ?, ?, ?)',
                    (teacher_id, permission, 1, datetime.now().isoformat(), datetime.now().isoformat())
                )
            
            conn.commit()
            flash(f'Teacher {name} updated successfully!', 'success')
            return redirect(url_for('admin_teachers'))
        except Exception as e:
            flash(f'Error updating teacher: {str(e)}', 'danger')
        finally:
            conn.close()

    return render_template('add_edit_teacher.html', 
                          semesters=semesters,
                          technologies=technologies,
                          employees=employees,
                          teacher=teacher_dict)

@app.route("/admin/teachers/delete/<int:teacher_id>", methods=['POST'])
@admin_required
def delete_teacher(teacher_id):
    conn = get_connection()
    try:
        # Get teacher name before deletion for flash message
        teacher = conn.execute('SELECT name FROM teachers WHERE id = ?', (teacher_id,)).fetchone()
        teacher_name = teacher['name'] if teacher else 'Teacher'
        
        # Delete permissions first (CASCADE should handle this, but being explicit)
        conn.execute('DELETE FROM teacher_permissions WHERE teacher_id = ?', (teacher_id,))
        # Delete activity logs
        conn.execute('DELETE FROM teacher_activity_log WHERE teacher_id = ?', (teacher_id,))
        # Delete teacher
        conn.execute('DELETE FROM teachers WHERE id = ?', (teacher_id,))
        conn.commit()
        flash(f'Teacher {teacher_name} deleted successfully!', 'success')
    except Exception as e:
        flash(f'Error deleting teacher: {str(e)}', 'danger')
    return redirect(url_for('admin_teachers'))

@app.route("/admin/teachers/deactivate/<int:teacher_id>", methods=['POST'])
@admin_required
def deactivate_teacher(teacher_id):
    conn = get_connection()
    try:
        conn.execute('UPDATE teachers SET status = ? WHERE id = ?', ('Inactive', teacher_id))
        conn.commit()
        flash('Teacher deactivated successfully!', 'success')
    except Exception as e:
        flash(f'Error deactivating teacher: {str(e)}', 'danger')
    finally:
        conn.close()
    return redirect(url_for('admin_teachers'))

@app.route("/admin/teachers/reset_password/<int:teacher_id>", methods=['POST'])
@admin_required
def reset_teacher_password(teacher_id):
    conn = get_connection()
    try:
        new_password = request.form.get('new_password', 'password123')  # Default password
        password_hash = generate_password_hash(new_password)
        conn.execute('UPDATE teachers SET password_hash = ?, updated_at = ? WHERE id = ?', 
                    (password_hash, datetime.now().isoformat(), teacher_id))
        conn.commit()
        flash('Password reset successfully!', 'success')
    except Exception as e:
        flash(f'Error resetting password: {str(e)}', 'danger')
    finally:
        conn.close()
    return redirect(url_for('admin_teachers'))

# API endpoint to get teacher's assigned semesters (for frontend use)
@app.route("/api/teacher_semesters", methods=['GET'])
@login_required
def get_teacher_semesters():
    if session.get('role') == 'admin':
        conn = get_connection()
        all_semesters = [row['name'] for row in conn.execute('SELECT name FROM semesters').fetchall()]
        conn.close()
        return jsonify(all_semesters)
    else:
        return jsonify(session.get('assigned_semesters', []))

# ==================== TEACHER MANAGEMENT API ENDPOINTS ====================

@app.route("/api/teachers", methods=['GET'])
@admin_required
def api_get_teachers():
    """Get all teachers with filters"""
    try:
        conn = get_connection()
        technology = request.args.get('technology', '').strip()
        semester = request.args.get('semester', '').strip()
        department = request.args.get('department', '').strip()
        status = request.args.get('status', '').strip()
        search = request.args.get('search', '').strip()
        
        query = '''
            SELECT t.id, t.username, t.name, t.role, t.assigned_semesters, t.employee_id, 
                   t.subject, t.technology, t.technology_assignments, t.status, t.email, t.phone, t.cnic,
                   t.created_at, t.updated_at,
                   e.name as employee_name, e.contact as employee_contact, e.cnic as employee_cnic,
                   d.name as department_name, des.name as designation_name
            FROM teachers t
            LEFT JOIN employees e ON t.employee_id = e.id
            LEFT JOIN departments d ON e.department_id = d.id
            LEFT JOIN designations des ON e.designation_id = des.id
            WHERE 1=1
        '''
        params = []
        
        if technology:
            query += ' AND (IFNULL(t.technology, \'\') LIKE ? OR IFNULL(t.technology_assignments, \'\') LIKE ?)'
            like_value = f'%{technology}%'
            params.extend([like_value, like_value])
        if semester:
            query += ' AND t.assigned_semesters LIKE ?'
            params.append(f'%{semester}%')
        if department:
            query += ' AND d.name = ?'
            params.append(department)
        if status:
            query += ' AND t.status = ?'
            params.append(status)
        if search:
            query += ' AND (t.name LIKE ? OR t.username LIKE ? OR IFNULL(e.name, \'\') LIKE ?)'
            like_value = f'%{search}%'
            params.extend([like_value, like_value, like_value])
        
        query += ' ORDER BY t.name'
        
        teachers = conn.execute(query, params).fetchall()
        
        teacher_ids = [teacher['id'] for teacher in teachers]
        permissions_map = {}
        if teacher_ids:
            placeholders = ','.join(['?'] * len(teacher_ids))
            perm_rows = conn.execute(f'''
                SELECT teacher_id, permission_name 
                FROM teacher_permissions 
                WHERE teacher_id IN ({placeholders}) AND granted = 1
            ''', teacher_ids).fetchall()
            for row in perm_rows:
                permissions_map.setdefault(row['teacher_id'], []).append(row['permission_name'])
        
        teachers_list = []
        for teacher in teachers:
            teacher_dict = dict(teacher)
            teacher_dict['assigned_semesters'] = parse_multi_value(teacher_dict.get('assigned_semesters'))
            technology_list = parse_multi_value(teacher_dict.get('technology_assignments') or teacher_dict.get('technology'))
            teacher_dict['technology_assignments'] = technology_list
            teacher_dict['technology'] = join_display(technology_list)
            teacher_dict['permissions'] = permissions_map.get(teacher_dict['id'], [])
            teacher_dict['permissions_count'] = len(teacher_dict['permissions'])
            teachers_list.append(teacher_dict)
        
        conn.close()
        return jsonify({'status': 'success', 'teachers': teachers_list})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route("/api/teachers/<int:teacher_id>", methods=['GET'])
@admin_required
def api_get_teacher(teacher_id):
    """Get single teacher details"""
    try:
        conn = get_connection()
        teacher = conn.execute('''
            SELECT t.*, e.name as employee_name, e.contact as employee_contact, e.cnic as employee_cnic,
                   d.name as department_name
            FROM teachers t
            LEFT JOIN employees e ON t.employee_id = e.id
            LEFT JOIN departments d ON e.department_id = d.id
            WHERE t.id = ?
        ''', (teacher_id,)).fetchone()
        
        if not teacher:
            return jsonify({'status': 'error', 'message': 'Teacher not found'}), 404
        
        teacher_dict = dict(teacher)
        teacher_dict['assigned_semesters'] = parse_multi_value(teacher_dict.get('assigned_semesters'))
        technology_list = parse_multi_value(teacher_dict.get('technology_assignments') or teacher_dict.get('technology'))
        teacher_dict['technology_assignments'] = technology_list
        teacher_dict['technology'] = join_display(technology_list)
        
        # Get permissions
        permissions = conn.execute(
            'SELECT permission_name FROM teacher_permissions WHERE teacher_id = ? AND granted = 1',
            (teacher_id,)
        ).fetchall()
        teacher_dict['permissions'] = [p['permission_name'] for p in permissions]
        
        conn.close()
        return jsonify({'status': 'success', 'teacher': teacher_dict})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route("/api/teachers/<int:teacher_id>/permissions", methods=['GET'])
@admin_required
def api_get_teacher_permissions(teacher_id):
    """Get teacher permissions"""
    try:
        conn = get_connection()
        permissions = conn.execute(
            'SELECT permission_name, granted FROM teacher_permissions WHERE teacher_id = ?',
            (teacher_id,)
        ).fetchall()
        conn.close()
        
        perm_dict = {p['permission_name']: bool(p['granted']) for p in permissions}
        return jsonify({'status': 'success', 'permissions': perm_dict})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route("/api/teachers/<int:teacher_id>/activity", methods=['GET'])
@admin_required
def api_get_teacher_activity(teacher_id):
    """Get teacher activity log"""
    try:
        limit = request.args.get('limit', 50, type=int)
        conn = get_connection()
        activities = conn.execute('''
            SELECT * FROM teacher_activity_log 
            WHERE teacher_id = ? 
            ORDER BY created_at DESC 
            LIMIT ?
        ''', (teacher_id, limit)).fetchall()
        conn.close()
        
        return jsonify({'status': 'success', 'activities': [dict(a) for a in activities]})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route("/api/teacher_permissions/options", methods=['GET'])
@admin_required
def api_teacher_permission_options():
    """Return available permission options"""
    return jsonify({'status': 'success', 'permissions': TEACHER_PERMISSION_OPTIONS})

@app.route("/api/teachers", methods=['POST'])
@admin_required
def api_create_teacher():
    """Create a new teacher record"""
    data = request.get_json() or {}
    name = (data.get('name') or '').strip()
    username = (data.get('username') or '').strip()
    password = data.get('password') or ''
    confirm_password = data.get('confirm_password') or ''
    status = (data.get('status') or 'Active').strip() or 'Active'
    employee_id = data.get('employee_id')
    subject = (data.get('subject') or '').strip()
    technologies = data.get('technologies')
    if technologies is None:
        technologies = data.get('technology_assignments')
    if technologies is None and data.get('technology'):
        technologies = parse_multi_value(data.get('technology'))
    email = (data.get('email') or '').strip()
    phone = normalize_phone(data.get('phone'))
    cnic = normalize_cnic(data.get('cnic'))
    assigned_semesters = data.get('assigned_semesters') or []
    permissions = data.get('permissions') or []
    role = data.get('role') or 'teacher'
    
    if not name or not username or not password:
        return jsonify({'status': 'error', 'message': 'Name, username, and password are required.'}), 400
    if password != confirm_password:
        return jsonify({'status': 'error', 'message': 'Passwords do not match.'}), 400
    if status not in ['Active', 'Inactive']:
        return jsonify({'status': 'error', 'message': 'Invalid status provided.'}), 400
    
    if not isinstance(assigned_semesters, list):
        assigned_semesters = []
    assigned_semesters = [str(item).strip() for item in assigned_semesters if str(item).strip()]
    assigned_semesters = list(dict.fromkeys(assigned_semesters))
    if not assigned_semesters:
        return jsonify({'status': 'error', 'message': 'Select at least one semester.'}), 400
    technology_list = parse_multi_value(technologies)
    technology_list = list(dict.fromkeys(technology_list))
    if not technology_list:
        return jsonify({'status': 'error', 'message': 'Select at least one technology/program.'}), 400
    valid_permissions = [perm for perm in permissions if perm in VALID_TEACHER_PERMISSIONS]
    
    conn = get_connection()
    try:
        existing = conn.execute('SELECT id FROM teachers WHERE username = ?', (username,)).fetchone()
        if existing:
            return jsonify({'status': 'error', 'message': 'Username already exists.'}), 400
        if teacher_email_exists(conn, email):
            return jsonify({'status': 'error', 'message': 'This teacher already exists in the system.'}), 400
        
        password_hash = generate_password_hash(password)
        assigned_semesters_json = serialize_multi_value(assigned_semesters)
        technology_assignments_json = serialize_multi_value(technology_list)
        technology_display = join_display(technology_list)
        now = datetime.now().isoformat()
        
        cursor = conn.cursor()
        cursor.execute(
            '''INSERT INTO teachers (username, password_hash, name, role, assigned_semesters, 
               employee_id, subject, technology, technology_assignments, status, email, phone, cnic, created_at, updated_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
            (
                username, password_hash, name, role, assigned_semesters_json,
                employee_id if employee_id else None, subject, technology_display,
                technology_assignments_json, status, email or None,
                phone or None, cnic or None, now, now
            )
        )
        teacher_id = cursor.lastrowid
        
        if valid_permissions:
            cursor.executemany(
                'INSERT INTO teacher_permissions (teacher_id, permission_name, granted, created_at, updated_at) VALUES (?, ?, 1, ?, ?)',
                [(teacher_id, perm, now, now) for perm in valid_permissions]
            )
        
        conn.commit()
        return jsonify({'status': 'success', 'message': 'Teacher added successfully.', 'teacher_id': teacher_id})
    except sqlite3.IntegrityError:
        conn.rollback()
        return jsonify({'status': 'error', 'message': 'Unable to save teacher. Please verify the information.'}), 400
    except Exception as e:
        conn.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        conn.close()

@app.route("/api/teachers/<int:teacher_id>", methods=['PUT'])
@admin_required
def api_update_teacher(teacher_id):
    """Update teacher details"""
    data = request.get_json() or {}
    name = (data.get('name') or '').strip()
    username = (data.get('username') or '').strip()
    status = (data.get('status') or 'Active').strip() or 'Active'
    employee_id = data.get('employee_id')
    subject = (data.get('subject') or '').strip()
    technologies = data.get('technologies')
    if technologies is None:
        technologies = data.get('technology_assignments')
    if technologies is None and data.get('technology'):
        technologies = parse_multi_value(data.get('technology'))
    email = (data.get('email') or '').strip()
    phone = normalize_phone(data.get('phone'))
    cnic = normalize_cnic(data.get('cnic'))
    assigned_semesters = data.get('assigned_semesters') or []
    permissions = data.get('permissions') or []
    role = data.get('role') or 'teacher'
    new_password = data.get('password') or ''
    confirm_password = data.get('confirm_password') or ''
    
    if not name or not username:
        return jsonify({'status': 'error', 'message': 'Name and username are required.'}), 400
    if status not in ['Active', 'Inactive']:
        return jsonify({'status': 'error', 'message': 'Invalid status provided.'}), 400
    if new_password and new_password != confirm_password:
        return jsonify({'status': 'error', 'message': 'Passwords do not match.'}), 400
    
    if not isinstance(assigned_semesters, list):
        assigned_semesters = []
    assigned_semesters = [str(item).strip() for item in assigned_semesters if str(item).strip()]
    assigned_semesters = list(dict.fromkeys(assigned_semesters))
    if not assigned_semesters:
        return jsonify({'status': 'error', 'message': 'Select at least one semester.'}), 400
    technology_list = parse_multi_value(technologies)
    technology_list = list(dict.fromkeys(technology_list))
    if not technology_list:
        return jsonify({'status': 'error', 'message': 'Select at least one technology/program.'}), 400
    valid_permissions = [perm for perm in permissions if perm in VALID_TEACHER_PERMISSIONS]
    
    conn = get_connection()
    try:
        existing = conn.execute('SELECT id FROM teachers WHERE username = ? AND id != ?', (username, teacher_id)).fetchone()
        if existing:
            return jsonify({'status': 'error', 'message': 'Username already exists.'}), 400
        if teacher_email_exists(conn, email, exclude_id=teacher_id):
            return jsonify({'status': 'error', 'message': 'This teacher already exists in the system.'}), 400
        
        assigned_semesters_json = serialize_multi_value(assigned_semesters)
        technology_assignments_json = serialize_multi_value(technology_list)
        technology_display = join_display(technology_list)
        now = datetime.now().isoformat()
        
        update_fields = [
            'name = ?', 'username = ?', 'role = ?', 'assigned_semesters = ?', 'employee_id = ?',
            'subject = ?', 'technology = ?', 'technology_assignments = ?', 'status = ?', 'email = ?', 'phone = ?', 'cnic = ?', 'updated_at = ?'
        ]
        params = [
            name, username, role, assigned_semesters_json,
            employee_id if employee_id else None, subject, technology_display,
            technology_assignments_json, status, email or None, phone or None, cnic or None, now
        ]
        
        if new_password:
            update_fields.insert(2, 'password_hash = ?')
            params.insert(2, generate_password_hash(new_password))
        
        params.append(teacher_id)
        conn.execute(f'UPDATE teachers SET {", ".join(update_fields)} WHERE id = ?', params)
        
        conn.execute('DELETE FROM teacher_permissions WHERE teacher_id = ?', (teacher_id,))
        if valid_permissions:
            conn.executemany(
                'INSERT INTO teacher_permissions (teacher_id, permission_name, granted, created_at, updated_at) VALUES (?, ?, 1, ?, ?)',
                [(teacher_id, perm, now, now) for perm in valid_permissions]
            )
        
        conn.commit()
        return jsonify({'status': 'success', 'message': 'Teacher updated successfully.'})
    except sqlite3.IntegrityError:
        conn.rollback()
        return jsonify({'status': 'error', 'message': 'Unable to update teacher.'}), 400
    except Exception as e:
        conn.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        conn.close()

@app.route("/api/teachers/<int:teacher_id>", methods=['DELETE'])
@admin_required
def api_delete_teacher(teacher_id):
    """Delete teacher"""
    conn = get_connection()
    try:
        teacher = conn.execute('SELECT name FROM teachers WHERE id = ?', (teacher_id,)).fetchone()
        if not teacher:
            return jsonify({'status': 'error', 'message': 'Teacher not found'}), 404
        
        conn.execute('DELETE FROM teacher_permissions WHERE teacher_id = ?', (teacher_id,))
        conn.execute('DELETE FROM teacher_activity_log WHERE teacher_id = ?', (teacher_id,))
        conn.execute('DELETE FROM teachers WHERE id = ?', (teacher_id,))
        conn.commit()
        return jsonify({'status': 'success', 'message': 'Teacher deleted successfully'})
    except Exception as e:
        conn.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        conn.close()

@app.route("/api/teachers/<int:teacher_id>/status", methods=['POST'])
@admin_required
def api_toggle_teacher_status(teacher_id):
    """Update teacher status"""
    data = request.get_json() or {}
    status = (data.get('status') or '').strip()
    if status not in ['Active', 'Inactive']:
        return jsonify({'status': 'error', 'message': 'Invalid status provided.'}), 400
    conn = get_connection()
    try:
        updated = conn.execute('UPDATE teachers SET status = ?, updated_at = ? WHERE id = ?', (status, datetime.now().isoformat(), teacher_id))
        conn.commit()
        if updated.rowcount == 0:
            return jsonify({'status': 'error', 'message': 'Teacher not found'}), 404
        return jsonify({'status': 'success', 'message': 'Status updated'})
    except Exception as e:
        conn.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        conn.close()

@app.route("/api/teachers/<int:teacher_id>/reset-password", methods=['POST'])
@admin_required
def api_reset_teacher_password(teacher_id):
    """Reset teacher password"""
    data = request.get_json() or {}
    new_password = data.get('password') or ''
    confirm_password = data.get('confirm_password') or ''
    if not new_password:
        return jsonify({'status': 'error', 'message': 'New password is required.'}), 400
    if new_password != confirm_password:
        return jsonify({'status': 'error', 'message': 'Passwords do not match.'}), 400
    conn = get_connection()
    try:
        updated = conn.execute(
            'UPDATE teachers SET password_hash = ?, updated_at = ? WHERE id = ?',
            (generate_password_hash(new_password), datetime.now().isoformat(), teacher_id)
        )
        conn.commit()
        if updated.rowcount == 0:
            return jsonify({'status': 'error', 'message': 'Teacher not found'}), 404
        return jsonify({'status': 'success', 'message': 'Password reset successfully'})
    except Exception as e:
        conn.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        conn.close()

@app.route("/api/teachers/<int:teacher_id>/permissions", methods=['PUT'])
@admin_required
def api_update_teacher_permissions(teacher_id):
    """Update teacher permissions only"""
    data = request.get_json() or {}
    permissions = data.get('permissions') or []
    valid_permissions = [perm for perm in permissions if perm in VALID_TEACHER_PERMISSIONS]
    now = datetime.now().isoformat()
    conn = get_connection()
    try:
        exists = conn.execute('SELECT id FROM teachers WHERE id = ?', (teacher_id,)).fetchone()
        if not exists:
            return jsonify({'status': 'error', 'message': 'Teacher not found'}), 404
        
        conn.execute('DELETE FROM teacher_permissions WHERE teacher_id = ?', (teacher_id,))
        if valid_permissions:
            conn.executemany(
                'INSERT INTO teacher_permissions (teacher_id, permission_name, granted, created_at, updated_at) VALUES (?, ?, 1, ?, ?)',
                [(teacher_id, perm, now, now) for perm in valid_permissions]
            )
        conn.commit()
        return jsonify({'status': 'success', 'message': 'Permissions updated'})
    except Exception as e:
        conn.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        conn.close()

# ==================== INVENTORY MANAGEMENT APIs ====================

@app.route("/api/inventory/dashboard", methods=['GET'])
@login_required
def api_inventory_dashboard():
    guard = ensure_inventory_permission(INVENTORY_REPORT_ROLES)
    if guard:
        return guard
    conn = get_connection()
    try:
        summary = inventory_counts(conn)
        return jsonify({'status': 'success', 'summary': summary})
    finally:
        conn.close()

@app.route("/api/inventory/items", methods=['GET'])
@login_required
def api_inventory_items():
    guard = ensure_inventory_permission(INVENTORY_REPORT_ROLES)
    if guard:
        return guard
    page = max(request.args.get('page', 1, type=int), 1)
    per_page = min(request.args.get('per_page', 20, type=int), 100)
    offset = (page - 1) * per_page
    search = (request.args.get('search') or '').strip()
    category = (request.args.get('category') or '').strip()
    status = (request.args.get('status') or '').strip()
    location = (request.args.get('location') or '').strip()

    conn = get_connection()
    try:
        base_query = 'FROM inventory_items WHERE 1=1'
        params = []
        if search:
            base_query += ' AND (LOWER(item_name) LIKE ? OR LOWER(category) LIKE ?)'
            like = f'%{search.lower()}%'
            params.extend([like, like])
        if category:
            base_query += ' AND LOWER(category) = ?'
            params.append(category.lower())
        if status:
            base_query += ' AND LOWER(status) = ?'
            params.append(status.lower())
        if location:
            base_query += ' AND LOWER(location) = ?'
            params.append(location.lower())

        total = conn.execute(f'SELECT COUNT(*) AS c {base_query}', params).fetchone()['c']
        rows = conn.execute(
            f'SELECT * {base_query} ORDER BY item_name ASC LIMIT ? OFFSET ?',
            (*params, per_page, offset)
        ).fetchall()
        items = [serialize_inventory_item(row) for row in rows]
        return jsonify({
            'status': 'success',
            'items': items,
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': total
            }
        })
    finally:
        conn.close()

@app.route("/api/inventory/items", methods=['POST'])
@login_required
def api_inventory_create_item():
    guard = ensure_inventory_permission()
    if guard:
        return guard
    data = request.get_json() or {}
    payload = inventory_item_payload(data)
    if not payload['item_name']:
        return jsonify({'status': 'error', 'message': 'Item name is required.'}), 400
    now = datetime.now().isoformat()
    payload['created_at'] = now
    payload['updated_at'] = now
    payload['quantity_available'] = payload['quantity_total']

    conn = get_connection()
    try:
        placeholders = ','.join(['?'] * len(payload))
        conn.execute(
            f"INSERT INTO inventory_items ({', '.join(payload.keys())}) VALUES ({placeholders})",
            tuple(payload.values())
        )
        conn.commit()
        return jsonify({'status': 'success', 'message': 'Inventory item added successfully.'})
    finally:
        conn.close()

@app.route("/api/inventory/items/<int:item_id>", methods=['GET'])
@login_required
def api_inventory_get_item(item_id):
    guard = ensure_inventory_permission(INVENTORY_REPORT_ROLES)
    if guard:
        return guard
    conn = get_connection()
    try:
        row = conn.execute('SELECT * FROM inventory_items WHERE item_id = ?', (item_id,)).fetchone()
        if not row:
            return jsonify({'status': 'error', 'message': 'Item not found'}), 404
        return jsonify({'status': 'success', 'item': serialize_inventory_item(row)})
    finally:
        conn.close()

@app.route("/api/inventory/items/<int:item_id>", methods=['PUT'])
@login_required
def api_inventory_update_item(item_id):
    guard = ensure_inventory_permission()
    if guard:
        return guard
    data = request.get_json() or {}
    conn = get_connection()
    try:
        existing = conn.execute('SELECT * FROM inventory_items WHERE item_id = ?', (item_id,)).fetchone()
        if not existing:
            return jsonify({'status': 'error', 'message': 'Item not found'}), 404
        existing_dict = serialize_inventory_item(existing)
        payload = inventory_item_payload(data)
        payload['updated_at'] = datetime.now().isoformat()
        # Adjust quantity_available relative to issued stock if necessary
        issued_out = existing_dict['quantity_total'] - existing_dict['quantity_available']
        if payload['quantity_total'] < issued_out:
            return jsonify({'status': 'error', 'message': 'Quantity total cannot be less than already issued quantity.'}), 400
        payload['quantity_available'] = payload['quantity_total'] - issued_out
        set_clause = ', '.join([f"{key} = ?" for key in payload.keys()])
        conn.execute(
            f"UPDATE inventory_items SET {set_clause} WHERE item_id = ?",
            (*payload.values(), item_id)
        )
        conn.commit()
        updated_row = conn.execute('SELECT * FROM inventory_items WHERE item_id = ?', (item_id,)).fetchone()
        return jsonify({'status': 'success', 'message': 'Inventory item updated.', 'item': serialize_inventory_item(updated_row)})
    finally:
        conn.close()

@app.route("/api/inventory/items/<int:item_id>", methods=['DELETE'])
@login_required
def api_inventory_delete_item(item_id):
    guard = ensure_inventory_permission()
    if guard:
        return guard
    conn = get_connection()
    try:
        conn.execute('DELETE FROM inventory_items WHERE item_id = ?', (item_id,))
        conn.commit()
        return jsonify({'status': 'success', 'message': 'Inventory item deleted.'})
    finally:
        conn.close()

@app.route("/api/inventory/items/<int:item_id>/issue", methods=['POST'])
@login_required
def api_inventory_issue_item(item_id):
    guard = ensure_inventory_permission()
    if guard:
        return guard
    data = request.get_json() or {}
    issued_to = (data.get('issued_to') or '').strip()
    issued_type = (data.get('issued_type') or '').strip().title()
    quantity = int(data.get('quantity') or 0)
    if not issued_to or issued_type not in INVENTORY_ISSUE_TYPES:
        return jsonify({'status': 'error', 'message': 'Issued to and issued type are required.'}), 400
    if quantity <= 0:
        return jsonify({'status': 'error', 'message': 'Quantity must be greater than zero.'}), 400

    conn = get_connection()
    try:
        item_row = conn.execute('SELECT * FROM inventory_items WHERE item_id = ?', (item_id,)).fetchone()
        if not item_row:
            return jsonify({'status': 'error', 'message': 'Item not found'}), 404
        if (item_row['status'] or '') in INVENTORY_INACTIVE_STATUSES:
            return jsonify({'status': 'error', 'message': 'Item is not available for issuing.'}), 400
        if (item_row['quantity_available'] or 0) < quantity:
            return jsonify({'status': 'error', 'message': 'Insufficient stock!'}), 400

        now = datetime.now().isoformat()
        conn.execute(
            '''
                INSERT INTO item_issue_records (item_id, issued_to, issued_type, issue_date, quantity, quantity_returned,
                                                status, remarks, receiver_signature, issuer_signature, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, 0, 'Issued', ?, ?, ?, ?, ?)
            ''',
            (
                item_id,
                issued_to,
                issued_type,
                data.get('issue_date') or now,
                quantity,
                (data.get('remarks') or '').strip(),
                data.get('receiver_signature'),
                data.get('issuer_signature'),
                now,
                now
            )
        )
        conn.execute(
            'UPDATE inventory_items SET quantity_available = quantity_available - ?, updated_at = ? WHERE item_id = ?',
            (quantity, now, item_id)
        )
        updated = conn.execute('SELECT * FROM inventory_items WHERE item_id = ?', (item_id,)).fetchone()
        maybe_create_auto_request(conn, serialize_inventory_item(updated))
        conn.commit()
        return jsonify({'status': 'success', 'message': 'Item issued successfully.', 'item': serialize_inventory_item(updated)})
    finally:
        conn.close()

@app.route("/api/inventory/issues", methods=['GET'])
@login_required
def api_inventory_issue_list():
    guard = ensure_inventory_permission(INVENTORY_REPORT_ROLES)
    if guard:
        return guard
    status = (request.args.get('status') or '').strip()
    issued_type = (request.args.get('issued_type') or '').strip().title()
    conn = get_connection()
    try:
        query = '''
            SELECT r.*, i.item_name, i.category, i.location
            FROM item_issue_records r
            LEFT JOIN inventory_items i ON r.item_id = i.item_id
            WHERE 1=1
        '''
        params = []
        if status:
            query += ' AND LOWER(r.status) = ?'
            params.append(status.lower())
        if issued_type in INVENTORY_ISSUE_TYPES:
            query += ' AND r.issued_type = ?'
            params.append(issued_type)
        query += ' ORDER BY r.issue_date DESC LIMIT 200'
        rows = conn.execute(query, params).fetchall()
        issues = [serialize_issue_row(row) for row in rows]
        return jsonify({'status': 'success', 'issues': issues})
    finally:
        conn.close()

@app.route("/api/inventory/issues/<int:issue_id>/return", methods=['POST'])
@login_required
def api_inventory_return_issue(issue_id):
    guard = ensure_inventory_permission()
    if guard:
        return guard
    data = request.get_json() or {}
    quantity = int(data.get('quantity') or 0)
    if quantity <= 0:
        return jsonify({'status': 'error', 'message': 'Returned quantity must be greater than zero.'}), 400

    conn = get_connection()
    try:
        issue_row = conn.execute('SELECT * FROM item_issue_records WHERE issue_id = ?', (issue_id,)).fetchone()
        if not issue_row:
            return jsonify({'status': 'error', 'message': 'Issue record not found'}), 404
        remaining = (issue_row['quantity'] or 0) - (issue_row['quantity_returned'] or 0)
        if remaining <= 0:
            return jsonify({'status': 'error', 'message': 'Item already fully returned.'}), 400
        if quantity > remaining:
            return jsonify({'status': 'error', 'message': 'Returned quantity exceeds outstanding quantity.'}), 400

        now = datetime.now().isoformat()
        new_status = 'Returned' if quantity == remaining else 'Partially Returned'
        conn.execute(
            '''
                UPDATE item_issue_records
                SET quantity_returned = quantity_returned + ?, status = ?, return_date = ?, remarks = COALESCE(remarks, '')
                    || CASE WHEN ? != '' THEN ' | Return: ' || ? ELSE '' END,
                    updated_at = ?
                WHERE issue_id = ?
            ''',
            (
                quantity,
                new_status,
                now,
                (data.get('remarks') or '').strip(),
                (data.get('remarks') or '').strip(),
                now,
                issue_id
            )
        )
        conn.execute(
            'UPDATE inventory_items SET quantity_available = quantity_available + ?, updated_at = ? WHERE item_id = ?',
            (quantity, now, issue_row['item_id'])
        )
        conn.commit()
        return jsonify({'status': 'success', 'message': 'Item return recorded.'})
    finally:
        conn.close()

@app.route("/api/inventory/requests", methods=['GET'])
@login_required
def api_inventory_requests():
    guard = ensure_inventory_permission(INVENTORY_REPORT_ROLES)
    if guard:
        return guard
    status = (request.args.get('status') or '').strip()
    conn = get_connection()
    try:
        query = 'SELECT * FROM inventory_requests WHERE 1=1'
        params = []
        if status:
            query += ' AND LOWER(status) = ?'
            params.append(status.lower())
        query += ' ORDER BY request_date DESC LIMIT 200'
        rows = conn.execute(query, params).fetchall()
        requests_list = [dict(row) for row in rows]
        return jsonify({'status': 'success', 'requests': requests_list})
    finally:
        conn.close()

@app.route("/api/inventory/requests", methods=['POST'])
@login_required
def api_inventory_create_request():
    guard = ensure_inventory_permission(INVENTORY_REPORT_ROLES)
    if guard:
        return guard
    data = request.get_json() or {}
    requested_by = (data.get('requested_by') or session.get('username') or 'Unknown').strip()
    item_name = (data.get('item_name') or '').strip()
    quantity = int(data.get('quantity') or 0)
    if not item_name or quantity <= 0:
        return jsonify({'status': 'error', 'message': 'Item name and quantity are required.'}), 400
    now = datetime.now().isoformat()
    conn = get_connection()
    try:
        conn.execute(
            '''
                INSERT INTO inventory_requests (requested_by, requested_role, item_name, quantity, status, request_date, remarks)
                VALUES (?, ?, ?, ?, 'Pending', ?, ?)
            ''',
            (
                requested_by,
                session.get('role'),
                item_name,
                quantity,
                now,
                (data.get('remarks') or '').strip()
            )
        )
        conn.commit()
        return jsonify({'status': 'success', 'message': 'Request submitted.'})
    finally:
        conn.close()

@app.route("/api/inventory/requests/<int:request_id>", methods=['PUT'])
@login_required
def api_inventory_update_request(request_id):
    guard = ensure_inventory_permission()
    if guard:
        return guard
    data = request.get_json() or {}
    status = (data.get('status') or '').strip().title()
    if status not in ('Pending', 'Approved', 'Rejected', 'Fulfilled'):
        return jsonify({'status': 'error', 'message': 'Invalid status.'}), 400
    conn = get_connection()
    try:
        request_row = conn.execute('SELECT * FROM inventory_requests WHERE request_id = ?', (request_id,)).fetchone()
        if not request_row:
            return jsonify({'status': 'error', 'message': 'Request not found'}), 404
        now = datetime.now().isoformat()
        conn.execute(
            '''
                UPDATE inventory_requests
                SET status = ?, approval_date = CASE WHEN ? IN ('Approved','Rejected') THEN ? ELSE approval_date END,
                    fulfilled_date = CASE WHEN ? = 'Fulfilled' THEN ? ELSE fulfilled_date END,
                    remarks = COALESCE(?, remarks)
                WHERE request_id = ?
            ''',
            (
                status,
                status,
                now,
                status,
                now,
                data.get('remarks'),
                request_id
            )
        )
        conn.commit()
        return jsonify({'status': 'success', 'message': 'Request updated.'})
    finally:
        conn.close()

@app.route("/inventory/report/<string:report_type>")
@login_required
@module_required('inventory')
def inventory_report(report_type):
    guard = ensure_inventory_permission(INVENTORY_REPORT_ROLES)
    if guard:
        return guard
    report_type = report_type.lower()
    valid_reports = {
        'stock': 'Stock Register Report',
        'category': 'Category-wise Report',
        'issued': 'Issued Items Report',
        'low-stock': 'Low Stock Report',
        'expired': 'Expired/Expiring Soon Report'
    }
    if report_type not in valid_reports:
        abort(404)

    conn = get_connection()
    try:
        rows = []
        if report_type == 'stock':
            rows = conn.execute('SELECT * FROM inventory_items ORDER BY category, item_name').fetchall()
        elif report_type == 'category':
            category = (request.args.get('category') or '').strip()
            if not category:
                flash('Category parameter is required for this report.', 'warning')
                return redirect(url_for('dashboard'))
            rows = conn.execute('SELECT * FROM inventory_items WHERE LOWER(category) = ? ORDER BY item_name', (category.lower(),)).fetchall()
        elif report_type == 'issued':
            rows = conn.execute('''
                SELECT r.*, i.item_name, i.category, i.location
                FROM item_issue_records r
                LEFT JOIN inventory_items i ON r.item_id = i.item_id
                ORDER BY r.issue_date DESC
            ''').fetchall()
        elif report_type == 'low-stock':
            rows = conn.execute(
                'SELECT * FROM inventory_items WHERE quantity_available < ? ORDER BY quantity_available ASC',
                (INVENTORY_LOW_STOCK_THRESHOLD,)
            ).fetchall()
        elif report_type == 'expired':
            rows = conn.execute(
                '''
                    SELECT * FROM inventory_items
                    WHERE (expiry_date IS NOT NULL AND expiry_date != '')
                      AND (DATE(expiry_date) < DATE('now') OR DATE(expiry_date) <= DATE('now', ?))
                    ORDER BY expiry_date ASC
                ''',
                (f'+{INVENTORY_EXPIRY_WINDOW_DAYS} days',)
            ).fetchall()

        context = {
            'report_title': valid_reports[report_type],
            'report_type': report_type,
            'generated_at': datetime.now().strftime('%d %B %Y'),
            'rows': rows,
            'serialize_inventory_item': serialize_inventory_item,
            'serialize_issue_row': serialize_issue_row
        }
        return render_template('inventory_report.html', **context)
    finally:
        conn.close()

@app.route("/teacher/dashboard")
@login_required
def teacher_dashboard():
    """Teacher dashboard page"""
    if session.get('role') != 'teacher':
        return redirect(url_for('index'))
    return render_template('teacher_dashboard.html')

@app.route("/api/teacher/dashboard", methods=['GET'])
@login_required
def api_teacher_dashboard():
    """Get teacher dashboard data"""
    try:
        teacher_id = session.get('teacher_id')
        if not teacher_id:
            return jsonify({'status': 'error', 'message': 'Not logged in'}), 401
        
        conn = get_connection()
        
        # Get teacher info
        teacher = conn.execute('SELECT * FROM teachers WHERE id = ?', (teacher_id,)).fetchone()
        if not teacher:
            return jsonify({'status': 'error', 'message': 'Teacher not found'}), 404
        
        teacher_dict = dict(teacher)
        assigned_semesters = parse_multi_value(teacher_dict.get('assigned_semesters'))
        technology_assignments = parse_multi_value(teacher_dict.get('technology_assignments') or teacher_dict.get('technology'))
        subject = teacher_dict.get('subject', '')
        technology = join_display(technology_assignments)
        
        # Get permissions
        permissions = conn.execute(
            'SELECT permission_name FROM teacher_permissions WHERE teacher_id = ? AND granted = 1',
            (teacher_id,)
        ).fetchall()
        permission_list = [p['permission_name'] for p in permissions]
        
        # Get assigned subjects (can be multiple if teacher has multiple subjects)
        assigned_subjects = []
        if subject:
            # Get students for each subject/semester combination
            for sem in assigned_semesters:
                student_count = conn.execute(
                    'SELECT COUNT(*) as count FROM students WHERE status = "Active" AND semester = ? AND technology = ?',
                    (sem, technology)
                ).fetchone()['count']
                assigned_subjects.append({
                    'subject': subject,
                    'technology': technology,
                    'semester': sem,
                    'student_count': student_count
                })
        else:
            # If no subject assigned, show all semesters
            for sem in assigned_semesters:
                student_count = conn.execute(
                    'SELECT COUNT(*) as count FROM students WHERE status = "Active" AND semester = ?',
                    (sem,)
                ).fetchone()['count']
                assigned_subjects.append({
                    'subject': 'All Subjects',
                    'technology': technology or 'All',
                    'semester': sem,
                    'student_count': student_count
                })
        
        # Get total students for this teacher
        student_query = 'SELECT COUNT(*) as count FROM students WHERE status = "Active"'
        student_params = []
        if assigned_semesters:
            semester_placeholders = ','.join(['?'] * len(assigned_semesters))
            student_query += f' AND semester IN ({semester_placeholders})'
            student_params.extend(assigned_semesters)
        if technology:
            student_query += ' AND technology = ?'
            student_params.append(technology)
        
        total_students = conn.execute(student_query, student_params).fetchone()['count']
        
        # Get upcoming exams
        exam_query = '''
            SELECT exam_id, title, subject, exam_date, start_time, status
            FROM midterm_exams
            WHERE status IN ('Scheduled', 'Active')
        '''
        exam_params = []
        if subject:
            exam_query += ' AND subject = ?'
            exam_params.append(subject)
        if technology:
            exam_query += ' AND technology = ?'
            exam_params.append(technology)
        if assigned_semesters:
            semester_placeholders = ','.join(['?'] * len(assigned_semesters))
            exam_query += f' AND semester IN ({semester_placeholders})'
            exam_params.extend(assigned_semesters)
        
        exam_query += ' ORDER BY exam_date ASC LIMIT 5'
        upcoming_exams = conn.execute(exam_query, exam_params).fetchall()
        
        # Get attendance summary by subject
        current_month = datetime.now().strftime('%Y-%m')
        attendance_summary = {}
        
        if assigned_semesters:
            # Get student IDs for assigned semesters
            student_ids_query = 'SELECT id FROM students WHERE status = "Active"'
            student_ids_params = []
            if assigned_semesters:
                semester_placeholders = ','.join(['?'] * len(assigned_semesters))
                student_ids_query += f' AND semester IN ({semester_placeholders})'
                student_ids_params.extend(assigned_semesters)
            if technology:
                student_ids_query += ' AND technology = ?'
                student_ids_params.append(technology)
            
            student_ids = [row['id'] for row in conn.execute(student_ids_query, student_ids_params).fetchall()]
            if student_ids:
                student_id_placeholders = ','.join(['?'] * len(student_ids))
                attendance_query = f'''
                    SELECT COUNT(DISTINCT student_id) as total,
                           SUM(CASE WHEN status = 'Present' THEN 1 ELSE 0 END) as present,
                           SUM(CASE WHEN status = 'Absent' THEN 1 ELSE 0 END) as absent
                    FROM attendance
                    WHERE strftime('%Y-%m', attendance_date) = ?
                    AND student_id IN ({student_id_placeholders})
                '''
                attendance_params = [current_month] + student_ids
                summary = conn.execute(attendance_query, attendance_params).fetchone()
                if summary:
                    attendance_summary[subject or 'All Subjects'] = {
                        'total': summary['total'] or 0,
                        'present': summary['present'] or 0,
                        'absent': summary['absent'] or 0
                    }
        
        # Calculate average attendance percentage
        attendance_percentage = 0
        if attendance_summary:
            total_present = sum(s['present'] for s in attendance_summary.values())
            total_records = sum(s['total'] for s in attendance_summary.values())
            if total_records > 0:
                attendance_percentage = round((total_present / total_records) * 100, 1)
        
        conn.close()
        
        return jsonify({
            'teacher_name': teacher_dict['name'],
            'assigned_subjects': assigned_subjects,
            'total_students': total_students,
            'upcoming_exams': [{
                'name': e.get('title', 'N/A'),
                'subject': e.get('subject', 'N/A'),
                'date': e.get('exam_date', '')
            } for e in upcoming_exams],
            'attendance_summary': attendance_summary,
            'attendance_percentage': attendance_percentage,
            'permissions': permission_list
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

UPLOAD_FOLDER = os.path.join(app.root_path, 'uploads')
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

@app.route("/api/campuses")
def get_campuses():
    conn = db.get_connection()
    campuses = conn.execute('SELECT name FROM campuses').fetchall()
    conn.close()
    # Filter out "Satellite Campus" from the list
    return jsonify([row['name'] for row in campuses if row['name'] != 'Satellite Campus'])

@app.route("/api/boards")
def get_boards():
    conn = db.get_connection()
    boards = conn.execute('SELECT name FROM boards').fetchall()
    conn.close()
    return jsonify([row['name'] for row in boards])

@app.route("/api/semesters")
def get_semesters():
    conn = db.get_connection()
    db_semesters = conn.execute('SELECT name FROM semesters').fetchall()
    conn.close()
    
    # Extract names from database results
    semesters = [row['name'] for row in db_semesters]
    
    # Add Sept-2023, Sept-2024, Sept-2025 to all semester lists (except Add/Edit Student form)
    additional_semesters = ['Sept-2023', 'Sept-2024', 'Sept-2025']
    for sem in additional_semesters:
        if sem not in semesters:
            semesters.append(sem)
    
    # Combine and ensure uniqueness, maintaining order if possible (or just append)
    # Using a set to maintain uniqueness, then converting back to list
    all_semesters = list(set(semesters))
    all_semesters.sort() # Optional: sort them alphabetically or numerically if they follow a pattern
    
    return jsonify(all_semesters)

@app.route("/api/technologies")
def get_technologies():
    campus = request.args.get('campus', '')
    board = request.args.get('board', '')
    
    conn = db.get_connection()
    
    # Build query - if campus and board are provided, filter technologies by students with those values
    if campus and campus != 'All' and board and board != 'All':
        query = '''
            SELECT DISTINCT s.technology as name 
            FROM students s 
            WHERE s.campus = ? AND s.board = ? AND s.technology IS NOT NULL AND s.technology != ''
        '''
        technologies = conn.execute(query, (campus, board)).fetchall()
    else:
        # Get all technologies from the technologies table
        technologies = conn.execute('SELECT name FROM technologies').fetchall()
    
    conn.close()
    
    # Filter out specific technologies: Pharmacy Technician, Dental Technician (case-insensitive)
    # Also filter out: Pharmacy technician, Dental technician, LHV, CMA
    excluded_technologies = [
        'Pharmacy Technician', 'Pharmacy technician', 'pharmacy technician', 'PHARMACY TECHNICIAN',
        'Dental Technician', 'Dental technician', 'dental technician', 'DENTAL TECHNICIAN',
        'LHV', 'CMA'
    ]
    # Case-insensitive filtering
    excluded_lower = [tech.lower() for tech in excluded_technologies]
    filtered_technologies = [
        row['name'] for row in technologies 
        if row['name'] and row['name'].lower() not in excluded_lower
    ]
    # Remove duplicates and sort
    filtered_technologies = sorted(list(set(filtered_technologies)))
    return jsonify(filtered_technologies)

@app.route("/api/students", methods=['GET'])
def get_students():
    search_query = request.args.get('search')
    admission_no_filter = request.args.get('admission_no')
    name_filter = request.args.get('name')
    father_name_filter = request.args.get('father_name')
    status_filter = request.args.get('status')
    campus_filter = request.args.get('campus')
    board_filter = request.args.get('board')
    semester_filter = request.args.get('semester')
    technology_filter = request.args.get('technology')
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)

    print(f"Received search query: {search_query}, status filter: {status_filter}, page: {page}")
    print(f"Campus: {campus_filter}, Board: {board_filter}, Semester: {semester_filter}, Technology: {technology_filter}")
    conn = db.get_connection()
    
    base_query = 'FROM students'
    count_query = 'SELECT COUNT(*) ' + base_query
    data_query = 'SELECT id, name, father_name, address, phone, sms_phone, admission_no, campus, board, semester, technology, gender, status, photo_path, student_type ' + base_query
    
    params = []
    conditions = []

    if search_query:
        conditions.append(' (admission_no LIKE ? OR name LIKE ? OR father_name LIKE ?) ')
        params.extend([f'%{search_query}%', f'%{search_query}%', f'%{search_query}%'])

    if admission_no_filter:
        conditions.append(' admission_no LIKE ? ')
        params.append(f'%{admission_no_filter}%')

    if name_filter:
        conditions.append(' name LIKE ? ')
        params.append(f'%{name_filter}%')

    if father_name_filter:
        conditions.append(' father_name LIKE ? ')
        params.append(f'%{father_name_filter}%')
    
    if status_filter:
        conditions.append(' status = ? ')
        params.append(status_filter)
    
    if campus_filter:
        conditions.append(' campus = ? ')
        params.append(campus_filter)
    
    if board_filter:
        conditions.append(' board = ? ')
        params.append(board_filter)
    
    if semester_filter:
        conditions.append(' semester = ? ')
        params.append(semester_filter)
    
    if technology_filter:
        conditions.append(' technology = ? ')
        params.append(technology_filter)

    if conditions:
        where_clause = ' WHERE ' + ' AND '.join(conditions)
        count_query += where_clause
        data_query += where_clause

    # Get total count
    total_students = conn.execute(count_query, params).fetchone()[0]
    
    # Get paginated data
    offset = (page - 1) * per_page
    data_query += f' LIMIT {per_page} OFFSET {offset}'
    
    students = conn.execute(data_query, params).fetchall()
    conn.close()
    
    return jsonify({
        'students': [dict(row) for row in students],
        'total': total_students,
        'page': page,
        'per_page': per_page
    })

@app.route("/import_excel_web", methods=['GET', 'POST'])
def import_excel_web():
    if request.method == 'POST':
        if 'excel_file' not in request.files:
            flash('No file part', 'danger')
            return redirect(request.url)
        file = request.files['excel_file']
        if file.filename == '':
            flash('No selected file', 'danger')
            return redirect(request.url)
        if file and (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
            try:
                # Read the Excel file into a pandas DataFrame
                df = pd.read_excel(file)
                
                # Normalize column names (case-insensitive, strip whitespace)
                df.columns = [col.strip().lower() if isinstance(col, str) else col for col in df.columns]
                
                # Process each row
                imported_count = 0
                updated_count = 0
                skipped_count = 0
                error_messages = []

                conn = get_connection()
                cursor = conn.cursor()

                for index, row in df.iterrows():
                    row_data = row.to_dict()
                    
                    # Get admission_no with flexible column name matching
                    admission_no = None
                    for key in row_data.keys():
                        if 'admission' in key and 'no' in key:
                            admission_no = str(row_data.get(key, '')).strip()
                            break
                    
                    if not admission_no:
                        skipped_count += 1
                        error_messages.append(f"Row {index + 2}: Skipped due to missing 'Admission No' column.")
                        continue

                    # Helper function to safely get value from row_data with flexible column matching
                    def get_value(row_dict, *possible_keys):
                        for key in possible_keys:
                            # Try exact match first
                            if key in row_dict:
                                val = row_dict.get(key, '')
                                if pd.isna(val):
                                    return ''
                                return str(val).strip()

                        # Try case-insensitive match
                        for key in possible_keys:
                            key_lower = str(key).lower().strip()
                            for dict_key in row_dict.keys():
                                if str(dict_key).lower().strip() == key_lower:
                                    val = row_dict.get(dict_key, '')
                                    if pd.isna(val):
                                        return ''
                                    return str(val).strip()
                        return ''

                    # Map Excel headers to database column names with flexible matching
                    try:
                        dob_val = get_value(row_data, 'dob', 'date of birth', 'birth date')
                        if dob_val:
                            try:
                                dob_formatted = pd.to_datetime(dob_val).strftime('%Y-%m-%d')
                            except:
                                dob_formatted = ''
                        else:
                            dob_formatted = ''

                        student_data = {
                            'admission_no': admission_no,
                            'name': get_value(row_data, 'name', 'student name'),
                            'father_name': get_value(row_data, "father's name", 'father name', 'fathers name'),
                            'address': get_value(row_data, 'address'),
                            'dob': dob_formatted,
                            'gender': get_value(row_data, 'gender'),
                            'nationality': get_value(row_data, 'nationality'),
                            'district': get_value(row_data, 'district'),
                            'phone': get_value(row_data, 'phone', 'phone #', 'phone number'),
                            'sms_phone': get_value(row_data, 'sms phone', 'sms_phone', 'sms number', 'sms #'),
                            'campus': get_value(row_data, 'campus'),
                            'board': get_value(row_data, 'board'),
                            'technology': get_value(row_data, 'technology', 'technology/program', 'program', 'course'),
                            'semester': get_value(row_data, 'semester', 'semester/year', 'year', 'session'),
                            'status': get_value(row_data, 'status', 'student status') or 'Active',
                            'photo_path': None,  # Not imported via web
                            'student_type': get_value(row_data, 'student type', 'student_type', 'type') or 'Paid',
                            'remarks': get_value(row_data, 'remarks', 'remarks & notes', 'notes', 'comments'),
                            'created_at': datetime.now().strftime('%Y-%m-%d')
                        }

                        # Check if student exists
                        cursor.execute("SELECT id FROM students WHERE admission_no=?", (admission_no,))
                        exists = cursor.fetchone()

                        if exists:
                            # Update existing record
                            update_data = {k: v for k, v in student_data.items() if k not in ['admission_no', 'created_at']}
                            
                            update_fields = [f"{k}=:{k}" for k in update_data.keys()]
                            query = f"UPDATE students SET {', '.join(update_fields)} WHERE admission_no=:admission_no"
                            cursor.execute(query, {**update_data, 'admission_no': admission_no})
                            updated_count += 1
                        else:
                            # Insert new record
                            fields = ', '.join(student_data.keys())
                            placeholders = ', '.join([f":{k}" for k in student_data.keys()])
                            query = f"INSERT INTO students ({fields}) VALUES ({placeholders})"
                            cursor.execute(query, student_data)
                            imported_count += 1
                        conn.commit()
                    except sqlite3.IntegrityError as e:
                        conn.rollback()
                        error_messages.append(f"Row {index + 2} (Admission No: {admission_no}): Integrity error - {str(e)}")
                        skipped_count += 1
                    except sqlite3.Error as e:
                        conn.rollback()
                        error_messages.append(f"Row {index + 2} (Admission No: {admission_no}): Database error - {str(e)}")
                        skipped_count += 1
                    except Exception as e:
                        conn.rollback()
                        error_messages.append(f"Row {index + 2} (Admission No: {admission_no}): Error - {str(e)}")
                        skipped_count += 1
                
                conn.close()

                summary = (
                    f"✅ Excel Import Complete!\n\n"
                    f"📊 Results:\n"
                    f"  • Successfully Imported: {imported_count}\n"
                    f"  • Successfully Updated: {updated_count}\n"
                    f"  • Skipped: {skipped_count}\n"
                    f"  • Total Processed: {imported_count + updated_count + skipped_count}\n"
                )
                if error_messages:
                    summary += f"\n⚠️ Errors encountered ({len(error_messages)}):\n" + "\n".join(error_messages[:10])  # Show first 10 errors
                    if len(error_messages) > 10:
                        summary += f"\n... and {len(error_messages) - 10} more errors"
                    flash(summary, 'warning')
                else:
                    flash(summary, 'success')

                return redirect(url_for('import_excel_web'))

            except Exception as e:
                flash(f"❌ Error processing Excel file: {str(e)}", 'danger')
                return redirect(request.url)
        else:
            flash('❌ Invalid file type. Please upload an Excel file (.xlsx or .xls).', 'danger')
            return redirect(request.url)
    return render_template('import_excel.html')

@app.route("/import_biodata_excel", methods=['GET', 'POST'])
def import_biodata_excel():
    if request.method == 'POST':
        if 'biodata_excel_file' not in request.files:
            flash('No file part for biodata', 'danger')
            return redirect(request.url)
        file = request.files['biodata_excel_file']
        if file.filename == '':
            flash('No selected biodata file', 'danger')
            return redirect(request.url)
        if file and (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
            try:
                df = pd.read_excel(file)
                df.columns = [col.strip().lower() if isinstance(col, str) else col for col in df.columns]

                imported_count = 0
                updated_count = 0
                skipped_count = 0
                error_messages = []

                conn = get_connection()
                cursor = conn.cursor()

                for index, row in df.iterrows():
                    row_data = row.to_dict()
                    
                    admission_no = None
                    for key in row_data.keys():
                        if 'admission' in key and 'no' in key:
                            admission_no = str(row_data.get(key, '')).strip()
                            break
                    
                    if not admission_no:
                        skipped_count += 1
                        error_messages.append(f"Row {index + 2}: Skipped due to missing 'Admission No' for biodata.")
                        continue

                    def get_value(row_dict, *possible_keys):
                        for key in possible_keys:
                            if key in row_dict:
                                val = row_dict.get(key, '')
                                if pd.isna(val):
                                    return ''
                                return str(val).strip()

                        for key in possible_keys:
                            key_lower = str(key).lower().strip()
                            for dict_key in row_dict.keys():
                                if str(dict_key).lower().strip() == key_lower:
                                    val = row_dict.get(dict_key, '')
                                    if pd.isna(val):
                                        return ''
                                    return str(val).strip()
                        return ''

                    try:
                        dob_val = get_value(row_data, 'dob', 'date of birth', 'birth date')
                        if dob_val:
                            try:
                                dob_formatted = pd.to_datetime(dob_val).strftime('%Y-%m-%d')
                            except:
                                dob_formatted = ''
                        else:
                            dob_formatted = ''

                        biodata_data = {
                            'admission_no': admission_no,
                            'name': get_value(row_data, 'name', 'student name'),
                            'father_name': get_value(row_data, "father's name", 'father name', 'fathers name'),
                            'address': get_value(row_data, 'address'),
                            'dob': dob_formatted,
                            'gender': get_value(row_data, 'gender'),
                            'nationality': get_value(row_data, 'nationality'),
                            'district': get_value(row_data, 'district'),
                            'phone': get_value(row_data, 'phone', 'phone #', 'phone number'),
                            'sms_phone': get_value(row_data, 'sms phone', 'sms_phone', 'sms number', 'sms #'),
                            'campus': get_value(row_data, 'campus'),
                            'board': get_value(row_data, 'board'),
                            'technology': get_value(row_data, 'technology', 'technology/program', 'program', 'course'),
                            'semester': get_value(row_data, 'semester', 'semester/year', 'year', 'session'),
                            'status': get_value(row_data, 'status', 'student status') or 'Active',
                            'student_type': get_value(row_data, 'student type', 'student_type', 'type') or 'Paid',
                            'remarks': get_value(row_data, 'remarks', 'remarks & notes', 'notes', 'comments'),
                        }

                        cursor.execute("SELECT id FROM students WHERE admission_no=?", (admission_no,))
                        exists = cursor.fetchone()

                        if exists:
                            update_data = {k: v for k, v in biodata_data.items() if k not in ['admission_no']}
                            update_fields = [f"{k}=:{k}" for k in update_data.keys()]
                            query = f"UPDATE students SET {', '.join(update_fields)} WHERE admission_no=:admission_no"
                            cursor.execute(query, {**update_data, 'admission_no': admission_no})
                            updated_count += 1
                        else:
                            # If student doesn't exist, insert as a new student
                            # Add created_at for new entries
                            biodata_data['created_at'] = datetime.now().strftime('%Y-%m-%d')
                            fields = ', '.join(biodata_data.keys())
                            placeholders = ', '.join([f":{k}" for k in biodata_data.keys()])
                            query = f"INSERT INTO students ({fields}) VALUES ({placeholders})"
                            cursor.execute(query, biodata_data)
                            imported_count += 1
                        conn.commit()
                    except sqlite3.IntegrityError as e:
                        conn.rollback()
                        error_messages.append(f"Row {index + 2} (Admission No: {admission_no}): Integrity error - {str(e)}")
                        skipped_count += 1
                    except sqlite3.Error as e:
                        conn.rollback()
                        error_messages.append(f"Row {index + 2} (Admission No: {admission_no}): Database error - {str(e)}")
                        skipped_count += 1
                    except Exception as e:
                        conn.rollback()
                        error_messages.append(f"Row {index + 2} (Admission No: {admission_no}): Error - {str(e)}")
                        skipped_count += 1
                
                conn.close()

                summary = (
                    f"✅ Biodata Excel Import Complete!\n\n"
                    f"📊 Results:\n"
                    f"  • Successfully Imported: {imported_count}\n"
                    f"  • Successfully Updated: {updated_count}\n"
                    f"  • Skipped: {skipped_count}\n"
                    f"  • Total Processed: {imported_count + updated_count + skipped_count}\n"
                )
                if error_messages:
                    summary += f"\n⚠️ Errors encountered ({len(error_messages)}):\n" + "\n".join(error_messages[:10])
                    if len(error_messages) > 10:
                        summary += f"\n... and {len(error_messages) - 10} more errors"
                    flash(summary, 'warning')
                else:
                    flash(summary, 'success')

                return redirect(url_for('import_excel_web')) # Redirect back to the import page

            except Exception as e:
                flash(f"❌ Error processing Biodata Excel file: {str(e)}", 'danger')
                return redirect(request.url)
        else:
            flash('❌ Invalid file type for biodata. Please upload an Excel file (.xlsx or .xls).', 'danger')
            return redirect(request.url)
    return render_template('import_excel.html')


@app.route("/api/students", methods=['POST'])
def add_student():
    try:
        data = request.form
        picture = request.files.get('admission_picture')
        picture_path = None

        if picture and picture.filename:
            filename = f"{data.get('admission_no')}_{picture.filename}"
            full_picture_path = os.path.join(UPLOAD_FOLDER, filename)
            picture.save(full_picture_path)
            # Store relative path from app root in DB for URL access
            picture_path = os.path.join('uploads', filename).replace('\\', '/')

        conn = db.get_connection()
        
        # Check for duplicate admission number
        existing_student = conn.execute('SELECT * FROM students WHERE admission_no = ?', (data.get('admission_no'),)).fetchone()
        if existing_student:
            return jsonify({'status': 'error', 'message': 'Student with this admission number already exists.'}), 400

        # Insert student
        conn.execute(
            'INSERT INTO students (name, father_name, address, dob, gender, nationality, district, phone, sms_phone, admission_no, campus, board, semester, technology, status, photo_path, student_type, remarks, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
            (
                data.get('name'),
                data.get('father_name'),
                data.get('address'),
                data.get('dob'),
                data.get('gender'),
                data.get('nationality'),
                data.get('district'),
                data.get('phone'),
                data.get('sms_phone'),  
                data.get('admission_no'),
                data.get('campus'),
                data.get('board'),
                data.get('semester'),
                data.get('technology'),
                data.get('status', 'Active'),
                picture_path, # Storing the path
                data.get('student_type', 'Paid'),
                data.get('remarks'),
                datetime.now().isoformat()
            )
        )
        
        # Auto-create student account if status is Active
        student_status = data.get('status', 'Active')
        if student_status == 'Active':
            student_id = conn.lastrowid
            admission_no = data.get('admission_no')
            technology = data.get('technology', 'student123')
            
            # Generate username (admission number) and password (technology name)
            username = admission_no
            password_hash = generate_password_hash(technology)
            
            try:
                conn.execute('''
                    UPDATE students 
                    SET username = ?, password_hash = ?, account_status = 'Active'
                    WHERE id = ?
                ''', (username, password_hash, student_id))
                
                # Log activity
                conn.execute(
                    'INSERT INTO student_activity_log (student_id, activity_type, activity_description, ip_address, created_at) VALUES (?, ?, ?, ?, ?)',
                    (student_id, 'account_created', f'Auto-created account with username: {username}', request.remote_addr, datetime.now().isoformat())
                )
            except Exception as e:
                print(f"Error auto-creating student account: {e}")
                # Don't fail the student creation if account creation fails

        conn.commit()
        conn.close()
        return jsonify({'status': 'success'})
    except Exception as e:
        print(f"Error adding student: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route("/api/students/<int:student_id>", methods=['GET'])
def get_student_by_id(student_id):
    """Fetch student data by ID for certificate generation"""
    try:
        conn = db.get_connection()
        student = conn.execute(
            'SELECT id, admission_no, name, father_name, technology, semester, campus, board, status, gender FROM students WHERE id = ?',
            (student_id,)
        ).fetchone()
        conn.close()

        if not student:
            return jsonify({'status': 'error', 'message': 'Student not found'}), 404

        return jsonify(dict(student))
    except Exception as e:
        print(f"Error fetching student: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route("/api/students/<int:student_id>", methods=['POST'])
def update_student(student_id):
    try:
        data = request.form
        picture = request.files.get('admission_picture')
        
        conn = db.get_connection()
        
        # Get existing photo_path
        existing_student = conn.execute('SELECT photo_path FROM students WHERE id = ?', (student_id,)).fetchone()
        current_photo_path = existing_student['photo_path'] if existing_student else None
        
        picture_path_to_save = current_photo_path

        if picture and picture.filename:
            # Save new picture
            filename = f"{data.get('admission_no')}_{picture.filename}"
            new_picture_full_path = os.path.join(UPLOAD_FOLDER, filename)
            picture.save(new_picture_full_path)
            # Store relative path from app root in DB for URL access
            picture_path_to_save = os.path.join('uploads', filename).replace('\\', '/')
            
            # Optionally, delete old picture if it exists and is different
            if current_photo_path and current_photo_path != picture_path_to_save:
                # Construct absolute path for deletion
                absolute_current_photo_path = os.path.join(app.root_path, current_photo_path)
                if os.path.exists(absolute_current_photo_path):
                    os.remove(absolute_current_photo_path)

        # Get current status before update
        current_student = conn.execute('SELECT status FROM students WHERE id = ?', (student_id,)).fetchone()
        old_status = dict(current_student)['status'] if current_student else None
        new_status = data.get('status')
        
        conn.execute(
            'UPDATE students SET name = ?, father_name = ?, address = ?, dob = ?, gender = ?, nationality = ?, district = ?, phone = ?, sms_phone = ?, admission_no = ?, campus = ?, board = ?, semester = ?, technology = ?, status = ?, photo_path = ?, student_type = ?, remarks = ? WHERE id = ?',
            (
                data.get('name'),
                data.get('father_name'),
                data.get('address'),
                data.get('dob'),
                data.get('gender'),
                data.get('nationality'),
                data.get('district'),
                data.get('phone'),
                data.get('sms_phone'),
                data.get('admission_no'),
                data.get('campus'),
                data.get('board'),
                data.get('semester'),
                data.get('technology'),
                data.get('status'),
                picture_path_to_save, # Update with new or existing path
                data.get('student_type', 'Paid'),
                data.get('remarks'),
                student_id
            )
        )
        
        # Auto-disable account if status changed to Left or Course Completed
        if new_status in ['Left', 'Course Completed'] and old_status != new_status:
            conn.execute('UPDATE students SET account_status = ? WHERE id = ?', ('Inactive', student_id))
            # Log activity
            from datetime import datetime
            conn.execute(
                'INSERT INTO student_activity_log (student_id, activity_type, activity_description, ip_address, created_at) VALUES (?, ?, ?, ?, ?)',
                (student_id, 'account_deactivated', f'Account auto-deactivated due to status change to {new_status}', request.remote_addr, datetime.now().isoformat())
            )
        
        conn.commit()
        conn.close()
        return jsonify({'status': 'success'})
    except Exception as e:
        print(f"Error updating student: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/uploads/<filename>')
def uploaded_file(filename):
    file_extension = filename.split('.')[-1].lower()
    if file_extension in ['jpg', 'jpeg', 'png', 'gif']:
        mimetype = f'image/{file_extension}'
    elif file_extension == 'pdf':
        mimetype = 'application/pdf'
    else:
        mimetype = 'application/octet-stream' # Default for unknown types
    return send_file(os.path.join(UPLOAD_FOLDER, filename), mimetype=mimetype)

@app.route("/api/students/<int:student_id>", methods=['DELETE'])
def delete_student(student_id):
    try:
        conn = db.get_connection()
        conn.execute('DELETE FROM students WHERE id = ?', (student_id,))
        conn.commit()
        conn.close()
        return jsonify({'status': 'success'})
    except Exception as e:
        print(f"Error deleting student: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route("/api/report1")
def report1():
    campus = request.args.get('campus')
    board = request.args.get('board')
    semester = request.args.get('semester')
    technology = request.args.get('technology')
    gender = request.args.get('gender')
    status = request.args.get('status')

    conn = db.get_connection()
    # Include 'technology' and 'admission_no' in the SELECT statement
    query = "SELECT admission_no, name, father_name, phone, technology, gender, student_type, status FROM students"
    params = []
    conditions = []

    if status == 'Free':
        conditions.append("student_type = 'Free'")
    elif status in ['Left', 'Course Completed', 'Active', 'Demoted']:
        conditions.append("status = ?")
        params.append(status)

    if campus and campus != 'All':
        conditions.append("campus = ?")
        params.append(campus)
    if board and board != 'All':
        conditions.append("board = ?")
        params.append(board)
    if semester and semester != 'All':
        conditions.append("semester = ?")
        params.append(semester)
    if technology and technology != 'All':
        conditions.append("technology = ?")
        params.append(technology)
    if gender:
        conditions.append("gender = ?")
        params.append(gender)

    if conditions:
        query += " WHERE " + " AND ".join(conditions)
        
    students = conn.execute(query, params).fetchall()
    conn.close()
    return jsonify([dict(row) for row in students])

@app.route("/api/report1/export_excel", methods=['GET'])
def export_report1_excel():
    campus = request.args.get('campus')
    board = request.args.get('board')
    semester = request.args.get('semester')
    technology = request.args.get('technology')
    gender = request.args.get('gender')
    status = request.args.get('status')

    conn = db.get_connection()
    # Include 'admission_no' and 'technology' in the SELECT statement for Excel export
    query = "SELECT admission_no, name, father_name, phone, technology, gender, student_type, status FROM students"
    params = []
    conditions = []

    if status == 'Free':
        conditions.append("student_type = 'Free'")
    elif status in ['Left', 'Course Completed', 'Active', 'Demoted']:
        conditions.append("status = ?")
        params.append(status)

    if campus and campus != 'All':
        conditions.append("campus = ?")
        params.append(campus)
    if board and board != 'All':
        conditions.append("board = ?")
        params.append(board)
    if semester and semester != 'All':
        conditions.append("semester = ?")
        params.append(semester)
    if technology and technology != 'All':
        conditions.append("technology = ?")
        params.append(technology)
    if gender:
        conditions.append("gender = ?")
        params.append(gender)

    if conditions:
        query += " WHERE " + " AND ".join(conditions)
        
    students = conn.execute(query, params).fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Report 1 Students"
    # Add "Technology" and "Admission No" to Excel headers
    ws.append(["S.NO", "Admission No", "Name", "Father Name", "Phone #", "Technology", "Gender", "Student Type", "Status"])
    for index, student in enumerate(students):
        # Add student['technology'] and student['admission_no'] to the row data
        ws.append([index + 1, student['admission_no'], student['name'], student['father_name'], student['phone'], student['technology'], student['gender'], student['student_type'], student['status']])

    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='report1_students.xlsx'
    )

@app.route("/api/report1/export_pdf", methods=['GET'])
def export_report1_pdf():
    campus = request.args.get('campus')
    board = request.args.get('board')
    semester = request.args.get('semester')
    technology = request.args.get('technology')
    gender = request.args.get('gender')
    status = request.args.get('status')

    conn = db.get_connection()
    # Include 'admission_no' and 'technology' in the SELECT statement for PDF export
    query = "SELECT admission_no, name, father_name, phone, technology, gender, student_type, status FROM students"
    params = []
    conditions = []

    if status == 'Free':
        conditions.append("student_type = 'Free'")
    elif status in ['Left', 'Course Completed', 'Active', 'Demoted']:
        conditions.append("status = ?")
        params.append(status)

    if campus and campus != 'All':
        conditions.append("campus = ?")
        params.append(campus)
    if board and board != 'All':
        conditions.append("board = ?")
        params.append(board)
    if semester and semester != 'All':
        conditions.append("semester = ?")
        params.append(semester)
    if technology and technology != 'All':
        conditions.append("technology = ?")
        params.append(technology)
    if gender:
        conditions.append("gender = ?")
        params.append(gender)

    if conditions:
        query += " WHERE " + " AND ".join(conditions)
        
    students = conn.execute(query, params).fetchall()
    conn.close()

    buffer = io.BytesIO()
    # Use landscape orientation for better table layout
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=0.5*inch, leftMargin=0.5*inch, topMargin=0.75*inch, bottomMargin=0.75*inch)
    styles = getSampleStyleSheet()

    # Create custom styles matching on-screen design
    institute_title_style = ParagraphStyle(
        'InstituteTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#00721c'),  # Green color matching on-screen
        spaceAfter=10,
        alignment=1,  # Center alignment
        fontName='Helvetica-Bold',
        leading=24
    )

    report_title_style = ParagraphStyle(
        'ReportTitle',
        parent=styles['Normal'],
        fontSize=14,
        textColor=colors.HexColor('#333333'),
        spaceAfter=15,
        alignment=1,  # Center alignment
        fontName='Helvetica',
        leading=18
    )

    elements = []

    # Add institute title - matching on-screen design
    semester_text = semester if semester and semester != 'All' else 'All'
    technology_text = technology if technology and technology != 'All' else 'All'
    report_title_text = f"{semester_text} {technology_text} Students List"
    if status:
        report_title_text += f" ({status})"
    
    elements.append(Paragraph("GHAZALI INSTITUTE OF MEDICAL SCIENCES", institute_title_style))
    elements.append(Paragraph(report_title_text, report_title_style))
    elements.append(Spacer(1, 0.15*inch))

    # Add table data - headers in uppercase to match on-screen
    data = [["S.NO", "ADMISSION NO", "NAME", "FATHER NAME", "PHONE #", "GENDER", "TECHNOLOGY", "STUDENT TYPE", "STATUS"]]
    for index, student in enumerate(students):
        data.append([
            str(index + 1),
            student['admission_no'] or '',
            student['name'] or '',
            student['father_name'] or '',
            student['phone'] or '',
            student['gender'] or '',
            student['technology'] or '',
            student['student_type'] or '',
            student['status'] or ''
        ])

    # Calculate column widths for landscape A4 (11.69 inches width - margins = ~10.69 inches usable)
    # Distribute columns proportionally for better spacing
    col_widths = [
        0.5 * inch,   # S.NO
        0.9 * inch,   # Admission No
        1.25 * inch,  # Name
        1.25 * inch,  # Father Name
        1.3 * inch,   # Phone #
        0.75 * inch,  # Gender
        1.2 * inch,   # Technology
        1.1 * inch,   # Student Type
        1.15 * inch   # Status
    ]
    
    table = Table(data, colWidths=col_widths)
    table.setStyle(TableStyle([
        # Header row - green gradient background matching on-screen
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#00721c')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),  # Left align for better readability
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        ('LEFTPADDING', (0, 0), (-1, 0), 8),
        ('RIGHTPADDING', (0, 0), (-1, 0), 8),
        # Data rows - alternate row shading
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ('LEFTPADDING', (0, 1), (-1, -1), 8),
        ('RIGHTPADDING', (0, 1), (-1, -1), 8),
        # Borders - soft borders matching on-screen
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
        ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#005a15')),  # Darker green border below header
    ]))
    elements.append(table)

    # Add footer with record count
    elements.append(Spacer(1, 0.2*inch))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#999999'),
        alignment=2  # Right alignment
    )
    elements.append(Paragraph(f"Total Records: {len(students)}", footer_style))

    doc.build(elements)
    buffer.seek(0)

    return send_file(
        buffer,
        mimetype='application/pdf',
        download_name='report1_students.pdf',
        as_attachment=False
    )

@app.route("/api/report2")
def report2():
    campus = request.args.get('campus')
    board = request.args.get('board')
    technology = request.args.get('technology')
    status = request.args.get('status')

    conn = db.get_connection()
    query = 'SELECT admission_no, name, father_name, phone, technology, status FROM students'
    params = []
    conditions = []

    if status == 'Free':
        conditions.append("student_type = 'Free'")
    elif status in ['Left', 'Course Completed', 'Active', 'Demoted']:
        conditions.append("status = ?")
        params.append(status)

    if campus and campus != 'All':
        conditions.append("campus = ?")
        params.append(campus)
    if board and board != 'All':
        conditions.append("board = ?")
        params.append(board)
    if technology and technology != 'All':
        conditions.append("technology = ?")
        params.append(technology)

    if conditions:
        query += " WHERE " + " AND ".join(conditions)

    students = conn.execute(query, params).fetchall()
    conn.close()
    return jsonify([dict(row) for row in students])

@app.route("/api/report2/export_excel", methods=['GET'])
def export_report2_excel():
    campus = request.args.get('campus')
    board = request.args.get('board')
    technology = request.args.get('technology')
    status = request.args.get('status')

    conn = db.get_connection()
    query = 'SELECT admission_no, name, father_name, phone, technology, status FROM students'
    params = []
    conditions = []

    if status == 'Free':
        conditions.append("student_type = 'Free'")
    elif status in ['Left', 'Course Completed', 'Active', 'Demoted']:
        conditions.append("status = ?")
        params.append(status)

    if campus and campus != 'All':
        conditions.append("campus = ?")
        params.append(campus)
    if board and board != 'All':
        conditions.append("board = ?")
        params.append(board)
    if technology and technology != 'All':
        conditions.append("technology = ?")
        params.append(technology)

    if conditions:
        query += " WHERE " + " AND ".join(conditions)

    students = conn.execute(query, params).fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Report 2 Students"
    ws.append(["S.NO", "Admission No", "Name", "Father Name", "Phone #", "Technology", "Status"])
    for index, student in enumerate(students):
        ws.append([index + 1, student['admission_no'], student['name'], student['father_name'], student['phone'], student['technology'], student['status']])

    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='report2_students.xlsx'
    )

@app.route("/api/report2/export_pdf", methods=['GET'])
def export_report2_pdf():
    campus = request.args.get('campus')
    board = request.args.get('board')
    technology = request.args.get('technology')
    status = request.args.get('status')

    conn = db.get_connection()
    query = 'SELECT admission_no, name, father_name, phone, technology, status FROM students'
    params = []
    conditions = []

    if status == 'Free':
        conditions.append("student_type = 'Free'")
    elif status in ['Left', 'Course Completed', 'Active', 'Demoted']:
        conditions.append("status = ?")
        params.append(status)

    if campus and campus != 'All':
        conditions.append("campus = ?")
        params.append(campus)
    if board and board != 'All':
        conditions.append("board = ?")
        params.append(board)
    if technology and technology != 'All':
        conditions.append("technology = ?")
        params.append(technology)

    if conditions:
        query += " WHERE " + " AND ".join(conditions)

    students = conn.execute(query, params).fetchall()
    conn.close()

    buffer = io.BytesIO()
    # Use landscape orientation for better table layout
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=0.5*inch, leftMargin=0.5*inch, topMargin=0.75*inch, bottomMargin=0.75*inch)
    styles = getSampleStyleSheet()

    # Create custom styles matching on-screen design
    institute_title_style = ParagraphStyle(
        'InstituteTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#00721c'),  # Green color matching on-screen
        spaceAfter=10,
        alignment=1,  # Center alignment
        fontName='Helvetica-Bold',
        leading=24
    )

    report_title_style = ParagraphStyle(
        'ReportTitle',
        parent=styles['Normal'],
        fontSize=14,
        textColor=colors.HexColor('#333333'),
        spaceAfter=15,
        alignment=1,  # Center alignment
        fontName='Helvetica',
        leading=18
    )

    elements = []

    # Add institute title - matching on-screen design
    campus_text = campus if campus and campus != 'All' else 'All'
    board_text = board if board and board != 'All' else 'All'
    technology_text = technology if technology and technology != 'All' else 'All'
    report_title_text = f"{campus_text} - {board_text} - {technology_text} Students List"
    if status:
        report_title_text += f" ({status})"
    
    elements.append(Paragraph("GHAZALI INSTITUTE OF MEDICAL SCIENCES", institute_title_style))
    elements.append(Paragraph(report_title_text, report_title_style))
    elements.append(Spacer(1, 0.15*inch))

    # Add table data - headers in uppercase
    data = [["S.NO", "ADMISSION NO", "NAME", "FATHER NAME", "PHONE #", "TECHNOLOGY", "STATUS"]]
    for index, student in enumerate(students):
        data.append([
            str(index + 1),
            student['admission_no'] or '',
            student['name'] or '',
            student['father_name'] or '',
            student['phone'] or '',
            student['technology'] or '',
            student['status'] or ''
        ])

    # Calculate column widths for landscape A4 (more horizontal space)
    col_widths = [
        0.5 * inch,   # S.NO
        1.0 * inch,   # Admission No
        1.45 * inch,  # Name
        1.45 * inch,  # Father Name
        1.5 * inch,   # Phone #
        1.5 * inch,   # Technology
        1.39 * inch   # Status
    ]
    
    table = Table(data, colWidths=col_widths)
    table.setStyle(TableStyle([
        # Header row - green background matching on-screen
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#00721c')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        ('LEFTPADDING', (0, 0), (-1, 0), 8),
        ('RIGHTPADDING', (0, 0), (-1, 0), 8),
        # Data rows - alternate row shading
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ('LEFTPADDING', (0, 1), (-1, -1), 8),
        ('RIGHTPADDING', (0, 1), (-1, -1), 8),
        # Borders - soft borders matching on-screen
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
        ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#005a15')),
    ]))
    elements.append(table)

    # Add footer with record count
    elements.append(Spacer(1, 0.2*inch))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#999999'),
        alignment=2  # Right alignment
    )
    elements.append(Paragraph(f"Total Records: {len(students)}", footer_style))

    doc.build(elements)
    buffer.seek(0)

    return send_file(
        buffer,
        mimetype='application/pdf',
        download_name='report2_students.pdf',
        as_attachment=False
    )

@app.route("/api/report3")
def report3():
    status = request.args.get('status')

    conn = db.get_connection()
    query = 'SELECT admission_no, name, father_name, phone, campus, board, semester, technology, status FROM students'
    params = []
    conditions = []

    if status == 'Free':
        conditions.append("student_type = 'Free'")
    elif status in ['Left', 'Course Completed', 'Active', 'Demoted']:
        conditions.append("status = ?")
        params.append(status)
    else: # Default for 'All Students' report: exclude 'Left' and include 'Active'
        conditions.append("status = 'Active'")
        conditions.append("status != 'Left'")

    if conditions:
        query += " WHERE " + " AND ".join(conditions)

    students = conn.execute(query, params).fetchall()
    conn.close()
    return jsonify([dict(row) for row in students])

@app.route("/api/report3/export_excel", methods=['GET'])
def export_report3_excel():
    status = request.args.get('status')

    conn = db.get_connection()
    query = 'SELECT admission_no, name, father_name, phone, campus, board, semester, technology, status FROM students'
    params = []
    conditions = []

    if status == 'Free':
        conditions.append("student_type = 'Free'")
    elif status in ['Left', 'Course Completed', 'Active', 'Demoted']:
        conditions.append("status = ?")
        params.append(status)
    else: # Default for 'All Students' report: exclude 'Left' and include 'Active'
        conditions.append("status = 'Active'")
        conditions.append("status != 'Left'")

    if conditions:
        query += " WHERE " + " AND ".join(conditions)

    students = conn.execute(query, params).fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "All Students"
    ws.append(["S.NO", "Admission No", "Name", "Father Name", "Phone #", "Campus", "Board", "Semester", "Technology", "Status"])
    for index, student in enumerate(students):
        ws.append([
            index + 1,
            student['admission_no'],
            student['name'],
            student['father_name'],
            student['phone'],
            student['campus'],
            student['board'],
            student['semester'],
            student['technology'],
            student['status']
        ])

    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='all_students.xlsx'
    )

@app.route("/api/report3/export_pdf", methods=['GET'])
def export_report3_pdf():
    status = request.args.get('status')

    conn = db.get_connection()
    query = 'SELECT admission_no, name, father_name, phone, campus, board, semester, technology, status FROM students'
    params = []
    conditions = []

    if status == 'Free':
        conditions.append("student_type = 'Free'")
    elif status in ['Left', 'Course Completed', 'Active', 'Demoted']:
        conditions.append("status = ?")
        params.append(status)
    else: # Default for 'All Students' report: exclude 'Left' and include 'Active'
        conditions.append("status = 'Active'")
        conditions.append("status != 'Left'")

    if conditions:
        query += " WHERE " + " AND ".join(conditions)

    students = conn.execute(query, params).fetchall()
    conn.close()

    buffer = io.BytesIO()
    # Use landscape orientation for better table layout
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=0.5*inch, leftMargin=0.5*inch, topMargin=0.75*inch, bottomMargin=0.75*inch)
    styles = getSampleStyleSheet()

    # Create custom styles matching on-screen design
    institute_title_style = ParagraphStyle(
        'InstituteTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#00721c'),  # Green color matching on-screen
        spaceAfter=10,
        alignment=1,  # Center alignment
        fontName='Helvetica-Bold',
        leading=24
    )

    report_title_style = ParagraphStyle(
        'ReportTitle',
        parent=styles['Normal'],
        fontSize=14,
        textColor=colors.HexColor('#333333'),
        spaceAfter=15,
        alignment=1,  # Center alignment
        fontName='Helvetica',
        leading=18
    )

    elements = []

    # Add institute title - matching on-screen design
    report_title_text = "All Students List"
    if status:
        report_title_text += f" ({status})"
    
    elements.append(Paragraph("GHAZALI INSTITUTE OF MEDICAL SCIENCES", institute_title_style))
    elements.append(Paragraph(report_title_text, report_title_style))
    elements.append(Spacer(1, 0.15*inch))

    # Add table data - headers in uppercase
    data = [["S.NO", "ADMISSION NO", "NAME", "FATHER NAME", "PHONE #", "CAMPUS", "BOARD", "SEMESTER", "TECHNOLOGY", "STATUS"]]
    for index, student in enumerate(students):
        data.append([
            str(index + 1),
            student['admission_no'] or '',
            student['name'] or '',
            student['father_name'] or '',
            student['phone'] or '',
            student['campus'] or '',
            student['board'] or '',
            student['semester'] or '',
            student['technology'] or '',
            student['status'] or ''
        ])

    # Calculate column widths for landscape A4 - adjust for 10 columns (more horizontal space)
    col_widths = [
        0.5 * inch,   # S.NO
        0.9 * inch,   # Admission No
        1.2 * inch,   # Name
        1.2 * inch,   # Father Name
        1.3 * inch,   # Phone #
        0.85 * inch,  # Campus
        0.85 * inch,  # Board
        0.9 * inch,   # Semester
        1.05 * inch,  # Technology
        0.95 * inch   # Status
    ]
    
    table = Table(data, colWidths=col_widths)
    table.setStyle(TableStyle([
        # Header row - green background matching on-screen
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#00721c')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        ('LEFTPADDING', (0, 0), (-1, 0), 6),
        ('RIGHTPADDING', (0, 0), (-1, 0), 6),
        # Data rows - alternate row shading
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('LEFTPADDING', (0, 1), (-1, -1), 6),
        ('RIGHTPADDING', (0, 1), (-1, -1), 6),
        # Borders - soft borders matching on-screen
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
        ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#005a15')),
    ]))
    elements.append(table)

    # Add footer with record count
    elements.append(Spacer(1, 0.2*inch))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#999999'),
        alignment=2  # Right alignment
    )
    elements.append(Paragraph(f"Total Records: {len(students)}", footer_style))

    doc.build(elements)
    buffer.seek(0)

    return send_file(
        buffer,
        mimetype='application/pdf',
        download_name='all_students.pdf',
        as_attachment=False
    )

@app.route("/api/promote", methods=['POST'])
def promote_students():
    data = request.get_json()
    campus = data['campus']
    board = data['board']
    semester = data['semester']
    next_semester = data['next_semester']

    conn = db.get_connection()
    query = 'UPDATE students SET semester = ? WHERE campus = ? AND board = ? AND semester = ?'
    conn.execute(query, (next_semester, campus, board, semester))
    conn.commit()
    conn.close()
    return jsonify({'status': 'success'})

@app.route("/api/send_sms", methods=['POST'])
def send_sms():
    data = request.get_json()
    numbers = data.get('numbers')  # Phone numbers from frontend
    student_ids = data.get('student_ids')  # Student IDs (alternative)
    message = data.get('message')

    if not message:
        return jsonify({'status': 'error', 'message': 'Message is required.'}), 400

    if not numbers and not student_ids:
        return jsonify({'status': 'error', 'message': 'No phone numbers or student IDs provided.'}), 400

    sent_count = 0
    
    # If phone numbers are directly provided (from SMS tab)
    if numbers:
        for phone in numbers:
            if phone:
                print(f"Sending SMS to {phone}: {message}")
                # In a real application, you would integrate with an SMS API here
                sent_count += 1
    # If student IDs are provided (fetch phone numbers from database)
    elif student_ids:
        conn = db.get_connection()
        cursor = conn.cursor()
        
        for student_id in student_ids:
            # Fetch the student's phone or sms_phone
            student = cursor.execute('SELECT phone, sms_phone FROM students WHERE id = ?', (student_id,)).fetchone()
            if student:
                phone_to_use = student['sms_phone'] if student['sms_phone'] else student['phone']
                if phone_to_use:
                    print(f"Sending SMS to {phone_to_use}: {message}")
                    # In a real application, you would integrate with an SMS API here
                    sent_count += 1
                else:
                    print(f"No phone number found for student ID {student_id}")
            else:
                print(f"Student with ID {student_id} not found.")
        
        conn.close()
    
    return jsonify({'status': 'success', 'message': f'{sent_count} SMS messages queued for sending.'})

@app.route("/api/student_count")
def get_student_count():
    conn = db.get_connection()
    count = conn.execute('SELECT COUNT(*) FROM students').fetchone()[0]
    conn.close()
    return jsonify({'count': count})

@app.route("/api/active_student_count")
def get_active_student_count():
    """Get count of active students"""
    try:
        conn = db.get_connection()
        count = conn.execute("SELECT COUNT(*) FROM students WHERE status = 'Active'").fetchone()[0]
        conn.close()
        return jsonify({'count': count})
    except Exception as e:
        return jsonify({'count': 0, 'error': str(e)}), 500

@app.route("/api/students_by_board")
def get_students_by_board():
    """Get student counts grouped by board"""
    try:
        conn = db.get_connection()
        cur = conn.cursor()
        
        # Get counts for each board
        boards = ['KPK Medical Faculty', 'KMU', 'PNC Board', 'Pharmacy Council']
        board_counts = {}
        
        for board in boards:
            # Use 'PNC Board' in database but display as 'PNC'
            board_name = board
            if board == 'PNC Board':
                query = "SELECT COUNT(*) FROM students WHERE board = 'PNC Board' AND status = 'Active'"
            else:
                query = f"SELECT COUNT(*) FROM students WHERE board = ? AND status = 'Active'"
            
            if board == 'PNC Board':
                count = cur.execute(query).fetchone()[0]
            else:
                count = cur.execute(query, (board,)).fetchone()[0]
            
            # Map board name for display
            display_name = 'PNC' if board == 'PNC Board' else board
            board_counts[display_name] = count
        
        conn.close()
        return jsonify({'status': 'success', 'data': board_counts})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e), 'data': {}}), 500

# New API endpoint for semester-wise student counts
@app.route("/api/students_by_semester", methods=['GET'])
def get_students_by_semester():
    conn = None
    try:
        conn = db.get_connection()
        cursor = conn.cursor()
        # Adjust table and column names based on actual schema
        # Assuming 'semester' column stores semester names like '1st', '2nd', etc.
        cursor.execute("SELECT semester, COUNT(*) FROM students GROUP BY semester ORDER BY semester;")
        results = cursor.fetchall()
        
        semester_counts = {}
        for row in results:
            semester_name = row[0]
            count = row[1]
            semester_counts[semester_name] = count
        
        return jsonify(semester_counts)
    except Exception as e:
        print(f"Error fetching students by semester: {e}")
        return jsonify({"error": "Could not retrieve semester counts"}), 500
    finally:
        if conn:
            conn.close()

# New API endpoint for program-wise student counts
# New API endpoint for organization (board) and semester-wise student counts for dashboard cards
@app.route("/api/students_by_organization_semester", methods=['GET'])
def get_students_by_organization_semester():
    conn = None
    try:
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # Fetch total students per board and their latest semester
        # This query gets the total count for each board and the max semester (assuming '1st Semester', '2nd Semester' etc. can be ordered lexicographically for 'latest')
        # For a more robust "latest semester" if semesters are not strictly lexicographical, a custom sorting logic might be needed.
        cursor.execute("""
            SELECT board, COUNT(*) as total_students, MAX(semester) as latest_semester
            FROM students
            WHERE board IN ('KPK Medical Faculty', 'PNC Board', 'KMU', 'Pharmacy Council')
            GROUP BY board
            ORDER BY board;
        """)
        results = cursor.fetchall()
        
        org_semester_data = {}
        for row in results:
            board_name = row['board']
            org_semester_data[board_name] = {
                'total_students': row['total_students'],
                'latest_semester': row['latest_semester']
            }
        
        return jsonify(org_semester_data)
    except Exception as e:
        print(f"Error fetching students by organization and semester: {e}")
        return jsonify({"error": "Could not retrieve organization and semester counts"}), 500
    finally:
        if conn:
            conn.close()

# New API endpoint to get detailed student counts by semester for a specific board
@app.route("/api/students_by_board_semester_detail/<board_name>", methods=['GET'])
def get_students_by_board_semester_detail(board_name):
    conn = None
    try:
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # Fetch semester-wise counts for the given board
        cursor.execute("""
            SELECT semester, COUNT(*) as student_count
            FROM students
            WHERE board = ?
            GROUP BY semester
            ORDER BY semester;
        """, (board_name,))
        results = cursor.fetchall()
        
        semester_details = []
        for row in results:
            semester_details.append({
                'semester': row['semester'],
                'count': row['student_count']
            })
        
        return jsonify(semester_details)
    except Exception as e:
        print(f"Error fetching semester details for board {board_name}: {e}")
        return jsonify({"error": f"Could not retrieve semester details for {board_name}"}), 500
    finally:
        if conn:
            conn.close()

# Existing API endpoint for program-wise student counts (renamed to avoid confusion with board)
@app.route("/api/students_by_technology_semester", methods=['GET'])
def get_students_by_technology_semester():
    conn = None
    try:
        conn = db.get_connection()
        cursor = conn.cursor()
        # Fetch counts grouped by technology and semester
        cursor.execute("SELECT technology, semester, COUNT(*) FROM students GROUP BY technology, semester ORDER BY technology, semester;")
        results = cursor.fetchall()
        
        program_semester_counts = {}
        for row in results:
            technology = row[0]
            semester = row[1]
            count = row[2]
            
            if technology not in program_semester_counts:
                program_semester_counts[technology] = {'total': 0, 'semesters': {}}
            
            program_semester_counts[technology]['total'] += count
            program_semester_counts[technology]['semesters'][semester] = count
        
        return jsonify(program_semester_counts)
    except Exception as e:
        print(f"Error fetching students by technology and semester: {e}")
        return jsonify({"error": "Could not retrieve technology and semester counts"}), 500
    finally:
        if conn:
            conn.close()

@app.route("/api/students_by_technology_semester_detail", methods=['GET'])
def get_students_by_technology_semester_detail():
    technology = request.args.get('technology')
    semester = request.args.get('semester')

    if not technology or not semester:
        return jsonify({'status': 'error', 'message': 'Technology and Semester are required'}), 400

    conn = None
    try:
        conn = db.get_connection()
        cursor = conn.cursor()
        students = cursor.execute(
            "SELECT id, admission_no, name, father_name, phone, status FROM students WHERE technology = ? AND semester = ? ORDER BY name",
            (technology, semester)
        ).fetchall()
        return jsonify([dict(row) for row in students])
    except Exception as e:
        print(f"Error fetching students by technology and semester: {e}")
        return jsonify({"error": "Could not retrieve student details"}), 500
    finally:
        if conn:
            conn.close()

@app.route("/api/students_for_sms_group", methods=['GET'])
def get_students_for_sms_group():
    campus = request.args.get('campus')
    board = request.args.get('board')
    semester = request.args.get('semester')
    technology = request.args.get('technology') # Optional

    conn = db.get_connection()
    query = 'SELECT id, admission_no, name, father_name, phone, sms_phone FROM students WHERE campus = ? AND board = ? AND semester = ?'
    params = [campus, board, semester]

    if technology:
        query += ' AND technology = ?'
        params.append(technology)

    students = conn.execute(query, params).fetchall()
    conn.close()
    return jsonify([dict(row) for row in students])

@app.route("/api/search_students_for_sms", methods=['GET'])
def search_students_for_sms():
    query = request.args.get('query')
    conn = db.get_connection()
    students = conn.execute(
        "SELECT id, admission_no, name, father_name, phone, sms_phone FROM students WHERE name LIKE ? OR admission_no LIKE ?",
        (f'%{query}%', f'%{query}%')
    ).fetchall()
    conn.close()
    return jsonify([dict(row) for row in students])

@app.route("/api/students_for_promotion", methods=['GET'])
def get_students_for_promotion():
    campus = request.args.get('campus')
    board = request.args.get('board')
    semester = request.args.get('semester')
    search = (request.args.get('search') or '').strip()

    if not campus or not board or not semester:
        return jsonify([])

    conn = db.get_connection()
    try:
        query = '''
            SELECT id, admission_no, name, father_name, semester, status
            FROM students
            WHERE campus = ? AND board = ? AND semester = ?
        '''
        params = [campus, board, semester]
        if search:
            like_search = f'%{search}%'
            query += ' AND (admission_no LIKE ? OR name LIKE ? OR father_name LIKE ?)'
            params.extend([like_search, like_search, like_search])
        query += ' ORDER BY name COLLATE NOCASE'

        students = conn.execute(query, params).fetchall()
        return jsonify([dict(row) for row in students])
    finally:
        conn.close()

@app.route("/api/promote_selected_students", methods=['POST'])
def promote_selected_students():
    data = request.get_json()
    student_ids = data.get('student_ids')
    next_semester = data.get('next_semester')

    if not student_ids or not next_semester:
        return jsonify({'status': 'error', 'message': 'Missing student IDs or next semester.'}), 400

    conn = db.get_connection()
    cursor = conn.cursor()
    try:
        for student_id in student_ids:
            cursor.execute('UPDATE students SET semester = ? WHERE id = ?', (next_semester, student_id))
        conn.commit()
        return jsonify({'status': 'success', 'message': f'{len(student_ids)} students promoted successfully.'})
    except Exception as e:
        conn.rollback()
        print(f"Error promoting students: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        conn.close()

@app.route("/api/search_students_for_demotion", methods=['GET'])
def search_students_for_demotion():
    """Search students for demotion by admission number, name, or father name"""
    search_term = request.args.get('search', '')
    
    if not search_term:
        return jsonify([])
    
    conn = db.get_connection()
    try:
        students = conn.execute(
            """SELECT id, admission_no, name, father_name, semester, status 
               FROM students 
               WHERE (admission_no LIKE ? OR name LIKE ? OR father_name LIKE ?) 
               AND status != ?
               LIMIT 100""",
            (f'%{search_term}%', f'%{search_term}%', f'%{search_term}%', 'Left')
        ).fetchall()
        conn.close()
        return jsonify([dict(row) for row in students])
    except Exception as e:
        print(f"Error searching students for demotion: {e}")
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route("/api/demote_selected_students", methods=['POST'])
def demote_selected_students():
    """Demote selected students to a lower semester and mark them as Demoted"""
    data = request.get_json()
    student_ids = data.get('student_ids')
    previous_semester = data.get('previous_semester')

    if not student_ids or not previous_semester:
        return jsonify({'status': 'error', 'message': 'Missing student IDs or previous semester.'}), 400

    conn = db.get_connection()
    cursor = conn.cursor()
    try:
        for student_id in student_ids:
            cursor.execute('UPDATE students SET semester = ?, status = ? WHERE id = ?', (previous_semester, 'Demoted', student_id))
        conn.commit()
        return jsonify({'status': 'success', 'message': f'{len(student_ids)} students demoted successfully and marked as Demoted.'})
    except Exception as e:
        conn.rollback()
        print(f"Error demoting students: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        conn.close()

@app.route("/api/promote_students", methods=['POST'])
def promote_students_api():
    """Promote selected students or update their status."""
    data = request.get_json() or {}
    student_ids = data.get('student_ids') or []
    new_semester = data.get('new_semester')
    student_status = (data.get('student_status') or '').strip()

    if not student_ids:
        return jsonify({'status': 'error', 'message': 'Missing student IDs.'}), 400

    status_only = student_status in ('Left', 'Course Completed')
    if not status_only and not new_semester:
        return jsonify({'status': 'error', 'message': 'Please select the semester to promote to.'}), 400

    try:
        student_ids = [int(sid) for sid in student_ids]
    except (TypeError, ValueError):
        return jsonify({'status': 'error', 'message': 'Invalid student IDs provided.'}), 400

    conn = db.get_connection()
    cursor = conn.cursor()
    try:
        if status_only:
            for student_id in student_ids:
                cursor.execute('UPDATE students SET status = ? WHERE id = ?', (student_status, student_id))
            conn.commit()
            return jsonify({'status': 'success', 'message': f'{len(student_ids)} student(s) marked as {student_status}.'})
        else:
            for student_id in student_ids:
                if student_status:
                    cursor.execute('UPDATE students SET semester = ?, status = ? WHERE id = ?', (new_semester, student_status, student_id))
                else:
                    cursor.execute('UPDATE students SET semester = ? WHERE id = ?', (new_semester, student_id))
            conn.commit()
            return jsonify({'status': 'success', 'message': f'{len(student_ids)} student(s) promoted successfully.'})
    except Exception as e:
        conn.rollback()
        print(f"Error promoting students: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        conn.close()

@app.route("/api/demote_students", methods=['POST'])
def demote_students_api():
    """Demote selected students to a lower semester and mark them as Demoted"""
    data = request.get_json()
    student_ids = data.get('student_ids')
    new_semester = data.get('new_semester')

    if not student_ids or not new_semester:
        return jsonify({'status': 'error', 'message': 'Missing student IDs or new semester.'}), 400

    conn = db.get_connection()
    cursor = conn.cursor()
    try:
        for student_id in student_ids:
            cursor.execute('UPDATE students SET semester = ?, status = ? WHERE id = ?', (new_semester, 'Demoted', student_id))
        conn.commit()
        return jsonify({'status': 'success', 'message': f'{len(student_ids)} students demoted successfully and marked as Demoted.'})
    except Exception as e:
        conn.rollback()
        print(f"Error demoting students: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        conn.close()

# ==================== ATTENDANCE MODULE ====================

@app.route("/attendance")
@login_required
@module_required('attendance')
def attendance_page():
    """Render the daily attendance page"""
    teacher_name = session.get('teacher_name', 'Guest')
    teacher_role = session.get('role', 'teacher')
    return render_template('attendance.html', teacher_name=teacher_name, teacher_role=teacher_role)

@app.route("/api/attendance/students", methods=['GET'])
@login_required
def get_attendance_students():
    """Get students for attendance marking with filters, respecting teacher's assigned semesters"""
    search_query = request.args.get('search', '')
    technology = request.args.get('technology', '')
    semester = request.args.get('semester', '')
    board = request.args.get('board', '')
    campus = request.args.get('campus', '')
    attendance_date = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))

    teacher_role = session.get('role')
    assigned_semesters = session.get('assigned_semesters', [])

    conn = db.get_connection()

    query = '''
        SELECT s.id, s.admission_no, s.name, s.father_name, s.technology, s.semester, s.board, s.campus,
               COALESCE(a.status, 'Present') as attendance_status, a.notes
        FROM students s
        LEFT JOIN attendance a ON s.id = a.student_id AND a.attendance_date = ?
        WHERE s.status = 'Active'
    '''
    params = [attendance_date]
    conditions = []

    if search_query:
        conditions.append('(s.admission_no LIKE ? OR s.name LIKE ? OR s.father_name LIKE ?)')
        params.extend([f'%{search_query}%', f'%{search_query}%', f'%{search_query}%'])

    if technology:
        conditions.append('s.technology = ?')
        params.append(technology)

    if semester:
        conditions.append('s.semester = ?')
        params.append(semester)

    if board:
        conditions.append('s.board = ?')
        params.append(board)

    if campus:
        conditions.append('s.campus = ?')
        params.append(campus)

    # Apply teacher-specific semester restriction
    if teacher_role == 'teacher' and assigned_semesters:
        semester_placeholders = ','.join(['?' for _ in assigned_semesters])
        conditions.append(f's.semester IN ({semester_placeholders})')
        params.extend(assigned_semesters)
        # If a specific semester is filtered, ensure it's one of the assigned ones
        if semester and semester not in assigned_semesters:
            # If a teacher tries to filter by a semester not assigned to them, return empty
            conn.close()
            return jsonify([])

    if conditions:
        query += ' AND ' + ' AND '.join(conditions)

    query += ' ORDER BY s.name'

    students = conn.execute(query, params).fetchall()
    conn.close()

    return jsonify([dict(row) for row in students])

@app.route("/api/attendance/save", methods=['POST'])
@login_required
def save_attendance():
    """Save attendance for a single student with restrictions"""
    try:
        data = request.get_json()
        student_id = data.get('student_id')
        attendance_date_str = data.get('date', datetime.now().strftime('%Y-%m-%d'))
        status = data.get('status', 'Present')
        notes = data.get('notes', '')

        attendance_date = datetime.strptime(attendance_date_str, '%Y-%m-%d').date()
        today = datetime.now().date()
        teacher_role = session.get('role')
        teacher_id = session.get('teacher_id')

        conn = db.get_connection()
        cursor = conn.cursor()

        # 1. Past Data Editing Restriction
        if attendance_date < today and teacher_role != 'admin':
            conn.close()
            return jsonify({'status': 'error', 'message': 'Teachers cannot edit past attendance data.'}), 403
        
        # 2. 30-minute editing window restriction (for non-admins)
        if teacher_role != 'admin':
            lock_record = conn.execute(
                'SELECT locked_at FROM attendance_lock WHERE attendance_date = ?',
                (attendance_date_str,)
            ).fetchone()

            if lock_record:
                locked_at_time = datetime.fromisoformat(lock_record['locked_at'])
                # Allow editing until 30 minutes past 9:00 AM, or 30 minutes after first save if later
                # Assuming attendance starts at 9:00 AM
                attendance_start_time = datetime.strptime(f"{attendance_date_str} 09:00:00", '%Y-%m-%d %H:%M:%S')
                
                # Determine the effective lock time: 9:00 AM + 30 min, or locked_at_time + 30 min, whichever is later
                # This logic needs refinement based on exact requirement: "9:00 Am ko har din attendance start hogi"
                # and "jab teacher ne save all attendance par click kar deya to 30 minut ke bad koye bhe teacher apne apne panel main attendance main changes na kar pa saky"
                # Let's interpret this as: editing is allowed until 9:30 AM, OR 30 minutes after the first save, whichever is later.
                # For simplicity, let's enforce 30 minutes from the *first save* for that day.
                
                lock_expiry_time = locked_at_time + timedelta(minutes=30)
                
                if datetime.now() > lock_expiry_time:
                    conn.close()
                    return jsonify({'status': 'error', 'message': 'Attendance is locked for changes after 30 minutes from initial save.'}), 403
            # If no lock_record, it means this is the first save for the day, or no save_all has happened yet.
            # The lock will be created by save_all_attendance.

        # 3. Teacher-specific semester restriction (already handled by get_attendance_students, but double-check here)
        student_semester = conn.execute('SELECT semester FROM students WHERE id = ?', (student_id,)).fetchone()
        if student_semester and teacher_role == 'teacher':
            assigned_semesters = session.get('assigned_semesters', [])
            if student_semester['semester'] not in assigned_semesters:
                conn.close()
                return jsonify({'status': 'error', 'message': 'You are not authorized to mark attendance for this student\'s semester.'}), 403

        # Check if attendance record exists
        cursor.execute(
            'SELECT id FROM attendance WHERE student_id = ? AND attendance_date = ?',
            (student_id, attendance_date_str)
        )
        existing = cursor.fetchone()

        if existing:
            # Update existing record
            cursor.execute(
                'UPDATE attendance SET status = ?, notes = ? WHERE student_id = ? AND attendance_date = ?',
                (status, notes, student_id, attendance_date_str)
            )
        else:
            # Insert new record
            cursor.execute(
                'INSERT INTO attendance (student_id, attendance_date, status, notes, created_at) VALUES (?, ?, ?, ?, ?)',
                (student_id, attendance_date_str, status, notes, datetime.now().isoformat())
            )

        conn.commit()
        conn.close()

        return jsonify({'status': 'success', 'message': 'Attendance saved successfully'})
    except Exception as e:
        print(f"Error saving attendance: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route("/api/attendance/save_all", methods=['POST'])
@login_required
def save_all_attendance():
    """Save attendance for all students at once with restrictions"""
    try:
        data = request.get_json()
        attendance_records = data.get('records', [])
        
        if not attendance_records:
            return jsonify({'status': 'success', 'message': 'No records to save.'})

        attendance_date_str = attendance_records[0].get('date', datetime.now().strftime('%Y-%m-%d'))
        attendance_date = datetime.strptime(attendance_date_str, '%Y-%m-%d').date()
        today = datetime.now().date()
        teacher_role = session.get('role')
        teacher_id = session.get('teacher_id')

        conn = db.get_connection()
        cursor = conn.cursor()

        # 1. Past Data Editing Restriction
        if attendance_date < today and teacher_role != 'admin':
            conn.close()
            return jsonify({'status': 'error', 'message': 'Teachers cannot edit past attendance data.'}), 403
        
        # 2. 30-minute editing window restriction (for non-admins)
        if teacher_role != 'admin':
            lock_record = conn.execute(
                'SELECT locked_at FROM attendance_lock WHERE attendance_date = ?',
                (attendance_date_str,)
            ).fetchone()

            if lock_record:
                locked_at_time = datetime.fromisoformat(lock_record['locked_at'])
                lock_expiry_time = locked_at_time + timedelta(minutes=30)
                
                if datetime.now() > lock_expiry_time:
                    conn.close()
                    return jsonify({'status': 'error', 'message': 'Attendance is locked for changes after 30 minutes from initial save.'}), 403
            else:
                # This is the first save_all for this date, create a lock record
                cursor.execute(
                    'INSERT INTO attendance_lock (attendance_date, locked_at, locked_by_teacher_id) VALUES (?, ?, ?)',
                    (attendance_date_str, datetime.now().isoformat(), teacher_id)
                )
                conn.commit() # Commit the lock record immediately

        # 3. Teacher-specific semester restriction (double-check here for all records)
        if teacher_role == 'teacher':
            assigned_semesters = session.get('assigned_semesters', [])
            for record in attendance_records:
                student_id = record.get('student_id')
                student_semester = conn.execute('SELECT semester FROM students WHERE id = ?', (student_id,)).fetchone()
                if student_semester and student_semester['semester'] not in assigned_semesters:
                    conn.close()
                    return jsonify({'status': 'error', 'message': f'You are not authorized to mark attendance for student {student_id}\'s semester.'}), 403

        for record in attendance_records:
            student_id = record.get('student_id')
            status = record.get('status', 'Present')
            notes = record.get('notes', '')

            # Check if attendance record exists
            cursor.execute(
                'SELECT id FROM attendance WHERE student_id = ? AND attendance_date = ?',
                (student_id, attendance_date_str)
            )
            existing = cursor.fetchone()

            if existing:
                cursor.execute(
                    'UPDATE attendance SET status = ?, notes = ? WHERE student_id = ? AND attendance_date = ?',
                    (status, notes, student_id, attendance_date_str)
                )
            else:
                cursor.execute(
                    'INSERT INTO attendance (student_id, attendance_date, status, notes, created_at) VALUES (?, ?, ?, ?, ?)',
                    (student_id, attendance_date_str, status, notes, datetime.now().isoformat())
                )

        conn.commit()
        conn.close()

        return jsonify({'status': 'success', 'message': f'{len(attendance_records)} attendance records saved'})
    except Exception as e:
        conn.rollback() # Rollback if any error occurs during the loop
        print(f"Error saving attendance: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route("/api/attendance/bulk_upload", methods=['POST'])
@login_required
@admin_required
def bulk_upload_attendance():
    """Bulk upload attendance from Excel file (Admin only)"""
    try:
        if 'file' not in request.files:
            return jsonify({'status': 'error', 'message': 'No file provided'}), 400

        file = request.files['file']
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'status': 'error', 'message': 'Invalid file format. Please upload Excel file'}), 400

        df = pd.read_excel(file)
        df.columns = [col.strip().lower() for col in df.columns]

        conn = db.get_connection()
        cursor = conn.cursor()

        success_count = 0
        error_count = 0
        errors = []

        for idx, row in df.iterrows():
            try:
                admission_no = str(row.get('admission_no', '')).strip()
                attendance_date_str = str(row.get('date', datetime.now().strftime('%Y-%m-%d'))).strip()
                status = str(row.get('status', 'Present')).strip()
                notes = str(row.get('notes', '')).strip()

                # Get student ID from admission number
                cursor.execute('SELECT id FROM students WHERE admission_no = ?', (admission_no,))
                student = cursor.fetchone()

                if not student:
                    error_count += 1
                    errors.append(f"Row {idx + 2}: Student with admission no {admission_no} not found")
                    continue

                student_id = student[0]

                # Check if record exists
                cursor.execute(
                    'SELECT id FROM attendance WHERE student_id = ? AND attendance_date = ?',
                    (student_id, attendance_date_str)
                )
                existing = cursor.fetchone()

                if existing:
                    cursor.execute(
                        'UPDATE attendance SET status = ?, notes = ? WHERE student_id = ? AND attendance_date = ?',
                        (status, notes, student_id, attendance_date_str)
                    )
                else:
                    cursor.execute(
                        'INSERT INTO attendance (student_id, attendance_date, status, notes, created_at) VALUES (?, ?, ?, ?, ?)',
                        (student_id, attendance_date_str, status, notes, datetime.now().isoformat())
                    )

                success_count += 1
            except Exception as e:
                error_count += 1
                errors.append(f"Row {idx + 2}: {str(e)}")

        conn.commit()
        conn.close()

        message = f"Uploaded: {success_count} records"
        if error_count > 0:
            message += f", Errors: {error_count}"

        return jsonify({
            'status': 'success',
            'message': message,
            'success_count': success_count,
            'error_count': error_count,
            'errors': errors[:10]  # Return first 10 errors
        })
    except Exception as e:
        print(f"Error in bulk upload: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

# ==================== ATTENDANCE REPORTS ====================

@app.route("/manage_attendance")
@login_required
@admin_required
def manage_attendance_page():
    """Render the manage attendance page (Admin only)"""
    return render_template('manage_attendance.html')

@app.route("/attendance_reports")
@login_required
def attendance_reports_page():
    """Render the attendance reports page"""
    teacher_role = session.get('role', 'teacher')
    return render_template('attendance_reports.html', teacher_role=teacher_role)

@app.route("/api/attendance/daily_report", methods=['GET'])
@login_required
def daily_attendance_report():
    """Get daily attendance report with teacher-specific semester filtering"""
    attendance_date = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    technology = request.args.get('technology', '')
    semester = request.args.get('semester', '')
    board = request.args.get('board', '')
    campus = request.args.get('campus', '')
    admission_no = request.args.get('admission_no', '') # New: Get admission number filter

    teacher_role = session.get('role')
    assigned_semesters = session.get('assigned_semesters', [])

    conn = db.get_connection()

    query = '''
        SELECT s.id, s.admission_no, s.name, s.father_name, s.technology, s.semester, s.board, s.campus,
               COALESCE(a.status, 'Present') as attendance_status, a.notes
        FROM students s
        LEFT JOIN attendance a ON s.id = a.student_id AND a.attendance_date = ?
        WHERE s.status = 'Active'
    '''
    params = [attendance_date]
    conditions = []

    if technology:
        conditions.append('s.technology = ?')
        params.append(technology)

    if semester:
        conditions.append('s.semester = ?')
        params.append(semester)

    if board:
        conditions.append('s.board = ?')
        params.append(board)

    if campus:
        conditions.append('s.campus = ?')
        params.append(campus)

    if admission_no: # New: Add admission number condition
        conditions.append('s.admission_no LIKE ?') # Use LIKE for partial matching
        params.append(f'%{admission_no}%') # Add wildcards for partial matching

    # Apply teacher-specific semester restriction
    if teacher_role == 'teacher' and assigned_semesters:
        semester_placeholders = ','.join(['?' for _ in assigned_semesters])
        conditions.append(f's.semester IN ({semester_placeholders})')
        params.extend(assigned_semesters)
        # If a specific semester is filtered, ensure it's one of the assigned ones
        if semester and semester not in assigned_semesters:
            conn.close()
            return jsonify([])

    if conditions:
        query += ' AND ' + ' AND '.join(conditions)

    query += ' ORDER BY s.name'

    students = conn.execute(query, params).fetchall()
    conn.close()

    return jsonify([dict(row) for row in students])

@app.route("/api/attendance/monthly_report", methods=['GET'])
@login_required
def monthly_attendance_report():
    """Get monthly attendance report with percentages and teacher-specific semester filtering"""
    year_month = request.args.get('month', datetime.now().strftime('%Y-%m'))
    campus = request.args.get('campus', '')
    board = request.args.get('board', '')
    semester = request.args.get('semester', '')
    technology = request.args.get('technology', '')
    admission_no = request.args.get('admission_no', '') # New: Get admission number filter

    teacher_role = session.get('role')
    assigned_semesters = session.get('assigned_semesters', [])

    conn = db.get_connection()

    # Get all students
    query = 'SELECT id, admission_no, name, father_name, technology, semester, board, campus FROM students WHERE status = ?'
    params = ['Active']
    conditions = []

    if campus:
        conditions.append('campus = ?')
        params.append(campus)

    if board:
        conditions.append('board = ?')
        params.append(board)

    if semester:
        conditions.append('semester = ?')
        params.append(semester)

    if technology:
        conditions.append('technology = ?')
        params.append(technology)

    if admission_no: # New: Add admission number condition
        conditions.append('admission_no LIKE ?') # Use LIKE for partial matching
        params.append(f'%{admission_no}%') # Add wildcards for partial matching
    
    # Apply teacher-specific semester restriction
    if teacher_role == 'teacher' and assigned_semesters:
        semester_placeholders = ','.join(['?' for _ in assigned_semesters])
        conditions.append(f'semester IN ({semester_placeholders})')
        params.extend(assigned_semesters)
        if semester and semester not in assigned_semesters:
            conn.close()
            return jsonify([])

    if conditions:
        query += ' AND ' + ' AND '.join(conditions)

    students = conn.execute(query, params).fetchall()

    # Get attendance records for the month
    attendance_query = '''
        SELECT student_id, status, COUNT(*) as count
        FROM attendance
        WHERE strftime('%Y-%m', attendance_date) = ?
GROUP BY student_id, status
    '''
    attendance_records = conn.execute(attendance_query, (year_month,)).fetchall()

    # Build attendance map
    attendance_map = {}
    for record in attendance_records:
        student_id = record['student_id']
        if student_id not in attendance_map:
            attendance_map[student_id] = {'Present': 0, 'Absent': 0, 'Late': 0, 'Leave': 0}
        attendance_map[student_id][record['status']] = record['count']

    # Calculate percentages and build report
    report = []
    low_attendance_students = []

    for student in students:
        student_id = student['id']
        attendance = attendance_map.get(student_id, {'Present': 0, 'Absent': 0, 'Late': 0, 'Leave': 0})

        total_days = sum(attendance.values())
        present_count = attendance['Present']
        absent_count = attendance['Absent']
        late_count = attendance['Late']
        leave_count = attendance['Leave']

        if total_days > 0:
            present_percentage = (present_count / total_days) * 100
            absent_percentage = (absent_count / total_days) * 100
            late_percentage = (late_count / total_days) * 100
        else:
            present_percentage = absent_percentage = late_percentage = 0

        # Flag if attendance is below 70%
        is_low_attendance = present_percentage < 70

        report.append({
            'id': student_id,
            'admission_no': student['admission_no'],
            'name': student['name'],
            'father_name': student['father_name'],
            'technology': student['technology'],
            'semester': student['semester'],
            'board': student['board'],
            'campus': student['campus'],
            'present': present_count,
            'absent': absent_count,
            'late': late_count,
            'leave': leave_count,
            'total_days': total_days,
            'present_percentage': round(present_percentage, 2),
            'absent_percentage': round(absent_percentage, 2),
            'late_percentage': round(late_percentage, 2),
            'is_low_attendance': is_low_attendance
        })

    conn.close()

    return jsonify(report)

@app.route("/api/attendance/monthly_detail_report", methods=['GET'])
def monthly_student_detail_report():
    """Get daily attendance details for a specific student for a given month"""
    student_id = request.args.get('student_id')
    year_month = request.args.get('month', datetime.now().strftime('%Y-%m'))

    if not student_id:
        return jsonify({'status': 'error', 'message': 'Student ID is required'}), 400

    try:
        year = int(year_month.split('-')[0])
        month = int(year_month.split('-')[1])
    except (IndexError, ValueError):
        return jsonify({'status': 'error', 'message': 'Invalid month format. Use YYYY-MM.'}), 400

    conn = db.get_connection()
    try:
        student = conn.execute(
            'SELECT id, admission_no, name, father_name, technology, semester, campus FROM students WHERE id = ?',
            (student_id,)
        ).fetchone()
        if not student:
            return jsonify({'status': 'error', 'message': 'Student not found'}), 404

        attendance_rows = conn.execute(
            '''
                SELECT attendance_date, status, notes
                FROM attendance
                WHERE student_id = ? AND strftime('%Y-%m', attendance_date) = ?
            ''',
            (student_id, year_month)
        ).fetchall()

        days_in_month = calendar.monthrange(year, month)[1]

        if not attendance_rows:
            payload = {
                'status': 'success',
                'student': dict(student),
                'month': year_month,
                'month_label': datetime(year, month, 1).strftime('%B %Y'),
                'records': [],
                'summary': {
                    'total_days': days_in_month,
                    'present': 0,
                    'absent': 0,
                    'late': 0,
                    'leave': 0,
                    'attendance_percentage': 0
                },
                'message': 'No attendance data found for this student.'
            }
            return jsonify(payload)

        attendance_map = {
            row['attendance_date']: {'status': row['status'], 'notes': row['notes']}
            for row in attendance_rows
        }

        records = []
        status_totals = {'Present': 0, 'Absent': 0, 'Late': 0, 'Leave': 0}

        for day in range(1, days_in_month + 1):
            current_date = datetime(year, month, day)
            iso_date = current_date.strftime('%Y-%m-%d')
            display_date = current_date.strftime('%d %b %Y')
            record = attendance_map.get(iso_date)
            if record:
                status = record['status']
                notes = record['notes'] or ''
            else:
                status = 'Absent'
                notes = 'No attendance recorded for this date.'
            status_totals.setdefault(status, 0)
            status_totals[status] += 1
            records.append({
                'attendance_date': iso_date,
                'display_date': display_date,
                'status': status,
                'display_status': status,
                'notes': notes
            })

        attended_days = status_totals.get('Present', 0) + status_totals.get('Late', 0) + status_totals.get('Leave', 0)
        attendance_percentage = round((attended_days / days_in_month) * 100, 2) if days_in_month else 0

        payload = {
            'status': 'success',
            'student': dict(student),
            'month': year_month,
            'month_label': datetime(year, month, 1).strftime('%B %Y'),
            'records': records,
            'summary': {
                'total_days': days_in_month,
                'present': status_totals.get('Present', 0),
                'absent': status_totals.get('Absent', 0),
                'late': status_totals.get('Late', 0),
                'leave': status_totals.get('Leave', 0),
                'attendance_percentage': attendance_percentage
            }
        }
        return jsonify(payload)
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        conn.close()

@app.route("/api/attendance/yearly_summary", methods=['GET'])
@login_required
def yearly_attendance_summary():
    """Get yearly attendance summary with filters"""
    year = request.args.get('year', str(datetime.now().year))
    campus = request.args.get('campus', '')
    board = request.args.get('board', '')
    semester = request.args.get('semester', '')
    technology = request.args.get('technology', '')
    semester_window = request.args.get('semester_window', '').lower()
    
    teacher_role = session.get('role')
    assigned_semesters = session.get('assigned_semesters', [])
    
    conn = db.get_connection()
    
    # Build query for students
    query = '''
        SELECT s.id, s.admission_no, s.name, s.father_name, s.technology, s.semester, s.board, s.campus
        FROM students s
        WHERE s.status = 'Active'
    '''
    params = []
    conditions = []
    
    if technology:
        conditions.append('s.technology = ?')
        params.append(technology)
    
    if semester:
        conditions.append('s.semester = ?')
        params.append(semester)
    
    if board:
        conditions.append('s.board = ?')
        params.append(board)
    
    if campus:
        conditions.append('s.campus = ?')
        params.append(campus)
    
    # Apply teacher-specific semester restriction
    if teacher_role == 'teacher' and assigned_semesters:
        semester_placeholders = ','.join(['?'] * len(assigned_semesters))
        conditions.append(f's.semester IN ({semester_placeholders})')
        params.extend(assigned_semesters)
    
    if conditions:
        query += ' AND ' + ' AND '.join(conditions)
    
    query += ' ORDER BY s.name'
    
    students = conn.execute(query, params).fetchall()
    
    # Get attendance records for the year
    month_filters = ()
    if semester_window == 'spring':
        month_filters = ('01', '02', '03', '04', '05', '06')
    elif semester_window == 'fall':
        month_filters = ('07', '08', '09', '10', '11', '12')

    attendance_query = '''
        SELECT student_id, status, COUNT(*) as count
        FROM attendance
        WHERE strftime('%Y', attendance_date) = ?
    '''
    attendance_params = [year]
    if month_filters:
        placeholders = ','.join(['?'] * len(month_filters))
        attendance_query += f' AND strftime(\'%m\', attendance_date) IN ({placeholders})'
        attendance_params.extend(month_filters)

    attendance_query += ' GROUP BY student_id, status'
    attendance_records = conn.execute(attendance_query, attendance_params).fetchall()
    
    # Build attendance map
    attendance_map = {}
    for record in attendance_records:
        student_id = record['student_id']
        if student_id not in attendance_map:
            attendance_map[student_id] = {'Present': 0, 'Absent': 0, 'Late': 0, 'Leave': 0}
        attendance_map[student_id][record['status']] = record['count']
    
    # Calculate summary statistics
    report = []
    total_present_days = 0
    total_days_all = 0
    
    for student in students:
        student_id = student['id']
        attendance = attendance_map.get(student_id, {'Present': 0, 'Absent': 0, 'Late': 0, 'Leave': 0})
        
        present_days = attendance['Present']
        absent_days = attendance['Absent']
        late_days = attendance['Late']
        leave_days = attendance['Leave']
        total_days = present_days + absent_days + late_days + leave_days
        
        # Calculate attendance percentage (Present + Late + Leave count as attendance)
        attendance_percentage = 0
        if total_days > 0:
            attendance_percentage = round(((present_days + late_days + leave_days) / total_days) * 100, 2)
        
        total_present_days += present_days + late_days + leave_days
        total_days_all += total_days
        
        report.append({
            'id': student_id,
            'admission_no': student['admission_no'],
            'name': student['name'],
            'father_name': student['father_name'],
            'campus': student['campus'],
            'board': student['board'],
            'semester': student['semester'],
            'technology': student['technology'],
            'total_days': total_days,
            'present_days': present_days + late_days + leave_days,
            'attendance_percentage': attendance_percentage
        })
    
    # Calculate average attendance
    average_attendance = 0
    if total_days_all > 0:
        average_attendance = round((total_present_days / total_days_all) * 100, 2)
    
    conn.close()
    
    return jsonify({
        'total_students': len(report),
        'average_attendance': average_attendance,
        'students': report
    })

@app.route("/api/attendance/daily_report/export_pdf", methods=['GET'])
def export_daily_attendance_pdf():
    """Export daily attendance report to PDF"""
    attendance_date = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    technology = request.args.get('technology', '')
    semester = request.args.get('semester', '')
    board = request.args.get('board', '')
    campus = request.args.get('campus', '') # New: Get campus filter
    admission_no = request.args.get('admission_no', '') # New: Get admission number filter

    conn = db.get_connection()

    query = '''
        SELECT s.id, s.admission_no, s.name, s.father_name, s.technology, s.semester, s.board, s.campus,
               COALESCE(a.status, 'Present') as attendance_status, a.notes
        FROM students s
        LEFT JOIN attendance a ON s.id = a.student_id AND a.attendance_date = ?
        WHERE s.status = 'Active'
    '''
    params = [attendance_date]
    conditions = []

    if technology:
        conditions.append('s.technology = ?')
        params.append(technology)

    if semester:
        conditions.append('s.semester = ?')
        params.append(semester)

    if board:
        conditions.append('s.board = ?')
        params.append(board)

    if campus: # New: Add campus condition
        conditions.append('s.campus = ?')
        params.append(campus)

    if admission_no: # New: Add admission number condition
        conditions.append('s.admission_no = ?')
        params.append(admission_no)

    if conditions:
        query += ' AND ' + ' AND '.join(conditions)

    query += ' ORDER BY s.name'

    students = conn.execute(query, params).fetchall()
    conn.close()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=0.5*inch, leftMargin=0.5*inch, topMargin=0.75*inch, bottomMargin=0.75*inch)
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1f4788'),
        spaceAfter=6,
        alignment=1
    )

    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=11,
        textColor=colors.HexColor('#333333'),
        spaceAfter=12,
        alignment=1
    )

    elements = []

    elements.append(Paragraph("Ghazali Institute of Medical Sciences", title_style))
    elements.append(Paragraph("Daily Attendance Report", subtitle_style))
    elements.append(Spacer(1, 0.2*inch))

    filter_text = f"Date: {attendance_date}"
    if technology:
        filter_text += f" | Technology: {technology}"
    if semester:
        filter_text += f" | Semester: {semester}"
    if board:
        filter_text += f" | Board: {board}"
    if campus: # New: Add campus to filter text
        filter_text += f" | Campus: {campus}"

    filter_style = ParagraphStyle(
        'FilterInfo',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#666666'),
        spaceAfter=12,
        alignment=0
    )
    elements.append(Paragraph(filter_text, filter_style))
    elements.append(Spacer(1, 0.1*inch))

    data = [["S.NO", "Admission No", "Name", "Father Name", "Campus", "Technology", "Status", "Reason"]]
    for index, student in enumerate(students):
        status_color = '#28a745' if student['attendance_status'] == 'Present' else '#dc3545' if student['attendance_status'] == 'Absent' else '#ffc107'
        reason = student['notes'] if student['attendance_status'] == 'Leave' else ''
        data.append([
            str(index + 1),
            student['admission_no'] or '',
            student['name'] or '',
            student['father_name'] or '',
            student['campus'] or '',
            student['technology'] or '',
            student['attendance_status'] or 'Present',
            reason
        ])

    table = Table(data, colWidths=[0.4*inch, 0.8*inch, 1*inch, 1*inch, 0.8*inch, 0.8*inch, 0.7*inch, 1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f5f5f5')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
    ]))
    elements.append(table)

    elements.append(Spacer(1, 0.2*inch))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#999999'),
        alignment=2
    )
    elements.append(Paragraph(f"Total Students: {len(students)}", footer_style))

    doc.build(elements)
    buffer.seek(0)

    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'daily_attendance_{attendance_date}.pdf'
    )

@app.route("/api/attendance/monthly_report/export_pdf", methods=['GET'])
def export_monthly_attendance_pdf():
    """Export monthly attendance report to PDF"""
    year_month = request.args.get('month', datetime.now().strftime('%Y-%m'))
    campus = request.args.get('campus', '') # New: Get campus filter
    board = request.args.get('board', '')
    semester = request.args.get('semester', '')
    technology = request.args.get('technology', '')
    admission_no = request.args.get('admission_no', '') # New: Get admission number filter

    conn = db.get_connection()

    # Get all students
    query = 'SELECT id, admission_no, name, father_name, technology, semester, board, campus FROM students WHERE status = ?'
    params = ['Active']
    conditions = []

    if campus: # New: Add campus condition
        conditions.append('campus = ?')
        params.append(campus)

    if board:
        conditions.append('board = ?')
        params.append(board)

    if semester:
        conditions.append('semester = ?')
        params.append(semester)

    if technology:
        conditions.append('technology = ?')
        params.append(technology)

    if admission_no: # New: Add admission number condition
        conditions.append('admission_no = ?')
        params.append(admission_no)

    if conditions:
        query += ' AND ' + ' AND '.join(conditions)

    students = conn.execute(query, params).fetchall()

# Get attendance records for the month
    attendance_query = '''
        SELECT student_id, status, COUNT(*) as count
        FROM attendance
        WHERE strftime('%Y-%m', attendance_date) = ?
GROUP BY student_id, status
    '''
    attendance_records = conn.execute(attendance_query, (year_month,)).fetchall()

    # Build attendance map
    attendance_map = {}
    for record in attendance_records:
        student_id = record['student_id']
        if student_id not in attendance_map:
            attendance_map[student_id] = {'Present': 0, 'Absent': 0, 'Late': 0, 'Leave': 0}
        attendance_map[student_id][record['status']] = record['count']

    # Calculate percentages and build report
    report = []
    low_attendance_students = []

    for student in students:
        student_id = student['id']
        attendance = attendance_map.get(student_id, {'Present': 0, 'Absent': 0, 'Late': 0, 'Leave': 0})

        total_days = sum(attendance.values())
        present_count = attendance['Present']
        absent_count = attendance['Absent']
        late_count = attendance['Late']
        leave_count = attendance['Leave']

        if total_days > 0:
            present_percentage = (present_count / total_days) * 100
        else:
            present_percentage = 0

        is_low_attendance = present_percentage < 70

        student_data = {
            'id': student_id,
            'admission_no': student['admission_no'],
            'name': student['name'],
            'father_name': student['father_name'],
            'technology': student['technology'],
            'semester': student['semester'],
            'board': student['board'],
            'campus': student['campus'], # New: Include campus in report
            'present': present_count,
            'absent': absent_count,
            'late': late_count,
            'leave': leave_count,
            'total_days': total_days,
            'present_percentage': round(present_percentage, 2),
            'is_low_attendance': is_low_attendance
        }

        report.append(student_data)
        if is_low_attendance:
            low_attendance_students.append(student_data)

    conn.close()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=0.5*inch, leftMargin=0.5*inch, topMargin=0.75*inch, bottomMargin=0.75*inch)
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1f4788'),
        spaceAfter=6,
        alignment=1
    )

    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=11,
        textColor=colors.HexColor('#333333'),
        spaceAfter=12,
        alignment=1
    )

    elements = []

    elements.append(Paragraph("Ghazali Institute of Medical Sciences", title_style))
    elements.append(Paragraph("Monthly Attendance Report", subtitle_style))
    elements.append(Spacer(1, 0.2*inch))

    filter_text = f"Month: {year_month}"
    if campus: # New: Add campus to filter text
        filter_text += f" | Campus: {campus}"
    if board:
        filter_text += f" | Board: {board}"
    if semester:
        filter_text += f" | Semester: {semester}"
    if technology:
        filter_text += f" | Technology: {technology}"

    filter_style = ParagraphStyle(
        'FilterInfo',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#666666'),
        spaceAfter=12,
        alignment=0
    )
    elements.append(Paragraph(filter_text, filter_style))
    elements.append(Spacer(1, 0.1*inch))

    # All students table
    elements.append(Paragraph("All Students Attendance", ParagraphStyle('SectionTitle', parent=styles['Heading2'], fontSize=12, textColor=colors.HexColor('#1f4788'), spaceAfter=10)))

    data = [["S.NO", "Admission No", "Name", "Campus", "Board", "Semester", "Technology", "Present", "Absent", "Late", "Leave", "Total", "Percentage"]]
    for index, student in enumerate(report):
        data.append([
            str(index + 1),
            student['admission_no'] or '',
            student['name'] or '',
            student['campus'] or '', # New: Include campus in PDF
            student['board'] or '',
            student['semester'] or '',
            student['technology'] or '',
            str(student['present']),
            str(student['absent']),
            str(student['late']),
            str(student['leave']),
            str(student['total_days']),
            f"{student['present_percentage']}%"
        ])

    table = Table(data, colWidths=[0.3*inch, 0.7*inch, 0.9*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.8*inch, 0.5*inch, 0.5*inch, 0.5*inch, 0.5*inch, 0.5*inch, 0.7*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f5f5f5')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
    ]))
    elements.append(table)

    # Low attendance students section
    if low_attendance_students:
        elements.append(Spacer(1, 0.3*inch))
        elements.append(Paragraph("⚠️ Students with Attendance Below 70%", ParagraphStyle('WarningTitle', parent=styles['Heading2'], fontSize=12, textColor=colors.HexColor('#dc3545'), spaceAfter=10)))

        low_data = [["S.NO", "Admission No", "Name", "Present", "Total Days", "Percentage"]]
        for index, student in enumerate(low_attendance_students):
            low_data.append([
                str(index + 1),
                student['admission_no'] or '',
                student['name'] or '',
                str(student['present']),
                str(student['total_days']),
                f"{student['present_percentage']}%"
            ])

        low_table = Table(low_data, colWidths=[0.4*inch, 0.8*inch, 1.2*inch, 0.8*inch, 1*inch, 1*inch])
        low_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dc3545')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#ffe6e6')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#fff0f0')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dc3545')),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
        ]))
        elements.append(low_table)

    doc.build(elements)
    buffer.seek(0)

    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'monthly_attendance_{year_month}.pdf'
    )

@app.route("/api/attendance/monthly_report/export_excel", methods=['GET'])
def export_monthly_attendance_excel():
    """Export monthly attendance report to Excel"""
    year_month = request.args.get('month', datetime.now().strftime('%Y-%m'))
    campus = request.args.get('campus', '') # New: Get campus filter
    board = request.args.get('board', '')
    semester = request.args.get('semester', '')
    technology = request.args.get('technology', '')
    admission_no = request.args.get('admission_no', '') # New: Get admission number filter

    conn = db.get_connection()

    # Get all students
    query = 'SELECT id, admission_no, name, father_name, technology, semester, board, campus FROM students WHERE status = ?'
    params = ['Active']
    conditions = []

    if campus: # New: Add campus condition
        conditions.append('campus = ?')
        params.append(campus)

    if board:
        conditions.append('board = ?')
        params.append(board)

    if semester:
        conditions.append('semester = ?')
        params.append(semester)

    if technology:
        conditions.append('technology = ?')
        params.append(technology)

    if admission_no: # New: Add admission number condition
        conditions.append('admission_no = ?')
        params.append(admission_no)

    if conditions:
        query += ' AND ' + ' AND '.join(conditions)

    students = conn.execute(query, params).fetchall()

    # Get attendance records for the month
    attendance_query = '''
        SELECT student_id, status, COUNT(*) as count
        FROM attendance
        WHERE strftime('%Y-%m', attendance_date) = ?
        GROUP BY student_id, status
    '''
    attendance_records = conn.execute(attendance_query, (year_month,)).fetchall()

    # Build attendance map
    attendance_map = {}
    for record in attendance_records:
        student_id = record['student_id']
        if student_id not in attendance_map:
            attendance_map[student_id] = {'Present': 0, 'Absent': 0, 'Late': 0, 'Leave': 0}
        attendance_map[student_id][record['status']] = record['count']

    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Attendance"

    # Add headers
    ws.append(["S.NO", "Admission No", "Name", "Father Name", "Campus", "Board", "Semester", "Technology", "Present", "Absent", "Late", "Leave", "Total Days", "Attendance %"])

    # Add data
    for index, student in enumerate(students):
        student_id = student['id']
        attendance = attendance_map.get(student_id, {'Present': 0, 'Absent': 0, 'Late': 0, 'Leave': 0})

        total_days = sum(attendance.values())
        present_count = attendance['Present']

        if total_days > 0:
            present_percentage = (present_count / total_days) * 100
        else:
            present_percentage = 0

        ws.append([
            index + 1,
            student['admission_no'],
            student['name'],
            student['father_name'],
            student['campus'], # New: Include campus in Excel
            student['board'],
            student['semester'],
            student['technology'],
            attendance['Present'],
            attendance['Absent'],
            attendance['Late'],
            attendance['Leave'],
            total_days,
            f"{round(present_percentage, 2)}%"
        ])

    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'monthly_attendance_{year_month}.xlsx'
    )

@app.route("/api/search_students_for_certificates", methods=['GET'])
def search_students_for_certificates():
    """Search for students by admission number, name, or father's name for certificate generation."""
    admission_number = request.args.get('admission_number', '').strip()
    name = request.args.get('name', '').strip()
    father_name = request.args.get('father_name', '').strip()

    conn = None
    try:
        conn = db.get_connection()
        query = "SELECT id, admission_no, name, father_name, technology, semester, campus, board, status, gender FROM students"
        params = []
        conditions = []

        if admission_number:
            conditions.append("admission_no LIKE ?")
            params.append(f"%{admission_number}%")
        if name:
            conditions.append("name LIKE ?")
            params.append(f"%{name}%")
        if father_name:
            conditions.append("father_name LIKE ?")
            params.append(f"%{father_name}%")

        if not conditions:
            return jsonify([]) # No search criteria, return empty list

        query += " WHERE " + " AND ".join(conditions)
        students = conn.execute(query, params).fetchall()
        
        return jsonify([dict(row) for row in students])
    except Exception as e:
        print(f"Error searching students for certificates: {e}")
        return jsonify({'status': 'error', 'message': 'Error searching students. Please try again.'}), 500
    finally:
        if conn:
            conn.close()

@app.route("/certificates", methods=['GET'])
def certificates():
    admission_number = request.args.get('admission_number', '').strip()
    name = request.args.get('name', '').strip()
    father_name = request.args.get('father_name', '').strip()
    certificate_type = request.args.get('certificate_type', '').strip()

    conn = db.get_connection()
    query = "SELECT * FROM students"
    params = []
    search_conditions = []

    if admission_number:
        search_conditions.append("admission_no LIKE ?")
        params.append(f"%{admission_number}%")
    if name:
        search_conditions.append("name LIKE ?")
        params.append(f"%{name}%")
    if father_name:
        search_conditions.append("father_name LIKE ?")
        params.append(f"%{father_name}%")

    if search_conditions:
        query += " WHERE " + " AND ".join(search_conditions)
        students = conn.execute(query, params).fetchall()
    else:
        students = []
    
    conn.close()

    if len(students) == 1:
        return render_template("certificates.html", student=students[0], certificate_type=certificate_type)
    else:
        return render_template("certificates.html", students=students, certificate_type=certificate_type)


# ==================== END CERTIFICATES API ENDPOINTS ====================

# ==================== BONAFIDE AND COURSE COMPLETION CERTIFICATES ====================
def generate_bonafide_pdf(student, reference_number=''):
    buffer = io.BytesIO()
    # Use Portrait orientation (A4 is portrait by default)
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=0.75*inch, leftMargin=0.75*inch, topMargin=0.75*inch, bottomMargin=0.75*inch)
    styles = getSampleStyleSheet()
    
    elements = []
    
    # Title style
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#000000'),
        spaceAfter=20,
        alignment=1  # Center alignment
    )
    
    elements.append(Paragraph("BONAFIDE CERTIFICATE", title_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Date and Reference
    date_text = f"Date: {datetime.now().strftime('%d-%m-%Y')}"
    if reference_number:
        date_text += f"<br/>Ref: {reference_number}"
    elements.append(Paragraph(date_text, styles['Normal']))
    elements.append(Spacer(1, 0.3*inch))
    
    # Certificate text
    technology = student.get('technology', 'N/A') or 'N/A'
    semester = student.get('semester', 'N/A') or 'N/A'
    
    text = f"""
    This is to certify that <b>{student['name']}</b>, son/daughter of <b>{student['father_name']}</b>,
    is a bonafide student of our institution.
    <br/><br/>
    His/Her admission number is <b>{student['admission_no']}</b> and he/she is currently studying in
    <b>{technology}</b>, semester <b>{semester}</b>.
    <br/><br/>
    We wish him/her all the best for his/her future endeavors.
    """
    
    elements.append(Paragraph(text, styles['Normal']))
    
    doc.build(elements)
    buffer.seek(0)
    return buffer

def generate_course_completion_pdf(student, reference_number=''):
    buffer = io.BytesIO()
    # Use Portrait orientation (A4 is portrait by default)
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=0.75*inch, leftMargin=0.75*inch, topMargin=0.75*inch, bottomMargin=0.75*inch)
    styles = getSampleStyleSheet()
    
    elements = []
    
    # Title style
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#000000'),
        spaceAfter=20,
        alignment=1  # Center alignment
    )
    
    elements.append(Paragraph("COURSE COMPLETION CERTIFICATE", title_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Date and Reference
    date_text = f"Date: {datetime.now().strftime('%d-%m-%Y')}"
    if reference_number:
        date_text += f"<br/>Ref: {reference_number}"
    elements.append(Paragraph(date_text, styles['Normal']))
    elements.append(Spacer(1, 0.3*inch))
    
    # Certificate text
    technology = student.get('technology', 'N/A') or 'N/A'
    
    text = f"""
    This is to certify that <b>{student['name']}</b>, son/daughter of <b>{student['father_name']}</b>,
    with admission number <b>{student['admission_no']}</b>, has successfully completed the
    <b>{technology}</b> program.
    <br/><br/>
    We congratulate him/her on this achievement and wish him/her success in all future endeavors.
    """
    
    elements.append(Paragraph(text, styles['Normal']))
    
    doc.build(elements)
    buffer.seek(0)
    return buffer

@app.route("/generate_certificate", methods=['POST'])
def generate_certificate():
    admission_number = request.form.get('admission_number')
    certificate_type = request.form.get('certificate_type')

    conn = db.get_connection()
    student = conn.execute('SELECT * FROM students WHERE admission_no = ?', (admission_number,)).fetchone()
    conn.close()

    if not student:
        flash('Student not found!', 'danger')
        return redirect(url_for('certificates'))

    if certificate_type == 'bonafide':
        pdf_buffer = generate_bonafide_pdf(student)
        return send_file(
            pdf_buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"bonafide_{student['admission_no']}.pdf"
        )
    elif certificate_type == 'course_completion':
        pdf_buffer = generate_course_completion_pdf(student)
        return send_file(
            pdf_buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"course_completion_{student['admission_no']}.pdf"
        )
    else:
        flash('Invalid certificate type!', 'danger')
        return redirect(url_for('certificates'))

# ==================== CERTIFICATES API ENDPOINTS ====================

@app.route("/api/student_by_admission", methods=['GET'])
def get_student_by_admission():
    """Fetch student data by admission number for certificate generation"""
    admission_no = request.args.get('admission_no')

    if not admission_no:
        return jsonify({'status': 'error', 'message': 'Admission number is required'}), 400

    try:
        conn = db.get_connection()
        student = conn.execute(
            'SELECT id, admission_no, name, father_name, technology, semester, campus, board, status FROM students WHERE admission_no = ?',
            (admission_no,)
        ).fetchone()
        conn.close()

        if not student:
            return jsonify({'status': 'error', 'message': 'Student not found'}), 404

        return jsonify(dict(student))
    except Exception as e:
        print(f"Error fetching student: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route("/api/generate_certificate", methods=['POST'])
def api_generate_certificate():
    """Generate certificate PDF for a student"""
    try:
        data = request.get_json()
        student_id = data.get('student_id')
        admission_no = data.get('admission_no')
        certificate_type = data.get('certificate_type')
        reference_number = data.get('reference_number', '')

        if not all([student_id, admission_no, certificate_type]):
            return jsonify({'status': 'error', 'message': 'Missing required fields'}), 400

        if certificate_type not in ['bonafide', 'course_completion']:
            return jsonify({'status': 'error', 'message': 'Invalid certificate type'}), 400

        # Fetch student data - ensure technology and semester are included
        conn = db.get_connection()
        student = conn.execute(
            'SELECT * FROM students WHERE id = ? AND admission_no = ?',
            (student_id, admission_no)
        ).fetchone()
        conn.close()

        if not student:
            return jsonify({'status': 'error', 'message': 'Student not found'}), 404

        # Generate PDF based on certificate type
        if certificate_type == 'bonafide':
            pdf_buffer = generate_bonafide_pdf(student, reference_number)
            filename = f"bonafide_{admission_no}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        else:  # course_completion
            pdf_buffer = generate_course_completion_pdf(student, reference_number)
            filename = f"course_completion_{admission_no}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

        # Save PDF to uploads folder
        pdf_path = os.path.join(UPLOAD_FOLDER, filename)
        with open(pdf_path, 'wb') as f:
            f.write(pdf_buffer.getvalue())

        # Return file URL for download
        file_url = f'/uploads/{filename}'
        return jsonify({
            'status': 'success',
            'message': 'Certificate generated successfully',
            'file_url': file_url,
            'filename': filename
        })

    except Exception as e:
        print(f"Error generating certificate: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

# ==================== END CERTIFICATES API ENDPOINTS ====================

@app.route("/api/meeting_reports/monthly", methods=['GET'])
@login_required
@admin_required
def monthly_meeting_report():
    """
    Generates a monthly report of student strength, broken down by board and semester,
    including paid, free, and total students, with overall summaries.
    """
    year_month = request.args.get('month', datetime.now().strftime('%Y-%m'))
    
    conn = db.get_connection()
    cursor = conn.cursor()

    # Define the boards to report on
    target_boards = ['KPK Medical Faculty', 'PNC Board', 'KMU', 'Pharmacy Council']
    
    report_data = {}
    overall_paid = 0
    overall_free = 0
    overall_total = 0

    for board in target_boards:
        board_data = {}
        
        # Get all semesters for the current board with student counts
        cursor.execute(f"""
            SELECT semester, student_type, COUNT(*) as count
            FROM students
            WHERE board = ? AND status = 'Active'
            AND strftime('%Y-%m', created_at) <= ? -- Consider students enrolled up to or in the report month
            GROUP BY semester, student_type
            ORDER BY semester, student_type
        """, (board, year_month))
        
        semester_results = cursor.fetchall()
        
        semesters_summary = {}
        board_paid = 0
        board_free = 0
        board_total = 0

        for row in semester_results:
            semester_name = row['semester']
            student_type = row['student_type']
            count = row['count']

            if semester_name not in semesters_summary:
                semesters_summary[semester_name] = {'Paid': 0, 'Free': 0, 'Total': 0}
            
            semesters_summary[semester_name][student_type] = count
            semesters_summary[semester_name]['Total'] += count
            
            if student_type == 'Paid':
                board_paid += count
            else:
                board_free += count
            board_total += count
        
        board_data['semesters'] = semesters_summary
        board_data['board_summary'] = {
            'Paid': board_paid,
            'Free': board_free,
            'Total': board_total
        }
        report_data[board] = board_data

        overall_paid += board_paid
        overall_free += board_free
        overall_total += board_total

    conn.close()

    # Calculate overall percentages
    overall_paid_percent = round((overall_paid / overall_total) * 100, 2) if overall_total > 0 else 0
    overall_free_percent = round((overall_free / overall_total) * 100, 2) if overall_total > 0 else 0

    final_report = {
        'report_month': year_month,
        'boards_data': report_data,
        'overall_summary': {
            'Total Paid Students': overall_paid,
            'Total Free Students': overall_free,
            'Grand Total Students': overall_total,
            'Paid Percentage': overall_paid_percent,
            'Free Percentage': overall_free_percent
        }
    }

    return jsonify(final_report)

# ==================== EMPLOYEE MANAGEMENT ENDPOINTS ====================

@app.route('/api/employees', methods=['GET'])
def get_employees():
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute('''
            SELECT e.*, d.name as department_name, des.name as designation_name
            FROM employees e
            LEFT JOIN departments d ON e.department_id = d.id
            LEFT JOIN designations des ON e.designation_id = des.id
            ORDER BY e.name
        ''')
        employees = [dict(row) for row in cur.fetchall()]
        conn.close()
        return jsonify({'status': 'success', 'employees': employees})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/employees', methods=['POST'])
def create_employee():
    try:
        data = request.json
        conn = get_connection()
        cur = conn.cursor()
        cur.execute('''
            INSERT INTO employees (
                name, father_name, cnic, contact, email, campus,
                department_id, designation_id, joining_date,
                basic_salary, status, security_mode, security_amount, security_notes,
                created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            data.get('name'), data.get('father_name'), data.get('cnic'),
            data.get('contact'), data.get('email'), data.get('campus'),
            data.get('department_id'), data.get('designation_id'),
            data.get('joining_date'), data.get('basic_salary', 0),
            data.get('status', 'Active'),
            (data.get('security_mode') or 'none').lower(),
            float(data.get('security_amount') or 0),
            data.get('security_notes'),
            datetime.now().isoformat(), datetime.now().isoformat()
        ))
        conn.commit()
        conn.close()
        return jsonify({'status': 'success', 'message': 'Employee created successfully'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/employees/<int:emp_id>', methods=['DELETE'])
def delete_employee(emp_id):
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute('DELETE FROM employees WHERE id = ?', (emp_id,))
        conn.commit()
        conn.close()
        return jsonify({'status': 'success', 'message': 'Employee deleted successfully'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/employees/options', methods=['GET'])
@login_required
def employee_options():
    """Return lightweight employee data for dropdowns."""
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute('''
            SELECT id, name, father_name, campus, basic_salary
            FROM employees
            WHERE status = 'Active'
            ORDER BY name
        ''')
        employees = [
            {
                'id': row['id'],
                'name': row['name'],
                'father_name': row['father_name'],
                'campus': row['campus'],
                'basic_salary': row['basic_salary']
            } for row in cur.fetchall()
        ]
        conn.close()
        return jsonify({'status': 'success', 'employees': employees})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/departments', methods=['GET'])
def get_departments():
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute('SELECT * FROM departments ORDER BY name')
        departments = [dict(row) for row in cur.fetchall()]
        conn.close()
        return jsonify({'status': 'success', 'departments': departments})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/designations', methods=['GET'])
def get_designations():
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute('SELECT * FROM designations ORDER BY name')
        designations = [dict(row) for row in cur.fetchall()]
        conn.close()
        return jsonify({'status': 'success', 'designations': designations})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/departments', methods=['POST'])
def add_department():
    try:
        data = request.json
        name = data.get('name')
        if not name:
            return jsonify({'status': 'error', 'message': 'Department name is required'}), 400
        
        conn = get_connection()
        cur = conn.cursor()
        cur.execute('INSERT INTO departments (name) VALUES (?)', (name,))
        conn.commit()
        conn.close()
        return jsonify({'status': 'success', 'message': 'Department added successfully'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/departments/<int:dept_id>', methods=['DELETE'])
def delete_department(dept_id):
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute('DELETE FROM departments WHERE id = ?', (dept_id,))
        conn.commit()
        conn.close()
        return jsonify({'status': 'success', 'message': 'Department deleted successfully'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/designations', methods=['POST'])
def add_designation():
    try:
        data = request.json
        name = data.get('name')
        if not name:
            return jsonify({'status': 'error', 'message': 'Designation name is required'}), 400
        
        conn = get_connection()
        cur = conn.cursor()
        cur.execute('INSERT INTO designations (name) VALUES (?)', (name,))
        conn.commit()
        conn.close()
        return jsonify({'status': 'success', 'message': 'Designation added successfully'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/designations/<int:desig_id>', methods=['DELETE'])
def delete_designation(desig_id):
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute('DELETE FROM designations WHERE id = ?', (desig_id,))
        conn.commit()
        conn.close()
        return jsonify({'status': 'success', 'message': 'Designation deleted successfully'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

# ==================== EMPLOYEE ATTENDANCE ENDPOINTS ====================

@app.route('/api/employees/for_attendance', methods=['GET'])
def get_employees_for_attendance():
    """Get employees for attendance marking with existing attendance status"""
    try:
        date = request.args.get('date')
        campus = request.args.get('campus', '')
        department_id = request.args.get('department_id', '')
        
        conn = get_connection()
        cur = conn.cursor()
        
        # Build query
        query = '''
            SELECT e.id, e.name, e.campus, d.name as department_name, des.name as designation_name
            FROM employees e
            LEFT JOIN departments d ON e.department_id = d.id
            LEFT JOIN designations des ON e.designation_id = des.id
            WHERE e.status = 'Active'
        '''
        params = []
        
        if campus:
            query += ' AND e.campus = ?'
            params.append(campus)
        if department_id:
            query += ' AND e.department_id = ?'
            params.append(department_id)
        
        query += ' ORDER BY e.name'
        
        employees = cur.execute(query, params).fetchall()
        
        # Get existing attendance for the date
        existing_attendance = []
        if date:
            attendance_query = '''
                SELECT employee_id, status
                FROM employee_attendance
                WHERE attendance_date = ?
            '''
            existing_attendance = cur.execute(attendance_query, (date,)).fetchall()
        
        conn.close()
        
        return jsonify({
            'status': 'success',
            'employees': [dict(row) for row in employees],
            'existing_attendance': [dict(row) for row in existing_attendance]
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/employee_attendance/check_lock', methods=['GET'])
def check_attendance_lock():
    """Check if attendance is locked for a specific date"""
    try:
        date = request.args.get('date')
        if not date:
            return jsonify({'locked': False})
        
        conn = get_connection()
        cur = conn.cursor()
        
        # Check if there's a lock record
        lock_record = cur.execute(
            'SELECT locked_at FROM employee_attendance_lock WHERE attendance_date = ?',
            (date,)
        ).fetchone()
        
        conn.close()
        
        if lock_record:
            locked_at = datetime.fromisoformat(lock_record['locked_at'])
            # Check if 30 minutes have passed
            time_diff = datetime.now() - locked_at
            if time_diff.total_seconds() >= 1800:  # 30 minutes
                return jsonify({'locked': True, 'locked_at': lock_record['locked_at']})
        
        return jsonify({'locked': False})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/employee_attendance', methods=['POST'])
def create_employee_attendance():
    try:
        data = request.json
        conn = get_connection()
        cur = conn.cursor()
        
        records = data.get('attendance_records', [])
        if not records:
            return jsonify({'status': 'error', 'message': 'No attendance records provided'}), 400
        
        attendance_date = records[0].get('attendance_date') if records else None
        
        # Check if already locked
        if attendance_date:
            lock_check = cur.execute(
                'SELECT locked_at FROM employee_attendance_lock WHERE attendance_date = ?',
                (attendance_date,)
            ).fetchone()
            
            if lock_check:
                locked_at = datetime.fromisoformat(lock_check['locked_at'])
                time_diff = datetime.now() - locked_at
                if time_diff.total_seconds() >= 1800:  # 30 minutes
                    conn.close()
                    return jsonify({'status': 'error', 'message': 'Attendance is locked for this date'}), 403
        
        # Insert/update attendance records
        for record in records:
            cur.execute('''
                INSERT OR REPLACE INTO employee_attendance 
                (employee_id, attendance_date, status, marked_at, marked_by)
                VALUES (?, ?, ?, ?, ?)
            ''', (
                record.get('employee_id'),
                record.get('attendance_date'),
                record.get('status'),
                datetime.now().isoformat(),
                session.get('teacher_id', 1)
            ))
        
        # Create or update lock record (lock starts from submission time)
        if attendance_date:
            cur.execute('''
                INSERT OR REPLACE INTO employee_attendance_lock
                (attendance_date, locked_at, locked_by_teacher_id)
                VALUES (?, ?, ?)
            ''', (
                attendance_date,
                datetime.now().isoformat(),
                session.get('teacher_id', 1)
            ))
        
        conn.commit()
        conn.close()
        return jsonify({'status': 'success', 'message': f'{len(records)} attendance records submitted'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

# ==================== LEAVE MANAGEMENT ENDPOINTS ====================

@app.route('/api/leave_types', methods=['GET'])
def get_leave_types():
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute('SELECT * FROM leave_types ORDER BY name')
        leave_types = [dict(row) for row in cur.fetchall()]
        conn.close()
        return jsonify({'status': 'success', 'leave_types': leave_types})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/leave_requests', methods=['GET'])
def get_leave_requests():
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute('''
            SELECT lr.*, e.name as employee_name, lt.name as leave_type_name
            FROM leave_requests lr
            JOIN employees e ON lr.employee_id = e.id
            JOIN leave_types lt ON lr.leave_type_id = lt.id
            ORDER BY lr.created_at DESC
        ''')
        requests = [dict(row) for row in cur.fetchall()]
        conn.close()
        return jsonify({'status': 'success', 'leave_requests': requests})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

# ==================== PAYROLL ENDPOINTS ====================

@app.route('/api/payroll/generate', methods=['POST'])
@admin_required
def generate_payroll():
    try:
        data = request.json or {}
        month = data.get('month')
        year = data.get('year')
        campus = data.get('campus', '').strip()
        department_id = data.get('department_id')
        status = data.get('status', '').strip()

        if month is None or year is None:
            return jsonify({'status': 'error', 'message': 'Month and Year are required.'}), 400

        month = int(month)
        year = int(year)

        conn = get_connection()
        cur = conn.cursor()

        query = 'SELECT * FROM employees WHERE 1=1'
        params = []
        if status:
            query += ' AND status = ?'
            params.append(status)
        if campus:
            query += ' AND campus = ?'
            params.append(campus)
        if department_id:
            query += ' AND department_id = ?'
            params.append(int(department_id))

        cur.execute(query, params)
        employees = [dict(row) for row in cur.fetchall()]

        generated_count = 0
        for emp in employees:
            upsert_employee_payroll(cur, emp, month, year)
            generated_count += 1

        conn.commit()
        conn.close()
        return jsonify({
            'status': 'success',
            'generated_count': generated_count,
            'message': f'Payroll recalculated for {generated_count} employee(s).'
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/payroll/salary_slips', methods=['GET'])
def get_salary_slips():
    try:
        employee_id = request.args.get('employee_id')
        employee_name = request.args.get('employee_name')
        father_name = request.args.get('father_name')
        month = request.args.get('month')
        year = request.args.get('year')
        
        conn = get_connection()
        cur = conn.cursor()
        
        query = '''
            SELECT p.*, e.name as employee_name, e.father_name, e.cnic
            FROM payroll p
            JOIN employees e ON p.employee_id = e.id
            WHERE 1=1
        '''
        params = []
        
        if employee_id:
            query += ' AND p.employee_id = ?'
            params.append(employee_id)
        
        if employee_name:
            query += ' AND LOWER(e.name) LIKE ?'
            params.append(f'%{employee_name.lower()}%')
        
        if father_name:
            query += ' AND LOWER(e.father_name) LIKE ?'
            params.append(f'%{father_name.lower()}%')
        
        if month:
            query += ' AND p.month = ?'
            params.append(month)
        
        if year:
            query += ' AND p.year = ?'
            params.append(year)
        
        query += ' ORDER BY e.name, p.year DESC, p.month DESC'
        
        cur.execute(query, params)
        salary_slips = [dict(row) for row in cur.fetchall()]
        conn.close()
        
        return jsonify({'status': 'success', 'salary_slips': salary_slips})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

# ==================== DEDUCTIONS ENDPOINTS ====================

@app.route('/api/deductions/generate', methods=['POST'])
@admin_required
def generate_deductions():
    """Generate monthly deductions for employees based on filters"""
    try:
        data = request.json or {}
        month = data.get('month')
        year = data.get('year')
        campus = data.get('campus', '')
        department_id = data.get('department_id', '')
        status = data.get('status', 'Active')

        if month is None or year is None:
            return jsonify({'status': 'error', 'message': 'Month and Year are required.'}), 400

        month = int(month)
        year = int(year)
        department_id = int(department_id) if department_id else None
        
        conn = get_connection()
        cur = conn.cursor()
        
        # Build query to get employees based on filters
        query = '''
            SELECT *
            FROM employees e
            WHERE 1=1
        '''
        params = []
        
        if status:
            query += ' AND e.status = ?'
            params.append(status)
        if campus:
            query += ' AND e.campus = ?'
            params.append(campus)
        if department_id:
            query += ' AND e.department_id = ?'
            params.append(int(department_id))
        
        cur.execute(query, params)
        employees = [dict(row) for row in cur.fetchall()]

        generated_count = 0
        for emp in employees:
            upsert_employee_payroll(cur, emp, month, year)
            generated_count += 1
        
        conn.commit()
        conn.close()
        return jsonify({'status': 'success', 'generated_count': generated_count, 
                       'message': f'Deductions successfully generated for {generated_count} employee(s).'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/deductions', methods=['GET'])
@login_required
def get_deductions():
    """Get deductions with filters"""
    try:
        employee_id = request.args.get('employee_id')
        employee_name = request.args.get('employee_name')
        father_name = request.args.get('father_name')
        month = request.args.get('month')
        year = request.args.get('year')
        campus = request.args.get('campus', '')
        deductions = fetch_deductions_data(employee_id, employee_name, father_name, month, year, campus)
        return jsonify({'status': 'success', 'deductions': deductions})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/deductions/manual', methods=['POST'])
@admin_required
def create_manual_deduction():
    """Record a manual deduction for an employee."""
    data = request.json or {}
    employee_id = data.get('employee_id')
    month = data.get('month')
    year = data.get('year')
    days = float(data.get('days', 0) or 0)
    amount = data.get('amount')
    reason = data.get('reason') or data.get('remarks') or ''
    deduction_type = normalize_deduction_type(data.get('deduction_type'))

    if not employee_id or month is None or year is None:
        return jsonify({'status': 'error', 'message': 'Employee, Month, and Year are required.'}), 400

    try:
        month = int(month)
        year = int(year)
    except ValueError:
        return jsonify({'status': 'error', 'message': 'Invalid month or year provided.'}), 400

    conn = get_connection()
    cur = conn.cursor()
    try:
        employee = cur.execute(
            'SELECT * FROM employees WHERE id = ?',
            (employee_id,)
        ).fetchone()
        if not employee:
            return jsonify({'status': 'error', 'message': 'Employee not found.'}), 404

        basic_salary = float(employee['basic_salary'] or 0)
        per_day_rate = basic_salary / 30 if basic_salary else 0
        calculated_amount = float(amount) if amount not in (None, '') else round(per_day_rate * days, 2)

        cur.execute(
            '''
                INSERT INTO employee_deductions (
                    employee_id, month, year, days_deducted, amount, reason, deduction_type, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''',
            (
                employee_id,
                month,
                year,
                days,
                calculated_amount,
                reason,
                deduction_type,
                datetime.now().isoformat()
            )
        )
        deduction_id = cur.lastrowid
        upsert_employee_payroll(cur, dict(employee), month, year)
        conn.commit()
        return jsonify({
            'status': 'success',
            'message': 'Deduction recorded successfully.',
            'deduction_id': deduction_id,
            'amount': calculated_amount,
            'per_day_rate': per_day_rate,
            'deduction_type': deduction_type
        })
    except Exception as e:
        conn.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        conn.close()

@app.route('/api/deductions/<int:deduction_id>', methods=['PUT'])
@admin_required
def update_manual_deduction(deduction_id):
    """Update an existing deduction entry."""
    data = request.json or {}
    conn = get_connection()
    cur = conn.cursor()
    try:
        deduction = cur.execute(
            'SELECT * FROM employee_deductions WHERE id = ?',
            (deduction_id,)
        ).fetchone()
        if not deduction:
            return jsonify({'status': 'error', 'message': 'Deduction not found.'}), 404

        new_employee_id = int(data.get('employee_id') or deduction['employee_id'])
        month = int(data.get('month') or deduction['month'])
        year = int(data.get('year') or deduction['year'])
        days = float(data.get('days', deduction['days_deducted']) or 0)
        reason = data.get('reason') or data.get('remarks') or deduction['reason'] or ''
        deduction_type = normalize_deduction_type(data.get('deduction_type') or deduction['deduction_type'])

        employee = cur.execute('SELECT * FROM employees WHERE id = ?', (new_employee_id,)).fetchone()
        if not employee:
            return jsonify({'status': 'error', 'message': 'Selected employee not found.'}), 404

        basic_salary = float(employee['basic_salary'] or 0)
        per_day_rate = basic_salary / 30 if basic_salary else 0
        amount = data.get('amount')
        calculated_amount = float(amount) if amount not in (None, '') else round(per_day_rate * days, 2)

        cur.execute(
            '''
                UPDATE employee_deductions
                SET employee_id = ?, month = ?, year = ?, days_deducted = ?, amount = ?, reason = ?, deduction_type = ?
                WHERE id = ?
            ''',
            (new_employee_id, month, year, days, calculated_amount, reason, deduction_type, deduction_id)
        )

        upsert_employee_payroll(cur, dict(employee), month, year)

        # Recalculate payroll for original employee if changed
        if (deduction['employee_id'], deduction['month'], deduction['year']) != (new_employee_id, month, year):
            original_emp = cur.execute('SELECT * FROM employees WHERE id = ?', (deduction['employee_id'],)).fetchone()
            if original_emp:
                upsert_employee_payroll(cur, dict(original_emp), deduction['month'], deduction['year'])

        conn.commit()
        return jsonify({
            'status': 'success',
            'message': 'Deduction updated successfully.',
            'amount': calculated_amount,
            'per_day_rate': per_day_rate,
            'deduction_type': deduction_type
        })
    except Exception as e:
        conn.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        conn.close()

@app.route('/api/deductions/<int:deduction_id>', methods=['DELETE'])
@admin_required
def delete_manual_deduction(deduction_id):
    """Delete a deduction and refresh payroll totals."""
    conn = get_connection()
    cur = conn.cursor()
    try:
        deduction = cur.execute(
            'SELECT * FROM employee_deductions WHERE id = ?',
            (deduction_id,)
        ).fetchone()
        if not deduction:
            return jsonify({'status': 'error', 'message': 'Deduction not found.'}), 404

        cur.execute('DELETE FROM employee_deductions WHERE id = ?', (deduction_id,))
        employee = cur.execute('SELECT * FROM employees WHERE id = ?', (deduction['employee_id'],)).fetchone()
        if employee:
            upsert_employee_payroll(cur, dict(employee), deduction['month'], deduction['year'])

        conn.commit()
        return jsonify({'status': 'success', 'message': 'Deduction removed successfully.'})
    except Exception as e:
        conn.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        conn.close()

@app.route('/api/deductions/employees_overview', methods=['GET'])
@login_required
def deductions_employees_overview():
    """Return active employees with campus, department, and attendance summary."""
    month = request.args.get('month')
    year = request.args.get('year')
    search_query = request.args.get('search', '').strip()
    try:
        if not month:
            month = datetime.now().month
        if not year:
            year = datetime.now().year
        month = int(month)
        year = int(year)
    except ValueError:
        return jsonify({'status': 'error', 'message': 'Invalid month or year.'}), 400

    month_str = f"{month:02d}"
    conn = get_connection()
    cur = conn.cursor()
    try:
        attendance_summary = {}
        cur.execute(
            '''
                SELECT employee_id,
                       SUM(CASE WHEN status = 'Present' THEN 1 ELSE 0 END) AS present_days,
                       SUM(CASE WHEN status = 'Absent' THEN 1 ELSE 0 END) AS absent_days,
                       SUM(CASE WHEN status = 'Leave' THEN 1 ELSE 0 END) AS leave_days,
                       SUM(CASE WHEN status = 'Late' THEN 1 ELSE 0 END) AS late_days
                FROM employee_attendance
                WHERE strftime('%Y', attendance_date) = ? AND strftime('%m', attendance_date) = ?
                GROUP BY employee_id
            ''',
            (str(year), month_str)
        )
        for row in cur.fetchall():
            attendance_summary[row['employee_id']] = {
                'present': row['present_days'] or 0,
                'absent': row['absent_days'] or 0,
                'leave': row['leave_days'] or 0,
                'late': row['late_days'] or 0,
            }

        deductions_map = {}
        cur.execute(
            '''
                SELECT employee_id,
                       SUM(days_deducted) AS total_days,
                       SUM(amount) AS total_amount,
                       MAX(created_at) AS last_entry
                FROM employee_deductions
                WHERE month = ? AND year = ?
                GROUP BY employee_id
            ''',
            (month, year)
        )
        for row in cur.fetchall():
            deductions_map[row['employee_id']] = {
                'days': float(row['total_days'] or 0),
                'amount': float(row['total_amount'] or 0),
                'last_entry': row['last_entry']
            }

        # Build query with optional search filter
        query = '''
                SELECT e.id, e.name, e.father_name, e.campus, e.status,
                       e.basic_salary, d.name AS department_name, des.name AS designation_name
                FROM employees e
                LEFT JOIN departments d ON e.department_id = d.id
                LEFT JOIN designations des ON e.designation_id = des.id
                WHERE e.status = 'Active'
        '''
        params = []
        
        if search_query:
            query += ' AND (LOWER(e.name) LIKE ? OR LOWER(e.father_name) LIKE ?)'
            search_like = f'%{search_query.lower()}%'
            params.extend([search_like, search_like])
        
        query += ' ORDER BY e.name'
        
        cur.execute(query, params)
        employees = []
        for row in cur.fetchall():
            emp = dict(row)
            summary = attendance_summary.get(emp['id'], {'present': 0, 'absent': 0, 'leave': 0, 'late': 0})
            deduction_info = deductions_map.get(emp['id'], {'days': 0, 'amount': 0, 'last_entry': None})
            per_day_rate = (float(emp.get('basic_salary') or 0) / 30) if emp.get('basic_salary') else 0.0
            emp.update({
                'attendance_summary': summary,
                'deduction_summary': deduction_info,
                'per_day_rate': round(per_day_rate, 2)
            })
            employees.append(emp)

        return jsonify({
            'status': 'success',
            'employees': employees,
            'month': month,
            'year': year
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        conn.close()

@app.route('/api/deductions/search_employees', methods=['GET'])
@login_required
def search_deduction_employees():
    """Search employees by name, father_name, or admission number for deductions module."""
    search_query = request.args.get('search', '').strip()
    
    if not search_query:
        return jsonify({'status': 'error', 'message': 'Search query is required.'}), 400
    
    conn = get_connection()
    cur = conn.cursor()
    try:
        search_like = f'%{search_query.lower()}%'
        
        query = '''
            SELECT e.id, e.name, e.father_name, e.campus, e.status,
                   e.basic_salary, d.name AS department_name, des.name AS designation_name
            FROM employees e
            LEFT JOIN departments d ON e.department_id = d.id
            LEFT JOIN designations des ON e.designation_id = des.id
            WHERE e.status = 'Active' 
                AND (LOWER(e.name) LIKE ? OR LOWER(e.father_name) LIKE ?)
            ORDER BY e.name
            LIMIT 50
        '''
        
        cur.execute(query, [search_like, search_like])
        employees = [dict(row) for row in cur.fetchall()]
        
        # Add per_day_rate for each employee
        for emp in employees:
            per_day_rate = (float(emp.get('basic_salary') or 0) / 30) if emp.get('basic_salary') else 0.0
            emp['per_day_rate'] = round(per_day_rate, 2)
        
        return jsonify({
            'status': 'success',
            'employees': employees,
            'count': len(employees)
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        conn.close()

@app.route('/api/deductions/export_excel', methods=['GET'])
@admin_required
def export_deductions_excel():
    """Export deductions to Excel"""
    try:
        employee_id = request.args.get('employee_id')
        employee_name = request.args.get('employee_name')
        father_name = request.args.get('father_name')
        month = request.args.get('month')
        year = request.args.get('year')
        campus = request.args.get('campus', '')

        deductions = fetch_deductions_data(employee_id, employee_name, father_name, month, year, campus)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Deductions Report"
        
        headers = [
            'Employee ID', 'Employee Name', 'Father Name', 'Designation',
            'Department', 'Campus', 'Month', 'Year', 'Deduction Type',
            'Days Deducted', 'Deduction Amount', 'Salary Before Deduction',
            'Salary After Deduction', 'Remarks', 'Date of Entry'
        ]
        ws.append(headers)
        
        month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June',
                       'July', 'August', 'September', 'October', 'November', 'December']
        for ded in deductions:
            month_name = month_names[ded['month']] if ded['month'] else ''
            ws.append([
                ded['employee_id'],
                ded['employee_name'] or '',
                ded['father_name'] or '',
                ded['designation_name'] or '',
                ded['department_name'] or '',
                ded['campus'] or '',
                month_name,
                ded['year'] or '',
                ded.get('deduction_type', 'Other'),
                ded.get('days_deducted', 0),
                ded['amount'] or 0,
                ded.get('salary_before', 0),
                ded.get('salary_after', 0),
                ded.get('reason', '') or '',
                ded.get('entry_date') or ''
            ])
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True, download_name=f'deductions_report_{datetime.now().strftime("%Y%m%d")}.xlsx')
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/deductions/export_pdf', methods=['GET'])
@admin_required
def export_deductions_pdf():
    """Export deductions to PDF"""
    try:
        employee_id = request.args.get('employee_id')
        employee_name = request.args.get('employee_name')
        father_name = request.args.get('father_name')
        month = request.args.get('month')
        year = request.args.get('year')
        campus = request.args.get('campus', '')
        
        deductions = fetch_deductions_data(employee_id, employee_name, father_name, month, year, campus)
        
        # Generate PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4,
                                rightMargin=0.5*inch, leftMargin=0.5*inch,
                                topMargin=0.75*inch, bottomMargin=0.75*inch)
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Institute Title
        title_style = ParagraphStyle(
            'InstituteTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#00721c'),
            alignment=1,  # Center
            fontName='Helvetica-Bold',
            spaceAfter=20
        )
        elements.append(Paragraph('GHAZALI INSTITUTE OF MEDICAL SCIENCES', title_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # Report Title
        report_title = Paragraph('Deductions Report', styles['Heading2'])
        elements.append(report_title)
        elements.append(Spacer(1, 0.3*inch))
        
        # Table data
        month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June',
                      'July', 'August', 'September', 'October', 'November', 'December']
        
        table_data = [['Emp ID', 'Employee Name', 'Department', 'Designation',
                       'Campus', 'Month/Year', 'Type', 'Days', 'Amount',
                       'Salary Before', 'Salary After', 'Remarks', 'Entry Date']]
        
        for ded in deductions:
            month_name = month_names[ded['month']] if ded['month'] else ''
            month_year = f"{month_name}/{ded['year']}" if month_name and ded['year'] else ''
            table_data.append([
                str(ded['employee_id'] or ''),
                ded['employee_name'] or '',
                ded['department_name'] or '',
                ded['designation_name'] or '',
                ded['campus'] or '',
                month_year,
                ded.get('deduction_type', 'Other') or 'Other',
                str(ded.get('days_deducted', 0)),
                f"{float(ded['amount'] or 0):.2f}",
                f"{float(ded.get('salary_before', 0) or 0):.2f}",
                f"{float(ded.get('salary_after', 0) or 0):.2f}",
                ded.get('reason', '') or '',
                ded.get('entry_date') or ''
            ])
        
        # Create table
        table = Table(table_data, colWidths=[0.5*inch, 1.1*inch, 0.9*inch, 0.9*inch,
                                            0.8*inch, 0.8*inch, 0.7*inch, 0.4*inch,
                                            0.7*inch, 0.8*inch, 0.8*inch, 0.9*inch, 0.9*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#00721c')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 0.2*inch))
        
        # Footer
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor('#999999'),
            alignment=2
        )
        elements.append(Paragraph(f"Total Records: {len(deductions)}", footer_style))
        
        doc.build(elements)
        buffer.seek(0)
        
        return send_file(buffer, mimetype='application/pdf',
                        as_attachment=True,
                        download_name=f'deductions_report_{datetime.now().strftime("%Y%m%d")}.pdf')
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

# ==================== EMPLOYEE REPORTS ENDPOINTS ====================

@app.route('/api/employee_reports/attendance_summary', methods=['GET'])
def get_attendance_summary():
    """Get attendance summary report"""
    try:
        month = request.args.get('month')
        year = request.args.get('year')
        campus = request.args.get('campus', '')
        
        if not month or not year:
            return jsonify({'status': 'error', 'message': 'Month and Year are required'}), 400
        
        conn = get_connection()
        cur = conn.cursor()
        
        # Build query to get attendance summary by department
        query = '''
            SELECT 
                d.name as department,
                COUNT(DISTINCT e.id) as total_employees,
                SUM(CASE WHEN ea.status = 'Present' THEN 1 ELSE 0 END) as present_days,
                SUM(CASE WHEN ea.status = 'Absent' THEN 1 ELSE 0 END) as absent_days,
                SUM(CASE WHEN ea.status = 'Leave' THEN 1 ELSE 0 END) as leave_days
            FROM employees e
            LEFT JOIN departments d ON e.department_id = d.id
            LEFT JOIN employee_attendance ea ON e.id = ea.employee_id 
                AND strftime('%Y-%m', ea.attendance_date) = ?
            WHERE e.status = 'Active'
        '''
        params = [f'{year}-{month}']
        
        if campus:
            query += ' AND e.campus = ?'
            params.append(campus)
        
        query += ' GROUP BY d.id, d.name ORDER BY d.name'
        
        results = cur.execute(query, params).fetchall()
        
        # Calculate attendance percentage
        data = []
        for row in results:
            total_days = (row['present_days'] or 0) + (row['absent_days'] or 0) + (row['leave_days'] or 0)
            attendance_pct = ((row['present_days'] or 0) / total_days * 100) if total_days > 0 else 0
            
            data.append({
                'department': row['department'],
                'total_employees': row['total_employees'] or 0,
                'present_days': row['present_days'] or 0,
                'absent_days': row['absent_days'] or 0,
                'leave_days': row['leave_days'] or 0,
                'attendance_percentage': attendance_pct
            })
        
        conn.close()
        return jsonify({'status': 'success', 'data': data})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/employee_reports/employee_list', methods=['GET'])
def get_employee_list_report():
    """Get employee list report"""
    try:
        campus = request.args.get('campus', '')
        department_id = request.args.get('department_id', '')
        
        conn = get_connection()
        cur = conn.cursor()
        
        query = '''
            SELECT e.*, d.name as department_name, des.name as designation_name
            FROM employees e
            LEFT JOIN departments d ON e.department_id = d.id
            LEFT JOIN designations des ON e.designation_id = des.id
            WHERE 1=1
        '''
        params = []
        
        if campus:
            query += ' AND e.campus = ?'
            params.append(campus)
        if department_id:
            query += ' AND e.department_id = ?'
            params.append(department_id)
        
        query += ' ORDER BY e.name'
        
        employees = cur.execute(query, params).fetchall()
        conn.close()
        
        return jsonify({
            'status': 'success',
            'employees': [dict(row) for row in employees]
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/employee_reports/payroll_summary', methods=['GET'])
def get_payroll_summary():
    """Get payroll summary report"""
    try:
        month = request.args.get('month')
        year = request.args.get('year')
        campus = request.args.get('campus', '')
        
        if not month or not year:
            return jsonify({'status': 'error', 'message': 'Month and Year are required'}), 400
        
        conn = get_connection()
        cur = conn.cursor()
        
        query = '''
            SELECT 
                e.id as employee_id,
                e.name,
                d.name as department,
                des.name as designation,
                p.basic_salary,
                p.allowances,
                p.deductions,
                p.net_salary
            FROM payroll p
            JOIN employees e ON p.employee_id = e.id
            LEFT JOIN departments d ON e.department_id = d.id
            LEFT JOIN designations des ON e.designation_id = des.id
            WHERE p.month = ? AND p.year = ?
        '''
        params = [month, year]
        
        if campus:
            query += ' AND e.campus = ?'
            params.append(campus)
        
        query += ' ORDER BY e.name'
        
        results = cur.execute(query, params).fetchall()
        conn.close()
        
        return jsonify({
            'status': 'success',
            'data': [dict(row) for row in results]
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/employee_reports/leave_report', methods=['GET'])
def get_leave_report():
    """Get leave report"""
    try:
        month = request.args.get('month')
        year = request.args.get('year')
        status = request.args.get('status', '')
        
        if not month or not year:
            return jsonify({'status': 'error', 'message': 'Month and Year are required'}), 400
        
        conn = get_connection()
        cur = conn.cursor()
        
        query = '''
            SELECT 
                lr.id,
                e.name as employee_name,
                lt.name as leave_type,
                lr.start_date,
                lr.end_date,
                lr.days,
                lr.reason,
                lr.status,
                t.name as approved_by_name,
                lr.approved_date
            FROM leave_requests lr
            JOIN employees e ON lr.employee_id = e.id
            JOIN leave_types lt ON lr.leave_type_id = lt.id
            LEFT JOIN teachers t ON lr.approved_by = t.id
            WHERE strftime('%Y-%m', lr.start_date) = ?
        '''
        params = [f'{year}-{month}']
        
        if status:
            query += ' AND lr.status = ?'
            params.append(status)
        
        query += ' ORDER BY lr.start_date DESC'
        
        leaves = cur.execute(query, params).fetchall()
        conn.close()
        
        return jsonify({
            'status': 'success',
            'leaves': [dict(row) for row in leaves]
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/employee_reports/attendance_report', methods=['GET'])
def get_attendance_report():
    """Get attendance report"""
    try:
        month = request.args.get('month')
        year = request.args.get('year')
        department_id = request.args.get('department_id', '')
        campus = request.args.get('campus', '')
        status = request.args.get('status', '')
        
        if not month or not year:
            return jsonify({'status': 'error', 'message': 'Month and Year are required'}), 400
        
        conn = get_connection()
        cur = conn.cursor()
        
        query = '''
            SELECT 
                ea.employee_id,
                e.name as employee_name,
                d.name as department,
                des.name as designation,
                e.campus,
                ea.attendance_date,
                ea.status,
                ea.check_in_time,
                ea.check_out_time
            FROM employee_attendance ea
            JOIN employees e ON ea.employee_id = e.id
            LEFT JOIN departments d ON e.department_id = d.id
            LEFT JOIN designations des ON e.designation_id = des.id
            WHERE strftime('%Y-%m', ea.attendance_date) = ?
        '''
        params = [f'{year}-{month}']
        
        if department_id:
            query += ' AND e.department_id = ?'
            params.append(department_id)
        if campus:
            query += ' AND e.campus = ?'
            params.append(campus)
        if status:
            query += ' AND ea.status = ?'
            params.append(status)
        
        query += ' ORDER BY ea.attendance_date DESC, e.name'
        
        attendance = cur.execute(query, params).fetchall()
        conn.close()
        
        return jsonify({
            'status': 'success',
            'attendance': [dict(row) for row in attendance]
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/employee_reports/late_absentee', methods=['GET'])
def get_late_absentee_report():
    """Get late arrival and absentee report"""
    try:
        report_type = request.args.get('type', 'monthly')
        month = request.args.get('month')
        year = request.args.get('year')
        campus = request.args.get('campus', '')
        department_id = request.args.get('department_id', '')
        
        if not month or not year:
            return jsonify({'status': 'error', 'message': 'Month and Year are required'}), 400
        
        conn = get_connection()
        cur = conn.cursor()
        
        # Build date filter based on report type
        if report_type == 'yearly':
            date_filter = f"strftime('%Y', ea.attendance_date) = '{year}'"
        else:
            date_filter = f"strftime('%Y-%m', ea.attendance_date) = '{year}-{month}'"
        
        query = f'''
            SELECT 
                e.id as employee_id,
                e.name,
                d.name as department,
                e.campus,
                SUM(CASE WHEN ea.status = 'Absent' THEN 1 ELSE 0 END) as total_absent_days,
                SUM(CASE WHEN ea.status = 'Late' THEN 1 ELSE 0 END) as total_late_arrivals,
                GROUP_CONCAT(CASE WHEN ea.status = 'Absent' THEN ea.attendance_date END, ', ') as details
            FROM employees e
            LEFT JOIN departments d ON e.department_id = d.id
            LEFT JOIN employee_attendance ea ON e.id = ea.employee_id AND {date_filter}
            WHERE e.status = 'Active'
        '''
        params = []
        
        if campus:
            query += ' AND e.campus = ?'
            params.append(campus)
        if department_id:
            query += ' AND e.department_id = ?'
            params.append(department_id)
        
        query += ' GROUP BY e.id, e.name, d.name, e.campus'
        query += ' HAVING total_absent_days > 0 OR total_late_arrivals > 0'
        query += ' ORDER BY total_absent_days DESC, total_late_arrivals DESC'
        
        results = cur.execute(query, params).fetchall()
        conn.close()
        
        return jsonify({
            'status': 'success',
            'data': [dict(row) for row in results]
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

# Export endpoints (stubs - can be implemented with Excel/PDF libraries)
@app.route('/api/employee_reports/employee_list/export_excel', methods=['GET'])
def export_employee_list_excel():
    # TODO: Implement Excel export
    return jsonify({'status': 'error', 'message': 'Excel export not yet implemented'}), 501

@app.route('/api/employee_reports/employee_list/export_pdf', methods=['GET'])
def export_employee_list_pdf():
    # TODO: Implement PDF export
    return jsonify({'status': 'error', 'message': 'PDF export not yet implemented'}), 501

@app.route('/api/employee_reports/payroll_summary/export_excel', methods=['GET'])
def export_payroll_summary_excel():
    # TODO: Implement Excel export
    return jsonify({'status': 'error', 'message': 'Excel export not yet implemented'}), 501

@app.route('/api/employee_reports/payroll_summary/export_pdf', methods=['GET'])
def export_payroll_summary_pdf():
    # TODO: Implement PDF export
    return jsonify({'status': 'error', 'message': 'PDF export not yet implemented'}), 501

@app.route('/api/employee_reports/leave_report/export_excel', methods=['GET'])
def export_leave_report_excel():
    # TODO: Implement Excel export
    return jsonify({'status': 'error', 'message': 'Excel export not yet implemented'}), 501

@app.route('/api/employee_reports/leave_report/export_pdf', methods=['GET'])
def export_leave_report_pdf():
    # TODO: Implement PDF export
    return jsonify({'status': 'error', 'message': 'PDF export not yet implemented'}), 501

@app.route('/api/employee_reports/attendance_report/export_excel', methods=['GET'])
def export_attendance_report_excel():
    # TODO: Implement Excel export
    return jsonify({'status': 'error', 'message': 'Excel export not yet implemented'}), 501

@app.route('/api/employee_reports/attendance_report/export_pdf', methods=['GET'])
def export_attendance_report_pdf():
    # TODO: Implement PDF export
    return jsonify({'status': 'error', 'message': 'PDF export not yet implemented'}), 501

@app.route('/api/employee_reports/late_absentee/export_excel', methods=['GET'])
def export_late_absentee_excel():
    # TODO: Implement Excel export
    return jsonify({'status': 'error', 'message': 'Excel export not yet implemented'}), 501

@app.route('/api/employee_reports/late_absentee/export_pdf', methods=['GET'])
def export_late_absentee_pdf():
    # TODO: Implement PDF export
    return jsonify({'status': 'error', 'message': 'PDF export not yet implemented'}), 501

# ==================== MASTER DATA MANAGEMENT ENDPOINTS ====================

@app.route('/api/master_data/boards', methods=['GET'])
@login_required
def get_master_boards():
    """Get all boards for master data management"""
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute('SELECT id, name FROM boards ORDER BY name')
        boards = [dict(row) for row in cur.fetchall()]
        conn.close()
        return jsonify({'status': 'success', 'data': boards})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/master_data/boards', methods=['POST'])
@admin_required
def add_master_board():
    """Add a new board"""
    try:
        data = request.json
        name = data.get('name', '').strip()
        if not name:
            return jsonify({'status': 'error', 'message': 'Board name is required'}), 400
        
        conn = get_connection()
        cur = conn.cursor()
        
        # Check for duplicate
        existing = cur.execute('SELECT id FROM boards WHERE name = ?', (name,)).fetchone()
        if existing:
            conn.close()
            return jsonify({'status': 'error', 'message': 'Board name already exists'}), 400
        
        cur.execute('INSERT INTO boards (name) VALUES (?)', (name,))
        conn.commit()
        conn.close()
        return jsonify({'status': 'success', 'message': 'Board added successfully'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/master_data/boards/<int:board_id>', methods=['PUT'])
@admin_required
def update_master_board(board_id):
    """Update a board"""
    try:
        data = request.json
        name = data.get('name', '').strip()
        if not name:
            return jsonify({'status': 'error', 'message': 'Board name is required'}), 400
        
        conn = get_connection()
        cur = conn.cursor()
        
        # Check for duplicate (excluding current record)
        existing = cur.execute('SELECT id FROM boards WHERE name = ? AND id != ?', (name, board_id)).fetchone()
        if existing:
            conn.close()
            return jsonify({'status': 'error', 'message': 'Board name already exists'}), 400
        
        cur.execute('UPDATE boards SET name = ? WHERE id = ?', (name, board_id))
        conn.commit()
        conn.close()
        return jsonify({'status': 'success', 'message': 'Board updated successfully'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/master_data/boards/<int:board_id>', methods=['DELETE'])
@admin_required
def delete_master_board(board_id):
    """Delete a board"""
    try:
        conn = get_connection()
        cur = conn.cursor()
        
        # Check if board is used in students table
        students_count = cur.execute('SELECT COUNT(*) as count FROM students WHERE board = (SELECT name FROM boards WHERE id = ?)', (board_id,)).fetchone()
        if students_count['count'] > 0:
            conn.close()
            return jsonify({'status': 'error', 'message': f'Cannot delete board. It is used by {students_count["count"]} student(s)'}), 400
        
        cur.execute('DELETE FROM boards WHERE id = ?', (board_id,))
        conn.commit()
        conn.close()
        return jsonify({'status': 'success', 'message': 'Board deleted successfully'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/master_data/technologies', methods=['GET'])
@login_required
def get_master_technologies():
    """Get all technologies for master data management"""
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute('SELECT id, name FROM technologies ORDER BY name')
        technologies = [dict(row) for row in cur.fetchall()]
        conn.close()
        return jsonify({'status': 'success', 'data': technologies})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/master_data/technologies', methods=['POST'])
@admin_required
def add_master_technology():
    """Add a new technology"""
    try:
        data = request.json
        name = data.get('name', '').strip()
        if not name:
            return jsonify({'status': 'error', 'message': 'Technology name is required'}), 400
        
        conn = get_connection()
        cur = conn.cursor()
        
        # Check for duplicate
        existing = cur.execute('SELECT id FROM technologies WHERE name = ?', (name,)).fetchone()
        if existing:
            conn.close()
            return jsonify({'status': 'error', 'message': 'Technology name already exists'}), 400
        
        cur.execute('INSERT INTO technologies (name) VALUES (?)', (name,))
        conn.commit()
        conn.close()
        return jsonify({'status': 'success', 'message': 'Technology added successfully'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/master_data/technologies/<int:tech_id>', methods=['PUT'])
@admin_required
def update_master_technology(tech_id):
    """Update a technology"""
    try:
        data = request.json
        name = data.get('name', '').strip()
        if not name:
            return jsonify({'status': 'error', 'message': 'Technology name is required'}), 400
        
        conn = get_connection()
        cur = conn.cursor()
        
        # Check for duplicate (excluding current record)
        existing = cur.execute('SELECT id FROM technologies WHERE name = ? AND id != ?', (name, tech_id)).fetchone()
        if existing:
            conn.close()
            return jsonify({'status': 'error', 'message': 'Technology name already exists'}), 400
        
        cur.execute('UPDATE technologies SET name = ? WHERE id = ?', (name, tech_id))
        conn.commit()
        conn.close()
        return jsonify({'status': 'success', 'message': 'Technology updated successfully'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/master_data/technologies/<int:tech_id>', methods=['DELETE'])
@admin_required
def delete_master_technology(tech_id):
    """Delete a technology"""
    try:
        conn = get_connection()
        cur = conn.cursor()
        
        # Check if technology is used in students table
        tech_name = cur.execute('SELECT name FROM technologies WHERE id = ?', (tech_id,)).fetchone()
        if tech_name:
            students_count = cur.execute('SELECT COUNT(*) as count FROM students WHERE technology = ?', (tech_name['name'],)).fetchone()
            if students_count['count'] > 0:
                conn.close()
                return jsonify({'status': 'error', 'message': f'Cannot delete technology. It is used by {students_count["count"]} student(s)'}), 400
        
        cur.execute('DELETE FROM technologies WHERE id = ?', (tech_id,))
        conn.commit()
        conn.close()
        return jsonify({'status': 'success', 'message': 'Technology deleted successfully'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/master_data/semesters', methods=['GET'])
@login_required
def get_master_semesters():
    """Get all semesters for master data management"""
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute('SELECT id, name FROM semesters ORDER BY name')
        semesters = [dict(row) for row in cur.fetchall()]
        conn.close()
        return jsonify({'status': 'success', 'data': semesters})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/master_data/semesters', methods=['POST'])
@admin_required
def add_master_semester():
    """Add a new semester"""
    try:
        data = request.json
        name = data.get('name', '').strip()
        if not name:
            return jsonify({'status': 'error', 'message': 'Semester name is required'}), 400
        
        conn = get_connection()
        cur = conn.cursor()
        
        # Check for duplicate
        existing = cur.execute('SELECT id FROM semesters WHERE name = ?', (name,)).fetchone()
        if existing:
            conn.close()
            return jsonify({'status': 'error', 'message': 'Semester name already exists'}), 400
        
        cur.execute('INSERT INTO semesters (name) VALUES (?)', (name,))
        conn.commit()
        conn.close()
        return jsonify({'status': 'success', 'message': 'Semester added successfully'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/master_data/semesters/<int:semester_id>', methods=['PUT'])
@admin_required
def update_master_semester(semester_id):
    """Update a semester"""
    try:
        data = request.json
        name = data.get('name', '').strip()
        if not name:
            return jsonify({'status': 'error', 'message': 'Semester name is required'}), 400
        
        conn = get_connection()
        cur = conn.cursor()
        
        # Check for duplicate (excluding current record)
        existing = cur.execute('SELECT id FROM semesters WHERE name = ? AND id != ?', (name, semester_id)).fetchone()
        if existing:
            conn.close()
            return jsonify({'status': 'error', 'message': 'Semester name already exists'}), 400
        
        cur.execute('UPDATE semesters SET name = ? WHERE id = ?', (name, semester_id))
        conn.commit()
        conn.close()
        return jsonify({'status': 'success', 'message': 'Semester updated successfully'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/master_data/semesters/<int:semester_id>', methods=['DELETE'])
@admin_required
def delete_master_semester(semester_id):
    """Delete a semester"""
    try:
        conn = get_connection()
        cur = conn.cursor()
        
        # Check if semester is used in students table
        semester_name = cur.execute('SELECT name FROM semesters WHERE id = ?', (semester_id,)).fetchone()
        if semester_name:
            students_count = cur.execute('SELECT COUNT(*) as count FROM students WHERE semester = ?', (semester_name['name'],)).fetchone()
            if students_count['count'] > 0:
                conn.close()
                return jsonify({'status': 'error', 'message': f'Cannot delete semester. It is used by {students_count["count"]} student(s)'}), 400
        
        cur.execute('DELETE FROM semesters WHERE id = ?', (semester_id,))
        conn.commit()
        conn.close()
        return jsonify({'status': 'success', 'message': 'Semester deleted successfully'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/master_data/campuses', methods=['GET'])
@login_required
def get_master_campuses():
    """Get all campuses for master data management"""
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute('SELECT id, name FROM campuses ORDER BY name')
        campuses = [dict(row) for row in cur.fetchall()]
        conn.close()
        return jsonify({'status': 'success', 'data': campuses})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/master_data/campuses', methods=['POST'])
@admin_required
def add_master_campus():
    """Add a new campus"""
    try:
        data = request.json
        name = data.get('name', '').strip()
        if not name:
            return jsonify({'status': 'error', 'message': 'Campus name is required'}), 400
        
        conn = get_connection()
        cur = conn.cursor()
        
        # Check for duplicate
        existing = cur.execute('SELECT id FROM campuses WHERE name = ?', (name,)).fetchone()
        if existing:
            conn.close()
            return jsonify({'status': 'error', 'message': 'Campus name already exists'}), 400
        
        cur.execute('INSERT INTO campuses (name) VALUES (?)', (name,))
        conn.commit()
        conn.close()
        return jsonify({'status': 'success', 'message': 'Campus added successfully'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/master_data/campuses/<int:campus_id>', methods=['PUT'])
@admin_required
def update_master_campus(campus_id):
    """Update a campus"""
    try:
        data = request.json
        name = data.get('name', '').strip()
        if not name:
            return jsonify({'status': 'error', 'message': 'Campus name is required'}), 400
        
        conn = get_connection()
        cur = conn.cursor()
        
        # Check for duplicate (excluding current record)
        existing = cur.execute('SELECT id FROM campuses WHERE name = ? AND id != ?', (name, campus_id)).fetchone()
        if existing:
            conn.close()
            return jsonify({'status': 'error', 'message': 'Campus name already exists'}), 400
        
        cur.execute('UPDATE campuses SET name = ? WHERE id = ?', (name, campus_id))
        conn.commit()
        conn.close()
        return jsonify({'status': 'success', 'message': 'Campus updated successfully'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/master_data/campuses/<int:campus_id>', methods=['DELETE'])
@admin_required
def delete_master_campus(campus_id):
    """Delete a campus"""
    try:
        conn = get_connection()
        cur = conn.cursor()
        
        # Check if campus is used in students table
        campus_name = cur.execute('SELECT name FROM campuses WHERE id = ?', (campus_id,)).fetchone()
        if campus_name:
            students_count = cur.execute('SELECT COUNT(*) as count FROM students WHERE campus = ?', (campus_name['name'],)).fetchone()
            if students_count['count'] > 0:
                conn.close()
                return jsonify({'status': 'error', 'message': f'Cannot delete campus. It is used by {students_count["count"]} student(s)'}), 400
        
        cur.execute('DELETE FROM campuses WHERE id = ?', (campus_id,))
        conn.commit()
        conn.close()
        return jsonify({'status': 'success', 'message': 'Campus deleted successfully'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/card_generator/students', methods=['GET'])
@admin_required
def get_students_for_cards():
    """Get filtered students for card generation"""
    try:
        campus = request.args.get('campus', '')
        board = request.args.get('board', '')
        technology = request.args.get('technology', '')
        semester = request.args.get('semester', '')
        status = request.args.get('status', '')

        conn = get_connection()
        cur = conn.cursor()

        query = 'SELECT id, admission_no, name, father_name, address, dob, cnic, phone, technology, semester, campus, board, status, photo_path FROM students WHERE 1=1'
        params = []

        if campus:
            query += ' AND campus = ?'
            params.append(campus)
        if board:
            query += ' AND board = ?'
            params.append(board)
        if technology:
            query += ' AND technology = ?'
            params.append(technology)
        if semester:
            query += ' AND semester = ?'
            params.append(semester)
        if status:
            query += ' AND status = ?'
            params.append(status)

        query += ' ORDER BY name'

        cur.execute(query, params)
        students = [dict(row) for row in cur.fetchall()]
        conn.close()

        return jsonify({'status': 'success', 'data': students})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/card_generator/employees', methods=['GET'])
@admin_required
def get_employees_for_cards():
    """Get filtered employees for card generation"""
    try:
        campus = request.args.get('campus', '')
        department_id = request.args.get('department_id', '')
        designation_id = request.args.get('designation_id', '')
        status = request.args.get('status', '')

        conn = get_connection()
        cur = conn.cursor()

        query = '''
            SELECT e.id, e.name, e.father_name, e.cnic, e.contact, e.campus, 
                   e.status, e.photo_path, d.name as department_name, 
                   des.name as designation_name
            FROM employees e
            LEFT JOIN departments d ON e.department_id = d.id
            LEFT JOIN designations des ON e.designation_id = des.id
            WHERE 1=1
        '''
        params = []

        if campus:
            query += ' AND e.campus = ?'
            params.append(campus)
        if department_id:
            query += ' AND e.department_id = ?'
            params.append(int(department_id))
        if designation_id:
            query += ' AND e.designation_id = ?'
            params.append(int(designation_id))
        if status:
            query += ' AND e.status = ?'
            params.append(status)

        query += ' ORDER BY e.name'

        cur.execute(query, params)
        employees = [dict(row) for row in cur.fetchall()]
        conn.close()

        return jsonify({'status': 'success', 'data': employees})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/card_generator/export_pdf', methods=['POST'])
@admin_required
def export_cards_to_pdf():
    """Export selected cards to PDF"""
    try:
        data = request.json
        card_type = data.get('type', 'students')
        ids = data.get('ids', [])

        if not ids:
            return jsonify({'status': 'error', 'message': 'No cards selected'}), 400

        conn = get_connection()
        cur = conn.cursor()

        if card_type == 'students':
            placeholders = ','.join(['?'] * len(ids))
            cur.execute(f'SELECT * FROM students WHERE id IN ({placeholders})', ids)
            records = [dict(row) for row in cur.fetchall()]
        else:
            placeholders = ','.join(['?'] * len(ids))
            cur.execute(f'''
                SELECT e.*, d.name as department_name, des.name as designation_name
                FROM employees e
                LEFT JOIN departments d ON e.department_id = d.id
                LEFT JOIN designations des ON e.designation_id = des.id
                WHERE e.id IN ({placeholders})
            ''', ids)
            records = [dict(row) for row in cur.fetchall()]

        conn.close()

        # Generate PDF using ReportLab
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, 
                                rightMargin=0.5*inch, leftMargin=0.5*inch,
                                topMargin=0.5*inch, bottomMargin=0.5*inch)
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Create cards (5 per page)
        cards_per_page = 5
        for i in range(0, len(records), cards_per_page):
            page_records = records[i:i+cards_per_page]
            # TODO: Create card layout using ReportLab Table/Paragraph
            # For now, return a simple PDF
            for record in page_records:
                name = record.get('name', 'N/A')
                p = Paragraph(f"<b>{name}</b>", styles['Normal'])
                elements.append(p)
                elements.append(Spacer(1, 0.2*inch))

        doc.build(elements)
        buffer.seek(0)

        return send_file(buffer, mimetype='application/pdf', 
                        as_attachment=True, 
                        download_name=f'{card_type}_cards.pdf')
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

# ==================== MIDTERM & TESTS MODULE API ROUTES ====================

# Student Login Routes
@app.route("/student/login", methods=['GET'])
def student_login_page():
    """Student login page"""
    if 'student_logged_in' in session and session.get('student_logged_in'):
        return redirect(url_for('student_dashboard'))
    return render_template('student_login.html')

@app.route("/student/dashboard")
def student_dashboard():
    """Student dashboard page"""
    if 'student_logged_in' not in session or not session.get('student_logged_in'):
        return redirect(url_for('student_login_page'))
    return render_template('student_dashboard.html')

@app.route("/student/exam/<int:exam_id>")
def student_exam_view(exam_id):
    """Render the exam taking interface for a specific exam"""
    if 'student_logged_in' not in session or not session.get('student_logged_in'):
        return redirect(url_for('student_login_page'))

    conn = get_connection()
    exam = conn.execute('SELECT exam_id, title, status FROM midterm_exams WHERE exam_id = ?', (exam_id,)).fetchone()
    conn.close()

    if not exam:
        flash('Exam not found or has been removed.', 'danger')
        return redirect(url_for('student_dashboard'))

    exam_config = {
        'examId': exam['exam_id'],
        'examTitle': exam['title'] or 'Exam Session',
        'mode': 'start'
    }
    return render_template('student_exam.html', exam_config=exam_config)

@app.route("/student/exam/continue/<int:instance_id>")
def student_exam_continue(instance_id):
    """Allow student to continue an in-progress instance by redirecting to the exam page"""
    if 'student_logged_in' not in session or not session.get('student_logged_in'):
        return redirect(url_for('student_login_page'))

    student_id = session.get('student_id')
    conn = get_connection()
    instance = conn.execute(
        'SELECT exam_id, status FROM midterm_instances WHERE instance_id = ? AND student_id = ?',
        (instance_id, student_id)
    ).fetchone()
    conn.close()

    if not instance:
        flash('Exam attempt not found or access denied.', 'danger')
        return redirect(url_for('student_dashboard'))

    if instance['status'] == 'Completed':
        flash('This exam attempt has already been submitted.', 'warning')
        return redirect(url_for('student_dashboard'))

    return redirect(url_for('student_exam_view', exam_id=instance['exam_id']))

@app.route("/student/logout")
def student_logout():
    """Logout student portal session"""
    if session.get('student_logged_in'):
        try:
            conn = get_connection()
            conn.execute(
                'INSERT INTO student_activity_log (student_id, activity_type, activity_description, ip_address, created_at) VALUES (?, ?, ?, ?, ?)',
                (
                    session.get('student_id'),
                    'logout',
                    'Student logged out',
                    request.remote_addr,
                    datetime.now().isoformat()
                )
            )
            conn.commit()
            conn.close()
        except Exception:
            # Ignore logging errors to avoid blocking logout
            pass

    for key in [
        'student_logged_in', 'student_id', 'student_name', 'student_admission_no',
        'student_technology', 'student_semester', 'student_campus', 'student_board'
    ]:
        session.pop(key, None)

    flash('You have been logged out.', 'info')
    return redirect(url_for('student_login_page'))

# Student Login Route
@app.route("/api/student_login", methods=['POST'])
def student_login():
    """Student login endpoint"""
    try:
        data = request.get_json()
        username = data.get('username')
        password = data.get('password')
        
        if not username or not password:
            return jsonify({'status': 'error', 'message': 'Username and password required'}), 400
        
        conn = get_connection()
        student = conn.execute('SELECT * FROM students WHERE username = ?', (username,)).fetchone()
        
        if not student:
            conn.close()
            return jsonify({'status': 'error', 'message': 'Invalid credentials'}), 401
        
        student_dict = dict(student)
        
        # Check if account is active
        if student_dict.get('status') not in ['Active', None] or student_dict.get('account_status') == 'Inactive':
            conn.close()
            return jsonify({'status': 'error', 'message': 'Account is deactivated. Please contact administrator.'}), 403
        
        # Check password
        if not student_dict.get('password_hash') or not check_password_hash(student_dict['password_hash'], password):
            conn.close()
            return jsonify({'status': 'error', 'message': 'Invalid credentials'}), 401
        
        # Update last login
        from datetime import datetime
        conn.execute('UPDATE students SET last_login = ? WHERE id = ?', (datetime.now().isoformat(), student_dict['id']))
        
        # Log activity
        conn.execute(
            'INSERT INTO student_activity_log (student_id, activity_type, activity_description, ip_address, created_at) VALUES (?, ?, ?, ?, ?)',
            (student_dict['id'], 'login', 'Student logged in', request.remote_addr, datetime.now().isoformat())
        )
        conn.commit()
        conn.close()
        
        session['student_logged_in'] = True
        session['student_id'] = student_dict['id']
        session['student_name'] = student_dict['name']
        session['student_admission_no'] = student_dict['admission_no']
        session['student_technology'] = student_dict.get('technology')
        session['student_semester'] = student_dict.get('semester')
        session['student_campus'] = student_dict.get('campus')
        session['student_board'] = student_dict.get('board')
        
        return jsonify({
            'status': 'success',
            'message': 'Login successful',
            'student': {
                'id': student_dict['id'],
                'name': student_dict['name'],
                'admission_no': student_dict['admission_no']
            }
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

# Exam Management Routes
@app.route("/api/exams", methods=['GET'])
@login_required
def get_exams():
    """Get all exams with filters"""
    try:
        role = session.get('role')
        teacher_id = session.get('teacher_id')
        campus = request.args.get('campus')
        semester = request.args.get('semester')
        status = request.args.get('status')
        
        conn = get_connection()
        cur = conn.cursor()
        
        query = "SELECT * FROM midterm_exams WHERE 1=1"
        params = []
        
        # Teachers can only see their own exams or published exams
        if role == 'teacher':
            query += " AND (created_by = ? OR status = 'Published')"
            params.append(teacher_id)
        
        if campus:
            query += " AND campus = ?"
            params.append(campus)
        if semester:
            query += " AND semester = ?"
            params.append(semester)
        if status:
            query += " AND status = ?"
            params.append(status)
        
        query += " ORDER BY created_at DESC"
        
        cur.execute(query, params)
        exams = [dict(row) for row in cur.fetchall()]
        conn.close()
        
        return jsonify({'status': 'success', 'data': exams})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route("/api/exams", methods=['POST'])
@login_required
def create_exam():
    """Create a new exam"""
    try:
        data = request.get_json()
        teacher_id = session.get('teacher_id')
        role = session.get('role')
        
        # Only admin and teachers can create exams
        if role not in ['admin', 'teacher']:
            return jsonify({'status': 'error', 'message': 'Unauthorized'}), 403
        
        conn = get_connection()
        cur = conn.cursor()
        
        cur.execute('''
            INSERT INTO midterm_exams (
                title, subject, campus, board, technology, semester,
                exam_date, start_time, end_time, duration, total_questions,
                marks_per_question, passing_marks, negative_marking,
                negative_marks_value, randomize_questions, randomize_options,
                allowed_devices, ip_restrictions, max_focus_losses,
                heartbeat_interval, auto_submit_violations, config_json,
                status, created_by, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            data.get('title'), data.get('subject'), data.get('campus'),
            data.get('board'), data.get('technology'), data.get('semester'),
            data.get('exam_date'), data.get('start_time'), data.get('end_time'),
            data.get('duration'), data.get('total_questions'),
            data.get('marks_per_question', 1.0), data.get('passing_marks'),
            data.get('negative_marking', 0), data.get('negative_marks_value', 0.0),
            data.get('randomize_questions', 0), data.get('randomize_options', 0),
            data.get('allowed_devices'), data.get('ip_restrictions'),
            data.get('max_focus_losses', 3), data.get('heartbeat_interval', 30),
            data.get('auto_submit_violations', 5), json.dumps(data.get('config_json', {})),
            data.get('status', 'Draft'), teacher_id,
            datetime.now().isoformat(), datetime.now().isoformat()
        ))
        
        exam_id = cur.lastrowid
        conn.commit()
        conn.close()
        
        return jsonify({'status': 'success', 'message': 'Exam created successfully', 'exam_id': exam_id})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route("/api/exams/<int:exam_id>", methods=['GET'])
@login_required
def get_exam(exam_id):
    """Get exam details"""
    try:
        conn = get_connection()
        exam = conn.execute('SELECT * FROM midterm_exams WHERE exam_id = ?', (exam_id,)).fetchone()
        conn.close()
        
        if not exam:
            return jsonify({'status': 'error', 'message': 'Exam not found'}), 404
        
        return jsonify({'status': 'success', 'data': dict(exam)})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route("/api/exams/<int:exam_id>", methods=['PUT'])
@login_required
def update_exam(exam_id):
    """Update exam"""
    try:
        data = request.get_json()
        teacher_id = session.get('teacher_id')
        role = session.get('role')
        
        conn = get_connection()
        cur = conn.cursor()
        
        # Check permissions
        exam = cur.execute('SELECT created_by FROM midterm_exams WHERE exam_id = ?', (exam_id,)).fetchone()
        if not exam:
            return jsonify({'status': 'error', 'message': 'Exam not found'}), 404
        
        if role != 'admin' and exam['created_by'] != teacher_id:
            return jsonify({'status': 'error', 'message': 'Unauthorized'}), 403
        
        cur.execute('''
            UPDATE midterm_exams SET
                title = ?, subject = ?, campus = ?, board = ?, technology = ?, semester = ?,
                exam_date = ?, start_time = ?, end_time = ?, duration = ?, total_questions = ?,
                marks_per_question = ?, passing_marks = ?, negative_marking = ?,
                negative_marks_value = ?, randomize_questions = ?, randomize_options = ?,
                allowed_devices = ?, ip_restrictions = ?, max_focus_losses = ?,
                heartbeat_interval = ?, auto_submit_violations = ?, config_json = ?,
                status = ?, updated_at = ?
            WHERE exam_id = ?
        ''', (
            data.get('title'), data.get('subject'), data.get('campus'),
            data.get('board'), data.get('technology'), data.get('semester'),
            data.get('exam_date'), data.get('start_time'), data.get('end_time'),
            data.get('duration'), data.get('total_questions'),
            data.get('marks_per_question'), data.get('passing_marks'),
            data.get('negative_marking'), data.get('negative_marks_value'),
            data.get('randomize_questions'), data.get('randomize_options'),
            data.get('allowed_devices'), data.get('ip_restrictions'),
            data.get('max_focus_losses'), data.get('heartbeat_interval'),
            data.get('auto_submit_violations'), json.dumps(data.get('config_json', {})),
            data.get('status'), datetime.now().isoformat(), exam_id
        ))
        
        conn.commit()
        conn.close()
        
        return jsonify({'status': 'success', 'message': 'Exam updated successfully'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route("/api/exams/<int:exam_id>", methods=['DELETE'])
@login_required
def delete_exam(exam_id):
    """Delete exam"""
    try:
        role = session.get('role')
        teacher_id = session.get('teacher_id')
        
        conn = get_connection()
        cur = conn.cursor()
        
        # Check permissions
        exam = cur.execute('SELECT created_by FROM midterm_exams WHERE exam_id = ?', (exam_id,)).fetchone()
        if not exam:
            return jsonify({'status': 'error', 'message': 'Exam not found'}), 404
        
        if role != 'admin' and exam['created_by'] != teacher_id:
            return jsonify({'status': 'error', 'message': 'Unauthorized'}), 403
        
        cur.execute('DELETE FROM midterm_exams WHERE exam_id = ?', (exam_id,))
        conn.commit()
        conn.close()
        
        return jsonify({'status': 'success', 'message': 'Exam deleted successfully'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

# Question Management Routes
@app.route("/api/exams/<int:exam_id>/questions", methods=['GET'])
@login_required
def get_questions(exam_id):
    """Get all questions for an exam"""
    try:
        try:
            page = int(request.args.get('page', 1))
        except (TypeError, ValueError):
            page = 1
        page = max(page, 1)

        try:
            per_page = int(request.args.get('per_page', 10))
        except (TypeError, ValueError):
            per_page = 10
        per_page = 10 if per_page <= 0 else min(per_page, 50)
        offset = (page - 1) * per_page

        conn = get_connection()
        cur = conn.cursor()

        total = cur.execute(
            'SELECT COUNT(*) as cnt FROM midterm_questions WHERE exam_id = ?',
            (exam_id,)
        ).fetchone()[0]
        selected_total = cur.execute(
            'SELECT COUNT(*) as cnt FROM midterm_questions WHERE exam_id = ? AND is_selected = 1',
            (exam_id,)
        ).fetchone()[0]

        cur.execute(
            '''
            SELECT * FROM midterm_questions
            WHERE exam_id = ?
            ORDER BY question_id DESC
            LIMIT ? OFFSET ?
            ''',
            (exam_id, per_page, offset)
        )
        questions = [dict(row) for row in cur.fetchall()]
        conn.close()

        total_pages = (total + per_page - 1) // per_page if total else 0

        return jsonify({
            'status': 'success',
            'data': questions,
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': total,
                'total_pages': total_pages
            },
            'selection_summary': {
                'selected': selected_total,
                'total': total
            }
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route("/api/questions/<int:question_id>/selection", methods=['PATCH'])
@login_required
def update_question_selection(question_id):
    """Toggle selection flag on a question"""
    try:
        role = session.get('role')
        teacher_id = session.get('teacher_id')
        if role not in ['admin', 'teacher']:
            return jsonify({'status': 'error', 'message': 'Unauthorized'}), 403

        data = request.get_json() or {}
        is_selected = 1 if data.get('is_selected') else 0

        conn = get_connection()
        cur = conn.cursor()

        question = cur.execute(
            'SELECT exam_id FROM midterm_questions WHERE question_id = ?',
            (question_id,)
        ).fetchone()

        if not question:
            conn.close()
            return jsonify({'status': 'error', 'message': 'Question not found'}), 404

        if role == 'teacher':
            exam = cur.execute(
                'SELECT created_by FROM midterm_exams WHERE exam_id = ?',
                (question['exam_id'],)
            ).fetchone()
            if exam and exam['created_by'] and exam['created_by'] != teacher_id:
                conn.close()
                return jsonify({'status': 'error', 'message': 'Unauthorized to update this question'}), 403

        cur.execute(
            'UPDATE midterm_questions SET is_selected = ? WHERE question_id = ?',
            (is_selected, question_id)
        )
        conn.commit()
        conn.close()

        return jsonify({
            'status': 'success',
            'message': 'Question selection updated',
            'question_id': question_id,
            'is_selected': bool(is_selected)
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route("/api/questions/<int:question_id>", methods=['PUT'])
@login_required
def update_question(question_id):
    """Update an existing question"""
    try:
        role = session.get('role')
        teacher_id = session.get('teacher_id')
        data = request.get_json() or {}

        conn = get_connection()
        cur = conn.cursor()

        question = cur.execute(
            'SELECT * FROM midterm_questions WHERE question_id = ?',
            (question_id,)
        ).fetchone()

        if not question:
            conn.close()
            return jsonify({'status': 'error', 'message': 'Question not found'}), 404

        exam = cur.execute(
            'SELECT created_by FROM midterm_exams WHERE exam_id = ?',
            (question['exam_id'],)
        ).fetchone()

        if role == 'teacher' and exam and exam['created_by'] and exam['created_by'] != teacher_id:
            conn.close()
            return jsonify({'status': 'error', 'message': 'Unauthorized'}), 403

        correct_option = (data.get('correct_option') or question['correct_option']).upper()
        correct_index_map = {'A': 0, 'B': 1, 'C': 2, 'D': 3}
        correct_index = correct_index_map.get(correct_option, question['correct_index'])

        cur.execute('''
            UPDATE midterm_questions
            SET question_text = ?, option_a = ?, option_b = ?, option_c = ?, option_d = ?,
                correct_option = ?, correct_index = ?, marks = ?, difficulty = ?, topic = ?, media_path = ?
            WHERE question_id = ?
        ''', (
            data.get('question_text', question['question_text']),
            data.get('option_a', question['option_a']),
            data.get('option_b', question['option_b']),
            data.get('option_c', question['option_c']),
            data.get('option_d', question['option_d']),
            correct_option,
            correct_index,
            data.get('marks', question['marks']),
            data.get('difficulty', question['difficulty']),
            data.get('topic', question['topic']),
            data.get('media_path', question['media_path']),
            question_id
        ))

        conn.commit()
        conn.close()

        return jsonify({'status': 'success', 'message': 'Question updated successfully'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route("/api/exams/<int:exam_id>/questions", methods=['POST'])
@login_required
def create_question(exam_id):
    """Create a new question"""
    try:
        data = request.get_json()
        conn = get_connection()
        cur = conn.cursor()
        
        # Determine correct_index from correct_option
        correct_option = data.get('correct_option', '').upper()
        correct_index_map = {'A': 0, 'B': 1, 'C': 2, 'D': 3}
        correct_index = correct_index_map.get(correct_option, 0)
        
        cur.execute('''
            INSERT INTO midterm_questions (
                exam_id, question_text, option_a, option_b, option_c, option_d,
                correct_option, correct_index, marks, media_path, difficulty, topic, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            exam_id, data.get('question_text'), data.get('option_a'),
            data.get('option_b'), data.get('option_c'), data.get('option_d'),
            correct_option, correct_index, data.get('marks', 1.0),
            data.get('media_path'), data.get('difficulty'), data.get('topic'),
            datetime.now().isoformat()
        ))
        
        question_id = cur.lastrowid
        conn.commit()
        conn.close()
        
        return jsonify({'status': 'success', 'message': 'Question created successfully', 'question_id': question_id})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route("/api/exams/<int:exam_id>/questions/count", methods=['GET'])
@login_required
def get_question_count(exam_id):
    """Get count of questions for an exam"""
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute('SELECT COUNT(*) as count FROM midterm_questions WHERE exam_id = ?', (exam_id,))
        result = cur.fetchone()
        conn.close()
        return jsonify({'status': 'success', 'count': result['count'] if result else 0})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route("/api/exams/<int:exam_id>/questions/bulk", methods=['POST'])
@login_required
def bulk_upload_questions(exam_id):
    """Bulk upload questions from Excel/CSV"""
    try:
        if 'file' not in request.files:
            return jsonify({'status': 'error', 'message': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'status': 'error', 'message': 'No file selected'}), 400
        
        # Read Excel file
        df = pd.read_excel(file)
        
        # Normalize column names: strip whitespace and convert to lowercase
        df.columns = [col.strip().lower() if isinstance(col, str) else col for col in df.columns]
        
        # Expected columns (normalized)
        required_columns = ['question', 'optiona', 'optionb', 'optionc', 'optiond', 'correctoption']
        
        # Check if all required columns are present after normalization
        if not all(col in df.columns for col in required_columns):
            missing_cols = [col for col in required_columns if col not in df.columns]
            return jsonify({'status': 'error', 'message': f'Missing required columns: {", ".join(missing_cols)}'}), 400
        
        conn = get_connection()
        cur = conn.cursor()
        
        correct_index_map = {'A': 0, 'B': 1, 'C': 2, 'D': 3} # Keep original for mapping
        success_count = 0
        
        for _, row in df.iterrows():
            try:
                # Access columns using their normalized names
                correct_option_raw = str(row['correctoption']).strip()
                correct_option = correct_option_raw.upper()
                correct_index = correct_index_map.get(correct_option, 0)
                
                cur.execute('''
                    INSERT INTO midterm_questions (
                        exam_id, question_text, option_a, option_b, option_c, option_d,
                        correct_option, correct_index, marks, media_path, created_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    exam_id, str(row['question']), str(row['optiona']),
                    str(row['optionb']), str(row['optionc']), str(row['optiond']),
                    correct_option, correct_index, float(row.get('marks', 1.0)),
                    str(row.get('media_link', '')), datetime.now().isoformat()
                ))
                success_count += 1
            except Exception as e:
                print(f"Error importing question: {e}")
                continue
        
        conn.commit()
        conn.close()
        
        return jsonify({'status': 'success', 'message': f'{success_count} questions imported successfully'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

# Student Exam Instance Routes
@app.route("/api/exams/<int:exam_id>/start", methods=['POST'])
def start_exam(exam_id):
    """Start exam for student - creates instance"""
    try:
        # Check if student is logged in
        if 'student_logged_in' not in session or not session['student_logged_in']:
            return jsonify({'status': 'error', 'message': 'Student login required'}), 401
        
        student_id = session.get('student_id')
        conn = get_connection()
        cur = conn.cursor()
        
        # Check if exam exists and is published
        exam = cur.execute('SELECT * FROM midterm_exams WHERE exam_id = ?', (exam_id,)).fetchone()
        if not exam:
            return jsonify({'status': 'error', 'message': 'Exam not found'}), 404
        
        if exam['status'] != 'Published':
            return jsonify({'status': 'error', 'message': 'Exam is not available'}), 400
        
        # Check if instance already exists
        existing = cur.execute(
            'SELECT * FROM midterm_instances WHERE exam_id = ? AND student_id = ?',
            (exam_id, student_id)
        ).fetchone()
        
        if existing:
            instance_id = existing['instance_id']
            # Return existing instance if not completed
            if existing['status'] != 'Completed':
                return jsonify({
                    'status': 'success',
                    'instance_id': instance_id,
                    'message': 'Resuming exam',
                    'duration': exam['duration'],
                    'exam_title': exam['title'],
                    'resume': True
                })
        
        # Create new instance
        import secrets
        token = secrets.token_urlsafe(32)
        ip_address = request.remote_addr
        
        cur.execute('''
            INSERT INTO midterm_instances (exam_id, student_id, start_time, status, token, ip_address, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (exam_id, student_id, datetime.now().isoformat(), 'In Progress', token, ip_address, datetime.now().isoformat()))
        
        instance_id = cur.lastrowid
        
        # Create exam attendance record
        cur.execute('''
            INSERT OR REPLACE INTO exam_attendance (exam_id, student_id, status, start_time, created_at)
            VALUES (?, ?, ?, ?, ?)
        ''', (exam_id, student_id, 'Present', datetime.now().isoformat(), datetime.now().isoformat()))
        
        # Load only selected questions when available
        selected_questions = cur.execute(
            'SELECT * FROM midterm_questions WHERE exam_id = ? AND is_selected = 1 ORDER BY question_id',
            (exam_id,)
        ).fetchall()

        if selected_questions:
            question_list = list(selected_questions)
        else:
            question_list = cur.execute(
                'SELECT * FROM midterm_questions WHERE exam_id = ? ORDER BY question_id',
                (exam_id,)
            ).fetchall()

        if not question_list:
            return jsonify({'status': 'error', 'message': 'No questions found for this exam'}), 400
        
        # Randomize questions if enabled
        import random
        if exam['randomize_questions']:
            random.shuffle(question_list)
        
        # Select total_questions
        total_questions = min(exam['total_questions'], len(question_list))
        selected_questions = question_list[:total_questions]
        
        # Create instance questions
        for idx, question in enumerate(selected_questions):
            option_order = [0, 1, 2, 3]
            if exam['randomize_options']:
                random.shuffle(option_order)
            
            cur.execute('''
                INSERT INTO midterm_instance_questions (instance_id, question_id, question_order, option_order_json, created_at)
                VALUES (?, ?, ?, ?, ?)
            ''', (instance_id, question['question_id'], idx + 1, json.dumps(option_order), datetime.now().isoformat()))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'status': 'success',
            'instance_id': instance_id,
            'token': token,
            'duration': exam['duration'],
            'total_questions': total_questions,
            'exam_title': exam['title'],
            'resume': False
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route("/api/exams/instances/<int:instance_id>/questions", methods=['GET'])
def get_instance_questions(instance_id):
    """Get questions for an exam instance"""
    try:
        student_id = session.get('student_id')
        if not student_id:
            return jsonify({'status': 'error', 'message': 'Unauthorized'}), 401
        
        conn = get_connection()
        cur = conn.cursor()
        
        # Verify instance belongs to student
        instance = cur.execute(
            'SELECT * FROM midterm_instances WHERE instance_id = ? AND student_id = ?',
            (instance_id, student_id)
        ).fetchone()
        
        if not instance:
            return jsonify({'status': 'error', 'message': 'Instance not found'}), 404
        
        # Get questions with their order
        cur.execute('''
            SELECT iq.question_order,
                   iq.option_order_json,
                   q.*,
                   mr.selected_index AS saved_selected_index,
                   mr.selected_option AS saved_selected_option,
                   sa.status AS answer_status
            FROM midterm_instance_questions iq
            JOIN midterm_questions q ON iq.question_id = q.question_id
            LEFT JOIN midterm_responses mr
                ON mr.instance_id = iq.instance_id AND mr.question_id = iq.question_id
            LEFT JOIN student_answers sa
                ON sa.instance_id = iq.instance_id AND sa.question_id = iq.question_id AND sa.student_id = ?
            WHERE iq.instance_id = ?
            ORDER BY iq.question_order
        ''', (student_id, instance_id))
        
        questions = []
        for row in cur.fetchall():
            row_dict = dict(row)
            option_order = json.loads(row_dict['option_order_json'])
            options = [
                row_dict['option_a'], row_dict['option_b'],
                row_dict['option_c'], row_dict['option_d']
            ]
            
            # Reorder options based on option_order
            reordered_options = [options[i] for i in option_order]
            
            questions.append({
                'question_id': row_dict['question_id'],
                'question_order': row_dict['question_order'],
                'question_text': row_dict['question_text'],
                'options': reordered_options,  # Don't expose correct answer
                'media_path': row_dict.get('media_path'),
                'marks': row_dict['marks'],
                'saved_selected_index': row_dict.get('saved_selected_index'),
                'saved_selected_option': row_dict.get('saved_selected_option'),
                'answer_status': row_dict.get('answer_status')
            })
        
        conn.close()
        
        return jsonify({'status': 'success', 'data': questions})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route("/api/exams/instances/<int:instance_id>/responses", methods=['POST'])
def save_response(instance_id):
    """Save student response to a question"""
    try:
        student_id = session.get('student_id')
        if not student_id:
            return jsonify({'status': 'error', 'message': 'Unauthorized'}), 401
        
        data = request.get_json()
        question_id = data.get('question_id')
        selected_option = data.get('selected_option')
        selected_index = data.get('selected_index')
        status = (data.get('status') or 'answered').lower()
        if status not in ('answered', 'skipped'):
            status = 'answered'
        
        if not question_id:
            return jsonify({'status': 'error', 'message': 'Question ID is required'}), 400
        
        if status == 'answered' and selected_index is None:
            return jsonify({'status': 'error', 'message': 'Please select an answer before saving.'}), 400
        
        conn = get_connection()
        cur = conn.cursor()
        
        # Verify instance
        instance = cur.execute(
            'SELECT * FROM midterm_instances WHERE instance_id = ? AND student_id = ?',
            (instance_id, student_id)
        ).fetchone()
        
        if not instance or instance['status'] == 'Completed':
            conn.close()
            return jsonify({'status': 'error', 'message': 'Invalid instance'}), 400

        exam_id = instance['exam_id']
        timestamp = datetime.now().isoformat()

        # Track answer status for skip/answer review
        cur.execute('''
            INSERT INTO student_answers (
                student_id, exam_id, instance_id, question_id,
                selected_option, selected_index, status, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(student_id, exam_id, question_id)
            DO UPDATE SET
                selected_option = excluded.selected_option,
                selected_index = excluded.selected_index,
                status = excluded.status,
                updated_at = excluded.updated_at,
                instance_id = excluded.instance_id
        ''', (
            student_id, exam_id, instance_id, question_id,
            selected_option, selected_index, status, timestamp
        ))

        if status == 'skipped':
            cur.execute(
                'DELETE FROM midterm_responses WHERE instance_id = ? AND question_id = ?',
                (instance_id, question_id)
            )
            conn.commit()
            conn.close()
            return jsonify({'status': 'success', 'message': 'Question skipped'})

        # Get correct answer (server-side only)
        question = cur.execute(
            '''
                SELECT q.correct_index, iq.option_order_json
                FROM midterm_questions q
                JOIN midterm_instance_questions iq ON iq.question_id = q.question_id
                WHERE q.question_id = ? AND iq.instance_id = ?
            ''',
            (question_id, instance_id)
        ).fetchone()

        if not question:
            conn.close()
            return jsonify({'status': 'error', 'message': 'Question not found'}), 404

        exam = cur.execute(
            'SELECT marks_per_question, negative_marking, negative_marks_value FROM midterm_exams WHERE exam_id = ?',
            (exam_id,)
        ).fetchone()

        option_order = json.loads(question['option_order_json']) if question['option_order_json'] else [0, 1, 2, 3]
        correct_index = question['correct_index']

        is_correct = 0
        marks_obtained = 0.0

        try:
            mapped_selected = option_order.index(selected_index) if selected_index is not None and selected_index < len(option_order) else -1
            mapped_correct = option_order.index(correct_index) if correct_index < len(option_order) else -1
            if mapped_selected == mapped_correct:
                is_correct = 1
                marks_obtained = exam['marks_per_question'] if exam else 1.0
            elif exam and exam['negative_marking']:
                marks_obtained = -(exam['negative_marks_value'] if exam else 0.0)
        except (ValueError, IndexError):
            if selected_index == correct_index:
                is_correct = 1
                marks_obtained = exam['marks_per_question'] if exam else 1.0
            elif exam and exam['negative_marking']:
                marks_obtained = -(exam['negative_marks_value'] if exam else 0.0)

        cur.execute('''
            INSERT OR REPLACE INTO midterm_responses (
                instance_id, question_id, selected_option, selected_index,
                is_correct, marks_obtained, timestamp
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (instance_id, question_id, selected_option, selected_index, is_correct, marks_obtained, datetime.now().isoformat()))
        
        conn.commit()
        conn.close()
        
        return jsonify({'status': 'success', 'message': 'Response saved'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route("/api/exams/instances/<int:instance_id>/submit", methods=['POST'])
def submit_exam(instance_id):
    """Submit exam and calculate results"""
    try:
        student_id = session.get('student_id')
        if not student_id:
            return jsonify({'status': 'error', 'message': 'Unauthorized'}), 401
        
        conn = get_connection()
        cur = conn.cursor()
        
        # Verify instance
        instance = cur.execute(
            'SELECT * FROM midterm_instances WHERE instance_id = ? AND student_id = ?',
            (instance_id, student_id)
        ).fetchone()
        
        if not instance:
            return jsonify({'status': 'error', 'message': 'Instance not found'}), 404
        
        if instance['status'] == 'Completed':
            return jsonify({'status': 'error', 'message': 'Exam already submitted'}), 400

        skipped_remaining = cur.execute(
            '''
                SELECT COUNT(*) FROM student_answers
                WHERE student_id = ? AND exam_id = ? AND status = 'skipped'
            ''',
            (student_id, instance['exam_id'])
        ).fetchone()[0]

        if skipped_remaining:
            conn.close()
            return jsonify({
                'status': 'error',
                'message': f'You have {skipped_remaining} skipped question(s). Please answer them before submitting.'
            }), 400

        total_questions = cur.execute(
            'SELECT COUNT(*) FROM midterm_instance_questions WHERE instance_id = ?',
            (instance_id,)
        ).fetchone()[0]
        answered_questions = cur.execute(
            'SELECT COUNT(*) FROM midterm_responses WHERE instance_id = ?',
            (instance_id,)
        ).fetchone()[0]

        if answered_questions < total_questions:
            remaining = total_questions - answered_questions
            conn.close()
            return jsonify({
                'status': 'error',
                'message': f'You still have {remaining} unanswered question(s). Please answer them before submitting.'
            }), 400
        
        # Calculate results
        responses = cur.execute(
            'SELECT * FROM midterm_responses WHERE instance_id = ?',
            (instance_id,)
        ).fetchall()
        
        total_marks = 0.0
        for response in responses:
            total_marks += response['marks_obtained']
        
        # Get exam details
        exam_row = cur.execute(
            'SELECT * FROM midterm_exams WHERE exam_id = ?',
            (instance['exam_id'],)
        ).fetchone()
        
        if not exam_row:
            conn.close()
            return jsonify({'status': 'error', 'message': 'Exam configuration not found'}), 404

        exam = dict(exam_row)
        
        max_marks = exam['total_questions'] * exam['marks_per_question']
        percentage = (total_marks / max_marks * 100) if max_marks > 0 else 0
        
        # Determine grade
        passing_threshold = exam['passing_marks'] if exam and 'passing_marks' in exam.keys() else None
        grade = calculate_grade_from_percentage(percentage, passing_threshold)
        
        # Update instance
        cur.execute('''
            UPDATE midterm_instances SET status = 'Completed', end_time = ?
            WHERE instance_id = ?
        ''', (datetime.now().isoformat(), instance_id))
        
        # Update attendance
        cur.execute('''
            UPDATE exam_attendance SET end_time = ?
            WHERE exam_id = ? AND student_id = ?
        ''', (datetime.now().isoformat(), instance['exam_id'], student_id))
        
        # Save result
        cur.execute('''
            INSERT OR REPLACE INTO midterm_results (
                instance_id, exam_id, student_id, total_marks, obtained_marks,
                percentage, grade, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            instance_id, instance['exam_id'], student_id, max_marks, total_marks,
            percentage, grade, datetime.now().isoformat(), datetime.now().isoformat()
        ))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'status': 'success',
            'message': 'Exam submitted successfully',
            'result': {
                'total_marks': max_marks,
                'obtained_marks': total_marks,
                'percentage': percentage,
                'grade': grade
            }
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

# Proctoring Routes
@app.route("/api/exams/instances/<int:instance_id>/heartbeat", methods=['POST'])
def heartbeat(instance_id):
    """Heartbeat for proctoring"""
    try:
        student_id = session.get('student_id')
        if not student_id:
            return jsonify({'status': 'error', 'message': 'Unauthorized'}), 401
        
        data = request.get_json()
        event_type = data.get('event_type', 'heartbeat')
        details = data.get('details', '')
        
        conn = get_connection()
        cur = conn.cursor()
        
        # Verify instance
        instance = cur.execute(
            'SELECT * FROM midterm_instances WHERE instance_id = ? AND student_id = ?',
            (instance_id, student_id)
        ).fetchone()
        
        if not instance:
            return jsonify({'status': 'error', 'message': 'Instance not found'}), 404
        
        # Log event
        cur.execute('''
            INSERT INTO exam_proctor_logs (instance_id, student_id, exam_id, event_type, details, timestamp)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (instance_id, student_id, instance['exam_id'], event_type, json.dumps(details), datetime.now().isoformat()))
        
        conn.commit()
        conn.close()
        
        return jsonify({'status': 'success'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

# Results & Attendance Routes
@app.route("/api/exams/<int:exam_id>/attendance", methods=['GET'])
@login_required
def get_exam_attendance(exam_id):
    """Get exam attendance"""
    try:
        conn = get_connection()
        cur = conn.cursor()
        
        cur.execute('''
            SELECT ea.*, s.name, s.admission_no, s.campus, s.semester
            FROM exam_attendance ea
            JOIN students s ON ea.student_id = s.id
            WHERE ea.exam_id = ?
            ORDER BY s.name
        ''', (exam_id,))
        
        attendance = [dict(row) for row in cur.fetchall()]
        conn.close()
        
        return jsonify({'status': 'success', 'data': attendance})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route("/api/exams/<int:exam_id>/attendance/students", methods=['GET'])
@login_required
def get_exam_attendance_students(exam_id):
    """Get students eligible for exam attendance marking"""
    try:
        role = session.get('role')
        teacher_id = session.get('teacher_id')

        conn = get_connection()
        cur = conn.cursor()

        exam = cur.execute('SELECT * FROM midterm_exams WHERE exam_id = ?', (exam_id,)).fetchone()
        if not exam:
            conn.close()
            return jsonify({'status': 'error', 'message': 'Exam not found'}), 404

        if role == 'teacher' and exam['created_by'] != teacher_id:
            conn.close()
            return jsonify({'status': 'error', 'message': 'Unauthorized'}), 403

        query = '''
            SELECT s.id AS student_id, s.name, s.admission_no, s.campus, s.board, s.technology, s.semester,
                   COALESCE(ea.status, 'Not Marked') AS status, ea.start_time, ea.end_time
            FROM students s
            LEFT JOIN exam_attendance ea ON ea.student_id = s.id AND ea.exam_id = ?
            WHERE 1=1
        '''
        params = [exam_id]
        if exam['campus']:
            query += ' AND s.campus = ?'
            params.append(exam['campus'])
        if exam['board']:
            query += ' AND s.board = ?'
            params.append(exam['board'])
        if exam['technology']:
            query += ' AND s.technology = ?'
            params.append(exam['technology'])
        if exam['semester']:
            query += ' AND s.semester = ?'
            params.append(exam['semester'])

        query += ' ORDER BY s.name'
        cur.execute(query, params)
        students = [dict(row) for row in cur.fetchall()]
        conn.close()

        return jsonify({'status': 'success', 'data': students})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route("/api/exams/<int:exam_id>/attendance", methods=['POST'])
@login_required
def save_exam_attendance_records(exam_id):
    """Save manual exam attendance"""
    try:
        role = session.get('role')
        teacher_id = session.get('teacher_id')
        data = request.get_json() or {}
        records = data.get('records', [])

        if not isinstance(records, list) or not records:
            return jsonify({'status': 'error', 'message': 'No attendance records provided'}), 400

        conn = get_connection()
        cur = conn.cursor()

        exam = cur.execute('SELECT * FROM midterm_exams WHERE exam_id = ?', (exam_id,)).fetchone()
        if not exam:
            conn.close()
            return jsonify({'status': 'error', 'message': 'Exam not found'}), 404

        if role == 'teacher' and exam['created_by'] != teacher_id:
            conn.close()
            return jsonify({'status': 'error', 'message': 'Unauthorized'}), 403

        now = datetime.now().isoformat()
        saved = 0
        for record in records:
            student_id = record.get('student_id')
            status = record.get('status')
            if not student_id or status not in ['Present', 'Absent']:
                continue
            start_time = record.get('start_time')
            end_time = record.get('end_time')
            if not start_time and status == 'Present':
                start_time = now

            cur.execute('''
                INSERT INTO exam_attendance (exam_id, student_id, status, start_time, end_time, created_at)
                VALUES (?, ?, ?, ?, ?, ?)
                ON CONFLICT(exam_id, student_id) DO UPDATE SET
                    status = excluded.status,
                    start_time = COALESCE(excluded.start_time, exam_attendance.start_time),
                    end_time = COALESCE(excluded.end_time, exam_attendance.end_time)
            ''', (exam_id, student_id, status, start_time, end_time, now))
            saved += 1

        conn.commit()
        conn.close()

        return jsonify({'status': 'success', 'message': f'Attendance saved for {saved} students.'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route("/api/exams/<int:exam_id>/results", methods=['GET'])
@login_required
def get_exam_results(exam_id):
    """Get exam results"""
    try:
        campus = request.args.get('campus', '').strip()
        technology = request.args.get('technology', '').strip()
        role = session.get('role')
        teacher_id = session.get('teacher_id')
        
        conn = get_connection()
        cur = conn.cursor()
        
        # Check permissions
        exam = cur.execute('SELECT * FROM midterm_exams WHERE exam_id = ?', (exam_id,)).fetchone()
        if not exam:
            return jsonify({'status': 'error', 'message': 'Exam not found'}), 404
        
        if role == 'teacher':
            permissions = set(session.get('permissions') or [])
            owns_exam = (exam['created_by'] == teacher_id)
            can_view = 'view_results' in permissions or 'create_exam' in permissions
            if not owns_exam and not can_view:
                return jsonify({'status': 'error', 'message': 'Unauthorized'}), 403
        
        results = fetch_exam_results_rows(cur, exam_id, campus or None, technology or None)
        conn.close()
        
        return jsonify({'status': 'success', 'data': results})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route("/api/exams/<int:exam_id>/results/export_pdf", methods=['GET'])
@login_required
def export_exam_results_pdf(exam_id):
    """Generate a PDF report for an exam's results."""
    campus = request.args.get('campus', '').strip()
    technology = request.args.get('technology', '').strip()
    role = session.get('role')
    teacher_id = session.get('teacher_id')

    conn = get_connection()
    try:
        cur = conn.cursor()
        exam = cur.execute('SELECT exam_id, title, subject, created_by FROM midterm_exams WHERE exam_id = ?', (exam_id,)).fetchone()
        if not exam:
            return jsonify({'status': 'error', 'message': 'Exam not found'}), 404
        if role == 'teacher' and exam['created_by'] != teacher_id:
            return jsonify({'status': 'error', 'message': 'Unauthorized'}), 403

        results = fetch_exam_results_rows(cur, exam_id, campus or None, technology or None)
        generated_at = datetime.now().strftime('%Y-%m-%d %H:%M')

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=30, rightMargin=30, topMargin=25, bottomMargin=25)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('ExamReportTitle', parent=styles['Title'], alignment=1, fontSize=18, textColor=colors.HexColor('#0a7b35'), spaceAfter=6)
        subtitle_style = ParagraphStyle('ExamReportSubtitle', parent=styles['Heading2'], alignment=1, fontSize=12, textColor=colors.HexColor('#111111'), spaceAfter=14)
        meta_style = ParagraphStyle('ExamReportMeta', parent=styles['Normal'], alignment=1, fontSize=9, textColor=colors.HexColor('#555555'), spaceAfter=12)

        elements = [
            Paragraph('GHAZALI INSTITUTE OF MEDICAL SCIENCES', title_style),
            Paragraph(f'Exam Report — {exam["title"] or exam["subject"] or "Exam"}', subtitle_style),
            Paragraph(f'Generated: {generated_at}', meta_style)
        ]

        filters_text = []
        if campus:
            filters_text.append(f'Campus: {campus}')
        if technology:
            filters_text.append(f'Technology: {technology}')
        if filters_text:
            elements.append(Paragraph('Filters — ' + ', '.join(filters_text), ParagraphStyle(
                'ExamReportFilters',
                parent=styles['Normal'],
                fontSize=9,
                textColor=colors.HexColor('#666666'),
                spaceAfter=10
            )))

        total_students = len(results)
        pass_count = sum(1 for row in results if (row.get('grade') or '').upper() != 'F')
        fail_count = total_students - pass_count
        avg_percentage = round(
            sum(float(row.get('percentage') or 0) for row in results) / total_students, 2
        ) if total_students else 0.0

        summary_data = [
            ['Total Students', str(total_students)],
            ['Passed', str(pass_count)],
            ['Failed', str(fail_count)],
            ['Average %', f'{avg_percentage:.2f}%']
        ]
        summary_table = Table(summary_data, colWidths=[1.2 * inch, 1.0 * inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0c2d48')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9fa')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 0.2 * inch))

        table_data = [['#', 'Admission No', 'Student Name', 'Campus', 'Technology', 'Obtained', 'Total', '%', 'Grade']]
        for idx, row in enumerate(results, start=1):
            obtained_marks = row.get('obtained_marks') or 0
            total_marks = row.get('total_marks') or 0
            percentage = row.get('percentage') or 0
            table_data.append([
                str(idx),
                row.get('admission_no') or '',
                row.get('name') or '',
                row.get('campus') or '',
                row.get('technology') or '',
                str(obtained_marks),
                str(total_marks),
                f'{float(percentage):.2f}%',
                row.get('grade') or ''
            ])

        results_table = Table(table_data, repeatRows=1, colWidths=[0.4*inch, 1*inch, 1.6*inch, 1*inch, 1.2*inch, 0.9*inch, 0.9*inch, 0.7*inch, 0.7*inch])
        results_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0c2d48')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (5, 1), (8, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f6f7fb')]),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#d7dce6')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
        ]))
        elements.append(results_table)

        doc.build(elements)
        buffer.seek(0)

        filename = f'exam_report_{exam_id}.pdf'
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        conn.close()

@app.route("/api/exams/<int:exam_id>/publish", methods=['POST'])
@login_required
def publish_exam(exam_id):
    """Publish exam to make it available to students"""
    try:
        conn = get_connection()
        cur = conn.cursor()
        
        # Get exam details
        cur.execute('SELECT * FROM midterm_exams WHERE exam_id = ?', (exam_id,))
        exam = cur.fetchone()
        if not exam:
            conn.close()
            return jsonify({'status': 'error', 'message': 'Exam not found'}), 404
        
        # Check if exam has at least one question
        cur.execute('SELECT COUNT(*) as count FROM midterm_questions WHERE exam_id = ?', (exam_id,))
        question_count = cur.fetchone()['count']
        
        if question_count == 0:
            conn.close()
            return jsonify({'status': 'error', 'message': 'Cannot publish exam. Please upload at least one question first.'}), 400
        
        # Update exam status to Published
        cur.execute('''
            UPDATE midterm_exams 
            SET status = 'Published', updated_at = ?
            WHERE exam_id = ?
        ''', (datetime.now().isoformat(), exam_id))
        
        conn.commit()
        conn.close()
        return jsonify({'status': 'success', 'message': 'Exam published successfully! It is now available to students.'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route("/api/exams/<int:exam_id>/publish-results", methods=['POST'])
@login_required
def publish_results(exam_id):
    """Publish exam results"""
    try:
        role = session.get('role')
        if role != 'admin':
            return jsonify({'status': 'error', 'message': 'Admin only'}), 403
        
        teacher_id = session.get('teacher_id')
        conn = get_connection()
        cur = conn.cursor()
        
        cur.execute('''
            UPDATE midterm_results SET published_flag = 1, published_at = ?, published_by = ?
            WHERE exam_id = ?
        ''', (datetime.now().isoformat(), teacher_id, exam_id))
        
        conn.commit()
        conn.close()
        
        return jsonify({'status': 'success', 'message': 'Results published successfully'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route("/api/dmc/generate", methods=['POST'])
@admin_required
def api_generate_dmc():
    data = request.get_json() or {}
    student_id = data.get('student_id')
    exam_ids = data.get('exam_ids') or []

    if not student_id:
        return jsonify({'status': 'error', 'message': 'Student ID is required.'}), 400
    if not isinstance(exam_ids, list) or not exam_ids:
        return jsonify({'status': 'error', 'message': 'Select at least one subject/exam.'}), 400

    exam_ids = [int(exam_id) for exam_id in exam_ids if str(exam_id).strip().isdigit()]
    if not exam_ids:
        return jsonify({'status': 'error', 'message': 'Invalid exam selection.'}), 400

    conn = get_connection()
    try:
        student = conn.execute(
            'SELECT id, name, father_name, admission_no, technology, semester, campus FROM students WHERE id = ?',
            (student_id,)
        ).fetchone()
        if not student:
            return jsonify({'status': 'error', 'message': 'Student not found.'}), 404

        subject_rows = []
        total_obtained = 0.0
        total_marks = 0.0
        now = datetime.now().isoformat()

        for exam_id in exam_ids:
            exam = conn.execute(
                'SELECT exam_id, title, subject, exam_date, passing_marks FROM midterm_exams WHERE exam_id = ?',
                (exam_id,)
            ).fetchone()
            result = conn.execute(
                'SELECT * FROM midterm_results WHERE exam_id = ? AND student_id = ?',
                (exam_id, student_id)
            ).fetchone()
            if not exam or not result:
                return jsonify({'status': 'error', 'message': f'Results not found for exam ID {exam_id}.'}), 404

            subject_label = exam['subject'] or exam['title'] or f'Exam {exam_id}'
            obtained = result['obtained_marks'] or 0.0
            total = result['total_marks'] or 0.0
            percentage = result['percentage'] or (obtained / total * 100 if total else 0)

            subject_rows.append({
                'subject': subject_label,
                'total_marks': total,
                'obtained_marks': obtained,
                'percentage': percentage,
                'grade': result.get('grade') or calculate_grade_from_percentage(percentage),
                'exam_date': exam.get('exam_date')
            })

            total_obtained += obtained
            total_marks += total

        overall_percentage = (total_obtained / total_marks * 100) if total_marks else 0
        overall_grade = calculate_grade_from_percentage(overall_percentage)
        result_status = 'Pass' if overall_percentage >= 50 else 'Fail'
        summary = {
            'obtained_marks': total_obtained,
            'total_marks': total_marks,
            'percentage': overall_percentage,
            'grade': overall_grade,
            'result_status': result_status
        }

        cursor = conn.cursor()
        cursor.execute(
            '''
                INSERT INTO dmc_records (student_id, exam_ids, subject_rows_json, summary_json, created_by, created_at)
                VALUES (?, ?, ?, ?, ?, ?)
            ''',
            (
                student_id,
                json.dumps(exam_ids),
                json.dumps(subject_rows),
                json.dumps(summary),
                session.get('teacher_id'),
                now
            )
        )
        record_id = cursor.lastrowid
        conn.commit()
        return jsonify({
            'status': 'success',
            'message': 'DMC generated successfully.',
            'dmc_id': record_id,
            'url': url_for('view_dmc_record', dmc_id=record_id)
        })
    except Exception as e:
        conn.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        conn.close()

@app.route("/dmc/<int:exam_id>/<int:student_id>")
def view_dmc_report(exam_id, student_id):
    """Render a printable Detailed Marks Certificate (DMC)"""
    if not can_view_student_dmc(student_id):
        abort(403)

    conn = get_connection()
    try:
        student_row = conn.execute(
            '''
                SELECT id, name, father_name, admission_no, technology, semester, campus
                FROM students WHERE id = ?
            ''',
            (student_id,)
        ).fetchone()

        exam_row = conn.execute(
            '''
                SELECT exam_id, title, subject, exam_date, semester, technology, passing_marks
                FROM midterm_exams WHERE exam_id = ?
            ''',
            (exam_id,)
        ).fetchone()

        result_row = conn.execute(
            '''
                SELECT * FROM midterm_results
                WHERE exam_id = ? AND student_id = ?
            ''',
            (exam_id, student_id)
        ).fetchone()

        student = dict(student_row) if student_row else None
        exam = dict(exam_row) if exam_row else None
        result = dict(result_row) if result_row else None

        error_message = None
        if not student:
            error_message = 'Student record not found.'
        elif not exam:
            error_message = 'Exam record not found.'
        elif not result:
            error_message = 'DMC not available at the moment. Please check later or contact admin.'

        subject_rows = []
        result_status = None
        result_summary = None
        selected_subjects = None
        if not error_message:
            subject_label = exam.get('subject') or exam.get('title') or 'Subject'
            selected_subjects = subject_label
            subject_rows.append({
                'subject': subject_label,
                'total_marks': result['total_marks'],
                'obtained_marks': result['obtained_marks'],
                'percentage': result['percentage'],
                'grade': result.get('grade') or calculate_grade_from_percentage(result.get('percentage'))
            })

            passing_marks = exam.get('passing_marks')
            result_status = determine_result_status(result['obtained_marks'], result['total_marks'], passing_marks)
            result_summary = {
                'obtained_marks': result['obtained_marks'],
                'total_marks': result['total_marks'],
                'percentage': result['percentage'],
                'grade': result.get('grade') or calculate_grade_from_percentage(result.get('percentage'), passing_marks),
                'result_status': result_status
            }

        generated_on = datetime.now()
        return render_template(
            'dmc.html',
            student=student,
            exam=exam,
            result_summary=result_summary,
            subject_rows=subject_rows,
            result_status=result_status,
            selected_subjects=selected_subjects,
            generated_on=generated_on,
            error_message=error_message,
            record_info=None
        )
    finally:
        conn.close()

@app.route("/dmc/record/<int:dmc_id>")
def view_dmc_record(dmc_id):
    conn = get_connection()
    try:
        record = conn.execute('SELECT * FROM dmc_records WHERE id = ?', (dmc_id,)).fetchone()
        if not record:
            abort(404)
        student_id = record['student_id']
        if not can_view_student_dmc(student_id):
            abort(403)

        student = conn.execute(
            'SELECT id, name, father_name, admission_no, technology, semester, campus FROM students WHERE id = ?',
            (student_id,)
        ).fetchone()

        subject_rows = json.loads(record['subject_rows_json']) if record['subject_rows_json'] else []
        result_summary = json.loads(record['summary_json']) if record['summary_json'] else None
        exam_ids = json.loads(record['exam_ids']) if record['exam_ids'] else []
        selected_subjects = ', '.join([row.get('subject', '') for row in subject_rows if row.get('subject')])

        exam_titles = []
        if exam_ids:
            placeholders = ','.join(['?'] * len(exam_ids))
            rows = conn.execute(
                f'SELECT exam_id, title, subject FROM midterm_exams WHERE exam_id IN ({placeholders})',
                exam_ids
            ).fetchall()
            exam_titles = [row['subject'] or row['title'] for row in rows]

        created_at = record['created_at']
        try:
            generated_on = datetime.fromisoformat(created_at) if created_at else datetime.now()
        except ValueError:
            generated_on = datetime.now()

        return render_template(
            'dmc.html',
            student=dict(student) if student else None,
            exam=None,
            subject_rows=subject_rows,
            result_summary=result_summary,
            result_status=(result_summary or {}).get('result_status'),
            selected_subjects=selected_subjects,
            generated_on=generated_on,
            error_message=None,
            record_info={'exam_titles': exam_titles}
        )
    finally:
        conn.close()

@app.route("/api/proctoring/logs", methods=['GET'])
@login_required
def get_proctoring_logs():
    """Get proctoring logs"""
    try:
        role = session.get('role')
        if role != 'admin':
            return jsonify({'status': 'error', 'message': 'Admin only'}), 403
        
        exam_id = request.args.get('exam_id')
        student_id = request.args.get('student_id')
        
        conn = get_connection()
        cur = conn.cursor()
        
        query = '''
            SELECT pl.*, s.name as student_name, s.admission_no, e.title as exam_title
            FROM exam_proctor_logs pl
            JOIN students s ON pl.student_id = s.id
            JOIN midterm_exams e ON pl.exam_id = e.exam_id
            WHERE 1=1
        '''
        params = []
        
        if exam_id:
            query += ' AND pl.exam_id = ?'
            params.append(exam_id)
        if student_id:
            query += ' AND pl.student_id = ?'
            params.append(student_id)
        
        query += ' ORDER BY pl.timestamp DESC LIMIT 1000'
        
        cur.execute(query, params)
        logs = [dict(row) for row in cur.fetchall()]
        conn.close()
        
        return jsonify({'status': 'success', 'data': logs})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route("/api/exams/<int:exam_id>/students/<int:student_id>/answers", methods=['GET'])
@login_required
def get_student_exam_answers(exam_id, student_id):
    role = session.get('role')
    teacher_id = session.get('teacher_id')
    if role not in ('admin', 'teacher'):
        return jsonify({'status': 'error', 'message': 'Unauthorized'}), 403

    conn = get_connection()
    try:
        exam = conn.execute('SELECT exam_id, title, subject, created_by FROM midterm_exams WHERE exam_id = ?', (exam_id,)).fetchone()
        if not exam:
            return jsonify({'status': 'error', 'message': 'Exam not found'}), 404
        if role == 'teacher' and exam['created_by'] != teacher_id:
            return jsonify({'status': 'error', 'message': 'Unauthorized'}), 403

        instance = conn.execute(
            '''
                SELECT * FROM midterm_instances
                WHERE exam_id = ? AND student_id = ?
                ORDER BY instance_id DESC LIMIT 1
            ''',
            (exam_id, student_id)
        ).fetchone()
        if not instance:
            return jsonify({'status': 'error', 'message': 'No attempt found for this student.'}), 404

        answers = conn.execute(
            '''
                SELECT
                    iq.question_order,
                    iq.option_order_json,
                    q.question_text,
                    q.correct_option,
                    q.correct_index,
                    q.option_a,
                    q.option_b,
                    q.option_c,
                    q.option_d,
                    q.marks AS question_marks,
                    mr.selected_option,
                    mr.selected_index,
                    mr.is_correct,
                    mr.marks_obtained,
                    mr.timestamp AS response_timestamp
                FROM midterm_instance_questions iq
                JOIN midterm_questions q ON iq.question_id = q.question_id
                LEFT JOIN midterm_responses mr
                    ON mr.instance_id = iq.instance_id AND mr.question_id = iq.question_id
                WHERE iq.instance_id = ?
                ORDER BY iq.question_order
            ''',
            (instance['instance_id'],)
        ).fetchall()

        answers_payload = []
        option_letter_map = {'A': 0, 'B': 1, 'C': 2, 'D': 3}
        for row in answers:
            selected_option = row['selected_option']
            selected_index = row['selected_index']
            try:
                option_order = json.loads(row['option_order_json']) if row['option_order_json'] else [0, 1, 2, 3]
            except (TypeError, json.JSONDecodeError):
                option_order = [0, 1, 2, 3]

            options = [row['option_a'], row['option_b'], row['option_c'], row['option_d']]
            if not selected_option and selected_index is not None:
                try:
                    base_index = option_order[selected_index]
                    if base_index is not None and 0 <= base_index < len(options):
                        selected_option = options[base_index]
                except (IndexError, TypeError, ValueError):
                    selected_option = None

            if selected_option:
                answer_display = selected_option
            elif selected_index is not None:
                answer_display = f"Option {chr(65 + selected_index)}"
            else:
                answer_display = 'Not Answered'

            correct_text = None
            correct_index = row['correct_index']
            if correct_index is not None and 0 <= correct_index < len(options):
                correct_text = options[correct_index]
            elif row['correct_option']:
                mapped_index = option_letter_map.get(str(row['correct_option']).strip().upper())
                if mapped_index is not None and 0 <= mapped_index < len(options):
                    correct_text = options[mapped_index]
            if not correct_text:
                correct_text = row['correct_option'] or 'Not Available'

            answers_payload.append({
                'question_order': row['question_order'],
                'question_text': row['question_text'],
                'selected_option': selected_option,
                'selected_index': selected_index,
                'answer_display': answer_display,
                'correct_answer': correct_text,
                'question_marks': row['question_marks'],
                'marks_obtained': row['marks_obtained'],
                'is_correct': row['is_correct'],
                'answered_at': row['response_timestamp']
            })

        student = conn.execute(
            'SELECT id, name, admission_no FROM students WHERE id = ?',
            (student_id,)
        ).fetchone()

        return jsonify({
            'status': 'success',
            'answers': answers_payload,
            'student': dict(student) if student else {},
            'exam': {'title': exam['title'], 'subject': exam['subject']},
            'instance_id': instance['instance_id'],
            'submitted_at': instance['end_time']
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        conn.close()

# Student available exams
@app.route("/api/student/exams", methods=['GET'])
def get_student_exams():
    """Get available exams for student"""
    try:
        if 'student_logged_in' not in session or not session['student_logged_in']:
            return jsonify({'status': 'error', 'message': 'Student login required'}), 401
        
        student_id = session.get('student_id')
        conn = get_connection()
        cur = conn.cursor()
        
        # Get student details
        student = cur.execute('SELECT * FROM students WHERE id = ?', (student_id,)).fetchone()
        if not student:
            return jsonify({'status': 'error', 'message': 'Student not found'}), 404
        
        student_dict = dict(student)
        student_campus = (student_dict.get('campus') or '').strip()
        student_board = (student_dict.get('board') or '').strip()
        student_technology = (student_dict.get('technology') or '').strip()
        student_semester = (student_dict.get('semester') or '').strip()

        try:
            tz = ZoneInfo('Asia/Karachi')
            tz_name = 'Asia/Karachi'
        except Exception:
            tz = ZoneInfo('UTC')
            tz_name = 'UTC'
        now = datetime.now(tz)

        def normalize_text(value):
            if not value:
                return ''
            return re.sub(r'[^a-z0-9]+', '', value.lower().strip())

        def field_matches(exam_value, student_value):
            exam_norm = normalize_text(exam_value)
            if not exam_norm or exam_norm in ('all', 'any'):
                return True
            if not student_value:
                return False
            student_norm = normalize_text(student_value)
            if student_norm == exam_norm:
                return True
            return SequenceMatcher(None, exam_norm, student_norm).ratio() >= 0.9

        def parse_exam_window(exam_row):
            exam_date = exam_row.get('exam_date')
            start_time = exam_row.get('start_time')
            end_time = exam_row.get('end_time')
            duration = exam_row.get('duration')
            start_dt = None
            end_dt = None
            exam_date_obj = None

            if exam_date:
                try:
                    exam_date_obj = datetime.strptime(exam_date, '%Y-%m-%d').date()
                except ValueError:
                    exam_date_obj = None

            if exam_date_obj:
                if start_time:
                    try:
                        start_time_obj = datetime.strptime(start_time, '%H:%M').time()
                    except ValueError:
                        start_time_obj = None
                else:
                    start_time_obj = datetime.strptime('00:00', '%H:%M').time()

                if start_time_obj:
                    start_dt = datetime.combine(exam_date_obj, start_time_obj).replace(tzinfo=tz)

                if end_time:
                    try:
                        end_time_obj = datetime.strptime(end_time, '%H:%M').time()
                        end_dt = datetime.combine(exam_date_obj, end_time_obj).replace(tzinfo=tz)
                    except ValueError:
                        end_dt = None

                if not end_dt and start_dt and duration:
                    try:
                        duration_minutes = int(duration)
                    except (ValueError, TypeError):
                        duration_minutes = 0
                    if duration_minutes > 0:
                        end_dt = start_dt + timedelta(minutes=duration_minutes)

            return start_dt, end_dt

        def build_time_window_label(start_dt, end_dt):
            if start_dt and end_dt:
                return f"{start_dt.strftime('%I:%M %p')} - {end_dt.strftime('%I:%M %p')} ({tz_name})"
            if start_dt:
                return f"Starts at {start_dt.strftime('%I:%M %p')} ({tz_name})"
            return 'Flexible timing'

        def build_window_message(start_dt, end_dt):
            if start_dt and now < start_dt:
                return 'scheduled', f"Opens on {start_dt.strftime('%d %b %Y %I:%M %p')} ({tz_name})"
            if start_dt and end_dt and now > end_dt:
                return 'closed', f"Closed on {end_dt.strftime('%d %b %Y %I:%M %p')} ({tz_name})"
            return 'open', 'Available now'

        cur.execute('''
            SELECT e.*,
                   t.name as teacher_name,
                   CASE WHEN mi.instance_id IS NOT NULL THEN mi.status ELSE 'Not Started' END as attempt_status,
                   mi.instance_id
            FROM midterm_exams e
            LEFT JOIN teachers t ON e.created_by = t.id
            LEFT JOIN midterm_instances mi ON e.exam_id = mi.exam_id AND mi.student_id = ?
            WHERE UPPER(TRIM(e.status)) = 'PUBLISHED'
            ORDER BY e.exam_date DESC, e.start_time DESC
        ''', (student_id,))
        
        raw_exams = [dict(row) for row in cur.fetchall()]
        conn.close()

        filtered_exams = []
        for exam in raw_exams:
            if not all([
                field_matches(exam.get('campus'), student_campus),
                field_matches(exam.get('board'), student_board),
                field_matches(exam.get('technology'), student_technology),
                field_matches(exam.get('semester'), student_semester)
            ]):
                continue

            start_dt, end_dt = parse_exam_window(exam)
            window_status, window_message = build_window_message(start_dt, end_dt)
            exam['start_datetime'] = start_dt.isoformat() if start_dt else None
            exam['end_datetime'] = end_dt.isoformat() if end_dt else None
            exam['time_window_label'] = build_time_window_label(start_dt, end_dt)
            exam['window_status'] = window_status
            exam['window_message'] = window_message
            exam['timezone'] = tz_name
            exam['display_date'] = start_dt.strftime('%d %b %Y') if start_dt else (exam.get('exam_date') or 'TBD')
            exam['can_start'] = window_status == 'open' and exam.get('attempt_status') == 'Not Started'
            exam['can_resume'] = window_status == 'open' and exam.get('attempt_status') == 'In Progress'
            exam['is_window_closed'] = window_status == 'closed'
            exam['is_scheduled'] = window_status == 'scheduled'
            filtered_exams.append(exam)

        if not filtered_exams:
            return jsonify({
                'status': 'success',
                'data': [],
                'message': 'No exams match your campus/semester filters right now. Please reach out to your teacher if you were expecting an exam.'
            })

        return jsonify({'status': 'success', 'data': filtered_exams})
    except Exception as e:
        print(f"Error getting student exams: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

# ==================== STUDENT ACCOUNT MANAGEMENT ROUTES ====================

@app.route("/admin/student-accounts")
@admin_required
def admin_student_accounts():
    """Admin page for managing student accounts"""
    return render_template('admin_student_accounts.html')

@app.route("/api/admin/student-accounts", methods=['GET'])
@admin_required
def api_get_student_accounts():
    """Get all student accounts with filters"""
    try:
        campus = request.args.get('campus', '')
        technology = request.args.get('technology', '')
        semester = request.args.get('semester', '')
        board = request.args.get('board', '')
        status = request.args.get('status', '')
        search = request.args.get('search', '')
        
        conn = get_connection()
        cur = conn.cursor()
        
        query = '''
            SELECT s.id, s.admission_no, s.name, s.father_name, s.campus, s.board, 
                   s.technology, s.semester, s.status, s.username, s.account_status, 
                   s.last_login, s.created_at
            FROM students s
            WHERE 1=1
        '''
        params = []
        
        # By default, show only active students (unless status filter is explicitly set)
        if not status:
            query += ' AND (s.status = ? OR s.status IS NULL)'
            params.append('Active')
        
        if campus:
            query += ' AND s.campus = ?'
            params.append(campus)
        if technology:
            query += ' AND s.technology = ?'
            params.append(technology)
        if semester:
            query += ' AND s.semester = ?'
            params.append(semester)
        if board:
            query += ' AND s.board = ?'
            params.append(board)
        if status:
            if status == 'Active':
                query += ' AND (s.status = ? OR s.status IS NULL) AND (s.account_status = ? OR s.account_status IS NULL)'
                params.extend(['Active', 'Active'])
            elif status == 'Inactive':
                query += ' AND (s.status != ? OR s.account_status = ?)'
                params.extend(['Active', 'Inactive'])
        if search:
            query += ' AND (s.admission_no LIKE ? OR s.name LIKE ? OR s.father_name LIKE ?)'
            search_term = f'%{search}%'
            params.extend([search_term, search_term, search_term])
        
        query += ' ORDER BY s.admission_no'
        cur.execute(query, params)
        students = [dict(row) for row in cur.fetchall()]
        
        # Get filter options
        cur.execute('SELECT DISTINCT campus FROM students WHERE campus IS NOT NULL ORDER BY campus')
        campuses = [row['campus'] for row in cur.fetchall()]
        cur.execute('SELECT DISTINCT technology FROM students WHERE technology IS NOT NULL ORDER BY technology')
        technologies = [row['technology'] for row in cur.fetchall()]
        cur.execute('SELECT DISTINCT semester FROM students WHERE semester IS NOT NULL ORDER BY semester')
        semesters = [row['semester'] for row in cur.fetchall()]
        cur.execute('SELECT DISTINCT board FROM students WHERE board IS NOT NULL ORDER BY board')
        boards = [row['board'] for row in cur.fetchall()]
        
        conn.close()
        
        return jsonify({
            'status': 'success',
            'data': students,
            'filters': {
                'campuses': campuses,
                'technologies': technologies,
                'semesters': semesters,
                'boards': boards
            }
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route("/api/admin/student-accounts/generate", methods=['POST'])
@admin_required
def api_generate_student_accounts():
    """Auto-generate accounts for active students without accounts"""
    try:
        conn = get_connection()
        cur = conn.cursor()
        
        # Get active students without accounts
        cur.execute('''
            SELECT id, admission_no, technology, status 
            FROM students 
            WHERE (status = ? OR status IS NULL) 
            AND (username IS NULL OR username = '')
            AND (account_status IS NULL OR account_status = 'Active')
        ''', ('Active',))
        
        students = cur.fetchall()
        generated = 0
        errors = []
        
        for student in students:
            try:
                student_dict = dict(student)
                username = student_dict['admission_no']
                # Default password is technology name
                default_password = student_dict.get('technology', 'student123')
                password_hash = generate_password_hash(default_password)
                
                cur.execute('''
                    UPDATE students 
                    SET username = ?, password_hash = ?, account_status = 'Active'
                    WHERE id = ?
                ''', (username, password_hash, student_dict['id']))
                generated += 1
            except Exception as e:
                errors.append(f"Error for {student_dict.get('admission_no', 'Unknown')}: {str(e)}")
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'status': 'success',
            'message': f'Generated {generated} student accounts',
            'generated': generated,
            'errors': errors
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route("/api/admin/student-accounts/<int:student_id>/reset-password", methods=['POST'])
@admin_required
def api_reset_student_password(student_id):
    """Reset student password"""
    try:
        data = request.get_json()
        new_password = data.get('password')
        
        if not new_password:
            return jsonify({'status': 'error', 'message': 'Password required'}), 400
        
        conn = get_connection()
        student = conn.execute('SELECT admission_no, technology FROM students WHERE id = ?', (student_id,)).fetchone()
        
        if not student:
            conn.close()
            return jsonify({'status': 'error', 'message': 'Student not found'}), 404
        
        password_hash = generate_password_hash(new_password)
        conn.execute('UPDATE students SET password_hash = ? WHERE id = ?', (password_hash, student_id))
        conn.commit()
        conn.close()
        
        return jsonify({'status': 'success', 'message': 'Password reset successfully'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route("/api/admin/student-accounts/<int:student_id>/deactivate", methods=['POST'])
@admin_required
def api_deactivate_student_account(student_id):
    """Deactivate student account"""
    try:
        conn = get_connection()
        conn.execute('UPDATE students SET account_status = ? WHERE id = ?', ('Inactive', student_id))
        conn.commit()
        conn.close()
        return jsonify({'status': 'success', 'message': 'Account deactivated successfully'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route("/api/admin/student-accounts/<int:student_id>/activate", methods=['POST'])
@admin_required
def api_activate_student_account(student_id):
    """Activate student account"""
    try:
        conn = get_connection()
        conn.execute('UPDATE students SET account_status = ? WHERE id = ?', ('Active', student_id))
        conn.commit()
        conn.close()
        return jsonify({'status': 'success', 'message': 'Account activated successfully'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route("/api/admin/student-accounts/<int:student_id>", methods=['GET'])
@admin_required
def api_get_student_account(student_id):
    """Get single student account details"""
    try:
        conn = get_connection()
        student = conn.execute('''
            SELECT id, admission_no, name, username, account_status, technology, semester
            FROM students WHERE id = ?
        ''', (student_id,)).fetchone()
        conn.close()
        
        if not student:
            return jsonify({'status': 'error', 'message': 'Student not found'}), 404
        
        return jsonify({'status': 'success', 'data': dict(student)})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route("/api/admin/student-accounts/<int:student_id>/edit", methods=['POST'])
@admin_required
def api_edit_student_account(student_id):
    """Edit student account (password and status)"""
    try:
        data = request.get_json()
        new_password = data.get('password')
        account_status = data.get('account_status', 'Active')
        
        conn = get_connection()
        student = conn.execute('SELECT admission_no, technology FROM students WHERE id = ?', (student_id,)).fetchone()
        
        if not student:
            conn.close()
            return jsonify({'status': 'error', 'message': 'Student not found'}), 404
        
        # Update password if provided
        if new_password:
            password_hash = generate_password_hash(new_password)
            conn.execute('UPDATE students SET password_hash = ? WHERE id = ?', (password_hash, student_id))
            
            # Log activity
            from datetime import datetime
            conn.execute(
                'INSERT INTO student_activity_log (student_id, activity_type, activity_description, ip_address, created_at) VALUES (?, ?, ?, ?, ?)',
                (student_id, 'password_changed', 'Password changed by admin', request.remote_addr, datetime.now().isoformat())
            )
        
        # Update account status
        conn.execute('UPDATE students SET account_status = ? WHERE id = ?', (account_status, student_id))
        
        # Log status change
        from datetime import datetime
        conn.execute(
            'INSERT INTO student_activity_log (student_id, activity_type, activity_description, ip_address, created_at) VALUES (?, ?, ?, ?, ?)',
            (student_id, 'account_updated', f'Account status changed to {account_status}', request.remote_addr, datetime.now().isoformat())
        )
        
        conn.commit()
        conn.close()
        
        return jsonify({'status': 'success', 'message': 'Account updated successfully'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route("/api/admin/student-accounts/<int:student_id>/create", methods=['POST'])
@admin_required
def api_create_student_account(student_id):
    """Create student account"""
    try:
        conn = get_connection()
        student = conn.execute('SELECT admission_no, technology FROM students WHERE id = ?', (student_id,)).fetchone()
        
        if not student:
            conn.close()
            return jsonify({'status': 'error', 'message': 'Student not found'}), 404
        
        student_dict = dict(student)
        
        # Check if account already exists
        existing = conn.execute('SELECT username FROM students WHERE id = ? AND username IS NOT NULL', (student_id,)).fetchone()
        if existing:
            conn.close()
            return jsonify({'status': 'error', 'message': 'Account already exists'}), 400
        
        # Generate username (admission number) and password (technology name)
        username = student_dict['admission_no']
        default_password = student_dict.get('technology', 'student123')
        password_hash = generate_password_hash(default_password)
        
        conn.execute('''
            UPDATE students 
            SET username = ?, password_hash = ?, account_status = 'Active'
            WHERE id = ?
        ''', (username, password_hash, student_id))
        
        # Log activity
        from datetime import datetime
        conn.execute(
            'INSERT INTO student_activity_log (student_id, activity_type, activity_description, ip_address, created_at) VALUES (?, ?, ?, ?, ?)',
            (student_id, 'account_created', f'Account created with username: {username}', request.remote_addr, datetime.now().isoformat())
        )
        
        conn.commit()
        conn.close()
        
        return jsonify({'status': 'success', 'message': 'Account created successfully'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route("/api/admin/student-accounts/<int:student_id>/activity", methods=['GET'])
@admin_required
def api_get_student_activity(student_id):
    """Get student activity log"""
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute('''
            SELECT * FROM student_activity_log 
            WHERE student_id = ? 
            ORDER BY created_at DESC 
            LIMIT 100
        ''', (student_id,))
        activities = [dict(row) for row in cur.fetchall()]
        conn.close()
        return jsonify({'status': 'success', 'data': activities})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

# ==================== END STUDENT ACCOUNT MANAGEMENT ROUTES ====================

# ==================== STUDENT DASHBOARD API ROUTES ====================

@app.route("/api/student/info", methods=['GET'])
def api_get_student_info():
    """Get logged-in student's information"""
    try:
        if 'student_logged_in' not in session or not session.get('student_logged_in'):
            return jsonify({'status': 'error', 'message': 'Student login required'}), 401
        
        student_id = session.get('student_id')
        conn = get_connection()
        student = conn.execute('SELECT id, admission_no, name, technology, semester, campus, board FROM students WHERE id = ?', (student_id,)).fetchone()
        conn.close()
        
        if not student:
            return jsonify({'status': 'error', 'message': 'Student not found'}), 404
        
        return jsonify({
            'status': 'success',
            'data': dict(student)
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route("/api/student/change-password", methods=['POST'])
def api_student_change_password():
    """Allow student to change their password"""
    try:
        if 'student_logged_in' not in session or not session.get('student_logged_in'):
            return jsonify({'status': 'error', 'message': 'Student login required'}), 401
        
        data = request.get_json()
        current_password = data.get('current_password')
        new_password = data.get('new_password')
        
        if not current_password or not new_password:
            return jsonify({'status': 'error', 'message': 'Current and new password required'}), 400
        
        student_id = session.get('student_id')
        conn = get_connection()
        student = conn.execute('SELECT password_hash FROM students WHERE id = ?', (student_id,)).fetchone()
        
        if not student:
            conn.close()
            return jsonify({'status': 'error', 'message': 'Student not found'}), 404
        
        # Verify current password
        if not check_password_hash(student['password_hash'], current_password):
            conn.close()
            return jsonify({'status': 'error', 'message': 'Current password is incorrect'}), 401
        
        # Update password
        new_password_hash = generate_password_hash(new_password)
        conn.execute('UPDATE students SET password_hash = ? WHERE id = ?', (new_password_hash, student_id))
        
        # Log activity
        from datetime import datetime
        conn.execute(
            'INSERT INTO student_activity_log (student_id, activity_type, activity_description, ip_address, created_at) VALUES (?, ?, ?, ?, ?)',
            (student_id, 'password_change', 'Student changed password', request.remote_addr, datetime.now().isoformat())
        )
        conn.commit()
        conn.close()
        
        return jsonify({'status': 'success', 'message': 'Password changed successfully'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

# ==================== END STUDENT DASHBOARD API ROUTES ====================

# ==================== END MIDTERM & TESTS MODULE API ROUTES ====================

def main():
    app.run(port=int(os.environ.get('PORT', 8080)), debug=True)

if __name__ == "__main__":
    main()
