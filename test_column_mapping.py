#!/usr/bin/env python
from openpyxl import load_workbook

# Test the column mapping logic
column_mapping = {
    'Admission No': 'admission_no',
    'Name': 'name',
    'Father\'s Name': 'father_name',
    'Address': 'address',
    'Phone #': 'phone',
    'SMS Number': 'sms_phone',
    'Gender': 'gender',
    'Campus': 'campus',
    'Board': 'board',
    'Semester/Year': 'semester',
    'Student Status': 'status',
    'Student Type': 'student_type',
    'Technology/Program': 'technology',
    'Remarks & Notes': 'remarks',
}

# Load your Excel file
wb = load_workbook('1ST SEMESTER.xlsx')
ws = wb.active
headers = [cell.value for cell in ws[1]]

print('Testing column mapping:')
print('=' * 60)

# Test row 2
row_data = {}
for col_idx, header in enumerate(headers):
    if header in column_mapping:
        db_col_name = column_mapping[header]
        cell_value = ws.cell(row=2, column=col_idx + 1).value
        row_data[db_col_name] = cell_value
        print(f'✓ {header:20} -> {db_col_name:15} = {cell_value}')

print()
print('Row 2 Data Summary:')
print(f'  Admission No: {row_data.get("admission_no")}')
print(f'  Name: {row_data.get("name")}')
print(f'  Phone: {row_data.get("phone")}')
print(f'  SMS Phone: {row_data.get("sms_phone")}')
print(f'  Semester: {row_data.get("semester")}')
print(f'  Technology: {row_data.get("technology")}')
print(f'  Remarks: {row_data.get("remarks")}')

